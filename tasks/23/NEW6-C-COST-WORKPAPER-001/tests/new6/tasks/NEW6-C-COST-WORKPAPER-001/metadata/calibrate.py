"""One bounded offline fixture batch; selective repair reruns by explicit case name."""
from pathlib import Path
from decimal import Decimal
import json,sys,openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
ROOT=Path(__file__).resolve().parents[1]
sys.path[:0]=[str(ROOT/'metadata'),str(ROOT/'tests')]
from ooxml_edit import edit
from evaluate import evaluate,original_model
ALL=['R001','R002','R003','R004','R005','R006']
F=ROOT/'fixtures';REF=ROOT/'solution/reference.xlsx'
def make(name,**kwargs):
 p=F/(name+'.xlsx')
 if not p.exists():edit(REF,p,clear_caches=True,**kwargs)
 return p
def build():
 # The fixture builder may use reference coordinates; the Judge never does.
 w=openpyxl.load_workbook(REF,data_only=False)
 names={s.title:'View '+str(i) for i,s in enumerate(w,1)};patch={};new={}
 for sh in w:
  patch[sh.title]={};d={}
  for row in sh:
   for c in row:
    if c.value is None:continue
    patch[sh.title][c.coordinate]=None;co=get_column_letter(c.column+1)+str(c.row+4);v=c.value
    if c.data_type=='f':
     v=Translator(v,origin=c.coordinate).translate_formula(co)
     for a,b in names.items():v=v.replace("'"+a+"'!","'"+b+"'!")
    d[co]=v
  new[names[sh.title]]=d
 make('equivalent_layout',patches=patch,new_sheets=new)
 make('equivalent_formula',patches={'Summary':{'E7':'=SUM(E5,E6)*D7','E11':'=SUM(E10*D11)','E16':'=E14*(1+D15)'},'Elements':{'D19':"=SUM(B19,-'Provisional sums'!B15,'Review decisions'!C5)"}})
 make('equivalent_rounded_final',patches={'Summary':{'E16':'=ROUND(SUM(E14:E15),0)','F16':'=E16-B16'}})
 original=original_model()['working'];rr={'building':5,'preliminaries':6,'overheads':7,'building_estimate':8,'base':10,'design_risk':11,'inflation_excluded':14,'inflation':15,'vat_excluded':16}
 p={'F4':'Difference from recomputed original','J4':'Recomputed original amount'}
 for key,r in rr.items():p['J'+str(r)]=float(original[key]);p['F'+str(r)]=f'=E{r}-J{r}'
 make('equivalent_recomputed_basis',patches={'Summary':p})
 make('equivalent_omitted_earlier_quote',patches={'Review decisions':{col+'6':None for col in 'ABCDE'}})
 make('duplicate_heating',patches={'Elements':{'D19':"=B19+'Review decisions'!C5"}})
 make('outdated_quote',patches={'Review decisions':{'C5':472000}})
 make('unapproved_option_included',patches={'Elements':{'D19':"=B19-'Provisional sums'!B15+'Review decisions'!C5+'Review decisions'!C7"}})
 make('overwritten_original',patches={'Summary':{'B16':2039472}})
 make('stale_reconciliation',patches={'Summary':{'F16':68193.588}})
 make('wrong_allowance_base',patches={'Summary':{'E11':'=E5*D11'}})
 make('mixed_final',clones={'Second final':('Summary',[])},patches={})
 mixed=F/'mixed_final.xlsx';marker=F/'mixed_final_mutated.json'
 if not marker.exists():
  edit(mixed,mixed,patches={'Second final':{'E16':'=SUM(E14:E15)+1000'}},clear_caches=True);marker.write_text('{"mutation":"one contradictory current final limit"}')
 make('duplicate_omission',patches={'Elements':{'A7':'Pitched Roof'}})
 # The exported reference caches were independently checked against native
 # Decimal parity; fixture construction has no machine-specific native path.
 cached=openpyxl.load_workbook(REF,data_only=True)
 static={sh.title:{c.coordinate:cached[sh.title][c.coordinate].value for row in sh for c in row if c.data_type=='f'} for sh in w}
 make('static_current',patches=static)
 make('legal_formula_limit',patches={'Summary':{'E16':'=_xlfn.PY("1")'}})
 unbound={'A1':'Unbound material review table','A4':'Mystery item','B4':'Price data','A5':'Heating review','B5':458000,'A6':'Earlier review','B6':472000,'A7':'Asbestos alternative','B7':18000}
 make('legal_layout_limit',patches={'Review decisions':{c.coordinate:None for row in w['Review decisions'] for c in row if c.value is not None}},new_sheets={'Alternative layout':unbound})
 (F/'malformed.xlsx').write_text('This is not an XLSX archive.')
 cases=[{'name':'reference','path':'solution/reference.xlsx','status':'SCORED','lose':[],'preserve':ALL}]
 def case(name,lose=(),preserve=None,status='SCORED'):
  cases.append({'name':name,'path':'fixtures/'+name+'.xlsx','status':status,'lose':list(lose),'preserve':list(preserve if preserve is not None else [k for k in ALL if k not in lose])})
 for n in ['equivalent_layout','equivalent_formula','equivalent_rounded_final','equivalent_recomputed_basis','equivalent_omitted_earlier_quote']:case(n)
 for n in ['duplicate_heating','wrong_allowance_base']:case(n,['R003','R004','R005'])
 case('outdated_quote',['R002','R003','R004','R005'])
 case('unapproved_option_included',['R002','R003','R004','R005'])
 case('overwritten_original',['R001','R004','R005'])
 case('stale_reconciliation',['R005'])
 case('mixed_final',['R003','R004','R005'])
 case('duplicate_omission',['R001','R003','R006'],['R002','R004','R005'])
 case('static_current',['R005'])
 for n in ['legal_formula_limit','legal_layout_limit']:case(n,[],[],status='JUDGE_ERROR')
 case('malformed',[],[],status='MALFORMED_OUTPUT');case('missing',[],[],status='OUTPUT_MISSING')
 (F/'manifest.json').write_text(json.dumps({'cases':cases,'source':'OOXML transformations from independently verified reference','fixture_names_do_not_determine_pass':True},indent=2))
 return cases
def main():
 cases=build();selected=set(sys.argv[1:])
 if selected:
  if selected-set(c['name'] for c in cases):raise ValueError('Unknown fixture name')
  cases=[c for c in cases if c['name'] in selected]
 out=ROOT/'validation/calibration';out.mkdir(exist_ok=True);rows=[];den=json.loads((ROOT/'tests/fact_contract.json').read_text())['denominators']
 for c in cases:
  r=evaluate(ROOT/c['path'],out/c['name'],True)
  scores=r.get('criterion_scores',{});checks=[r['evaluation_status']==c['status']]
  if c['status']=='SCORED':
   checks += [Decimal(scores[k])<1 for k in c['lose']]+[Decimal(scores[k])==1 for k in c['preserve']]
   checks += [r['evidence']['denominators']==den]
   if c['name']=='static_current':checks += [Decimal(scores['R005'])==0]
  row={'name':c['name'],'status':r['evaluation_status'],'score':r.get('score_decimal'),'facts':scores,'expected':c,'passed':all(checks),'error':r.get('evidence',{}).get('error')};rows.append(row)
  receipt={'planned':len(cases),'completed':len(rows),'passed':all(r['passed'] for r in rows),'results':rows,'api_calls':0,'fixed_denominators':den}
  name='receipt.json' if not selected else 'receipt_repair_'+'_'.join(sorted(selected))+'.json'
  (out/name).write_text(json.dumps(receipt,indent=2));print(json.dumps(row),flush=True)
  if not row['passed']:raise AssertionError(c['name'])
if __name__=='__main__':main()
