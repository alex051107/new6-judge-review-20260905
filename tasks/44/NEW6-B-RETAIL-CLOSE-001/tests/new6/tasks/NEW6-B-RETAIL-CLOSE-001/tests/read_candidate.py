"""Conservative label/identity reader, with actual OOXML chart evidence.

This reader discovers headers and locations from candidate contents. It does not
use reference cell coordinates or interpret candidate formulas. Uncached formulas
and chart reference forms outside the supported reader have explicit pending states.
"""
from pathlib import Path
from decimal import Decimal, InvalidOperation
from collections import Counter
import re, zipfile, xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.cell import range_boundaries,get_column_letter
from semantic_headers import headers as semantic_headers,october_row,apply_scope,month

class ParsePending(Exception):pass
def norm(v):return re.sub(r'[^a-z0-9]+','',str(v or '').lower())
ALIASES={
 'code':['geographycode','areacode','localcode','authoritycode','code'],
 'name':['geography','authority','areaname','geographyname','localauthority'],
 'employment_old':['employmentrateearlier','employmentrate2024','employment2024','earlieremploymentrate'],
 'employment_new':['employmentratelater','employmentrate2025','employment2025','lateremploymentrate'],
 'unemployment_old':['unemploymentrateearlier','unemploymentrate2024','unemployment2024','earlierunemploymentrate'],
 'unemployment_new':['unemploymentratelater','unemploymentrate2025','unemployment2025','laterunemploymentrate'],
 'employment_change':['employmentchangepp','employmentratechangepp','employmentchange','employmentppchange'],
 'unemployment_change':['unemploymentchangepp','unemploymentratechangepp','unemploymentchange','unemploymentppchange'],
 'comparison_status':['comparisonstatus','comparability','pairstatus'],
 'metric':['metric','indicator','measure','field','item'], 'value':['value','reportedvalue','result','finding'],
 'reason':['reason','exclusionreason','treatmentreason','exceptionreason'],
 'rank':['rank','position'], 'edition':['edition','release'],
 'raw':['rawvalue','storedvalue','underlyingvalue'], 'published':['publishedvalue','displayedrate','publishedrate'],
 'source':['sourcelocation','source','provenance','sourcereference'],
 'rowid':['sourcerowid','rowidentity','sourceid'], 'invoice':['invoiceno','invoice','invoiceid','invoicenumber'],
 'stock':['stockcode','productcode'], 'description':['description','productdescription'],
 'quantity':['quantity','qty'], 'date':['invoicedate','timestamp','transactiondate'],
 'price':['unitprice','unitpricegbp'], 'customer':['customerid','customer'], 'country':['country'],
 'class':['classification','status','category','entrytype'], 'amount':['recordedamountgbp','amountgbp','signedamount','recordedamount','amount'],
 'missing_customer':['missingcustomer','missingcustomerid'], 'count':['rowcount','recordcount','occurrences','count','linecount'],
 'issue':['issuetype','issue'], 'countries':['invoicecountries','countries'],
 'original_invoice':['originalinvoice','linkedinvoice','originalinvoiceno','credittooriginal']}
ALIASES.update({'sale_amount':['salesgbp','salesamount','salesvaluegbp'],'credit_amount':['creditsgbp','creditamount','creditsvaluegbp'],'exception_amount':['exceptionsgbp','exceptionamount'],'sale_count':['saleslines','salesrowcount'],'credit_count':['creditlines','creditrowcount'],'exception_count':['exceptionlines','exceptionrowcount']})
ALIASES.update({ 'component':['component','bridgecomponent','contribution','driver'],
 'bridge_amount':['bridgeamountgbp','bridgeamount','contributiongbp','contributionamount'],
 'cohort':['cohort','skucohort','membership'],
 'q_sep':['sepsalesquantity','septembersalesquantity','sepquantity','septemberquantity','qsep','q0'],
 'q_oct':['octsalesquantity','octobersalesquantity','octquantity','octoberquantity','qoct','q1'],
 'v_sep':['sepsalesvaluegbp','septembersalesvalue','sepsalesvalue','sepvalue','vsep','v0'],
 'v_oct':['octsalesvaluegbp','octobersalesvalue','octsalesvalue','octvalue','voct','v1'],
 'p_sep':['sepweightedunitprice','septemberweightedunitprice','sepweightedprice','psep','p0'],
 'p_oct':['octweightedunitprice','octoberweightedunitprice','octweightedprice','poct','p1'],
 'qty_effect':['quantityeffectgbp','quantityeffect','volumeeffectgbp','volumeeffect','quantitycontribution'],
 'price_effect':['priceandmixeffectgbp','priceandmixeffect','pricemixeffect','priceeffect','pricecontribution'],
 'entry_effect':['newskueffectgbp','newskueffect','newskusales','entryeffect'],
 'exit_effect':['exitedskueffectgbp','exitedskueffect','exitskueffect','exiteffect']})
ALIASES.update({'oct_report':['october'],'sep_report':['september']})
for key,more in {'sale_amount':['sales'],'credit_amount':['credits'],'exception_amount':['exceptions'],'sale_count':['salelines'],'q_sep':['sepqty'],'q_oct':['octqty'],'p_sep':['sepwtdprice'],'p_oct':['octwtdprice'],'qty_effect':['qtycontrib'],'price_effect':['pricemixcontrib'],'entry_effect':['newskucontrib'],'exit_effect':['exitskucontrib']}.items():ALIASES[key].extend(more)
LOOKUP={a:k for k,aliases in ALIASES.items() for a in aliases}

def tables(path,specs):
    w=openpyxl.load_workbook(path,read_only=True,data_only=True)
    result={k:[] for k in specs};text=[];key_bridges=[];key_changes=[]
    from evaluate_bridge import component,COMPONENTS
    for s in w:
        active=None;mapping=None;rows=[];start=None;side_mapping=None;side_rows=[];section_context='';pending_kpis=[]
        def flush():
            if active is not None:result[active].append({'sheet':s.title,'header_row':start,'rows':rows[:]})
            if side_mapping:result['bridge'].append({'sheet':s.title,'header_row':start,'rows':side_rows[:]})
        for rownum,row in enumerate(s.iter_rows(),1):
            values=[c.value for c in row]
            text.extend(str(x) for x in values if isinstance(x,str) and len(str(x))>1)
            # Keep visible row context for free-text/key-value reports without a
            # required "Metric / Value" header. Do not infer values from truth.
            populated=[v for v in values if v is not None]
            if pending_kpis:
                for col,label,bridge_label in pending_kpis:
                    if col<len(values) and num(values[col]) is not None:
                        if label:result['report'].append({'sheet':s.title,'header_row':rownum-1,'rows':[{'metric':label,'value':values[col],'_loc':f'{s.title}!{rownum}','_cells':{'value':f'{s.title}!{get_column_letter(col+1)}{rownum}'}}]})
                        if bridge_label:key_bridges.append({'component':bridge_label,'bridge_amount':values[col],'_loc':f'{s.title}!{rownum}','_cells':{'bridge_amount':f'{s.title}!{get_column_letter(col+1)}{rownum}'}})
                pending_kpis=[]
            if populated and all(isinstance(v,str) for v in populated):
                horizontal={'octobersales':('Gross sales',None),'sales':('Gross sales',None),'octobercredits':('Signed credits',None),'signedcredits':('Signed credits',None),'netrecordedvalue':('Net recorded value',None),'octobernetrecorded':('Net recorded value','October net recorded value'),'octobernet':('Net recorded value','October net recorded value'),'septembernetrecorded':(None,'September net recorded value'),'septembernet':(None,'September net recorded value'),'octoberexceptions':('Exception recorded amount',None)}
                pending_kpis=[(i,*horizontal[norm(v)]) for i,v in enumerate(values) if norm(v) in horizontal]
                if len(pending_kpis)<2:pending_kpis=[]
            if len(populated)<=3 or rownum<=30:text.append(('__BRIDGE__ ' if active=='bridge' or side_mapping else '__SEPTEMBER__ ' if month(section_context)=='sep' else '')+' || '.join(str(x) for x in populated))
            if len(populated)==1 and isinstance(populated[0],str) and len(populated[0])<100 and month(populated[0]):
                section_context=populated[0]
            if rownum<=60 and active not in ['records','queue','sku','bridge'] and not side_mapping:
                label=component(values[0]) if values else None
                if values and norm(values[0])=='change' and len(values)>1 and num(values[1]) is not None:
                    key_changes.append({'component':'Net recorded value change','bridge_amount':values[1],'_loc':f'{s.title}!{rownum}','_cells':{'bridge_amount':f'{s.title}!B{rownum}'}})
                if label in ['September net recorded value','October net recorded value','Net recorded value change','Bridge residual'] and values and len(values)>1 and num(values[1]) is not None:
                    key_bridges.append({'component':label,'bridge_amount':values[1],'_loc':f'{s.title}!{rownum}','_cells':{'bridge_amount':f'{s.title}!B{rownum}'}})
            headers=semantic_headers(values,LOOKUP,norm);new_side=None
            if 'component' in headers and 'metric' in headers and headers['component']>headers['metric']:
                split=headers['component'];new_side=semantic_headers([v if i>=split else None for i,v in enumerate(values)],LOOKUP,norm);headers=semantic_headers([v if i<split else None for i,v in enumerate(values)],LOOKUP,norm)
                if 'bridge_amount' not in new_side:new_side['bridge_amount']=new_side.get('value',new_side.get('amount'))
                if new_side['bridge_amount'] is None:
                    # A contribution heading followed by one otherwise unlabeled column
                    # is a two-column table; the next body cell still has to be numeric.
                    if split+1<len(values) and all(v is None for v in values[split+1:]):new_side['bridge_amount']=split+1
                    else:raise ParsePending('Side-by-side bridge amount column is ambiguous')
            if 'metric' in headers and 'amount' in headers and 'value' not in headers:headers['value']=headers['amount']
            if 'component' in headers and 'value' in headers and 'bridge_amount' not in headers:headers['bridge_amount']=headers['value']
            if 'component' in headers and 'amount' in headers and 'bridge_amount' not in headers:headers['bridge_amount']=headers['amount']
            if 'oct_report' in headers and 'sep_report' in headers:
                lead=min(headers['oct_report'],headers['sep_report'])-1
                if lead>=0:headers['metric']=lead
            matches=[k for k,required in specs.items() if set(required)<=set(headers)]
            if 'totals' in matches and 'invoices' not in matches and 'invoice' in norm(s.title) and not any(norm(v)=='invoicecount' for v in values):
                raise ParsePending('Invoice summary has classification, count and amount but its document identifier header cannot be safely bound.')
            if matches:
                flush()
                if 'queue' in matches and ('issue' in headers or re.search(r'exception|issue',s.title,re.I)):active='queue'
                elif any(k.startswith('invoices') for k in matches):active=max((k for k in matches if k.startswith('invoices')),key=lambda k:len(specs[k]))
                elif any(k.startswith('countries') for k in matches):active=max((k for k in matches if k.startswith('countries')),key=lambda k:len(specs[k]))
                else:active=max(matches,key=lambda k:len(specs[k]))
                if set(headers.get('_duplicate_roles',[])) & set(specs[active]):raise ParsePending('Repeated required header roles require explicit disambiguation: '+str(headers['_duplicate_roles']))
                mapping={k:v for k,v in headers.items() if not k.startswith('_')};rows=[];start=rownum;side_mapping=new_side;side_rows=[];continue
            if active is None and populated:text.append('__UNBOUND__ '+' || '.join(str(x) for x in populated))
            if active is not None and any(v is not None for v in values):
                d={k:(values[i] if i<len(values) else None) for k,i in mapping.items()}
                d['_loc']=f'{s.title}!{rownum}';d['_cells']={k:f'{s.title}!{get_column_letter(i+1)}{rownum}' for k,i in mapping.items()}
                if side_mapping:
                    side_mapping={k:v for k,v in side_mapping.items() if not k.startswith('_')}
                    side={k:(values[i] if i<len(values) else None) for k,i in side_mapping.items()};side['_loc']=d['_loc'];side['_cells']={k:f'{s.title}!{get_column_letter(i+1)}{rownum}' for k,i in side_mapping.items()}
                    if side.get('component') is not None:side_rows.append(side)
                if active in ['records','invoices','countries','invoices_wide','invoices_amounts','countries_wide','countries_amounts','queue','totals','period_totals'] and not october_row(d,s.title+' '+section_context):continue
                if active in ['records','queue'] and norm(d.get('rowid')) in ['total','totals','grandtotal']:continue
                if active=='records':apply_scope(d,norm)
                if active=='records' and d.get('class') is None and d.get('amount') is None:continue
                if active=='sku' and (norm(d.get('stock')).startswith('total') or norm(d.get('stock'))=='grandtotal'):continue
                rows.append(d)
        flush()
    w.close()
    if key_bridges:result['bridge'].append({'sheet':'explicit labelled key/value claims','header_row':None,'rows':key_bridges})
    # Wide tables retain candidate values; class counts may be recoverable from delivered detail.
    for wide,target,key in [('invoices_wide','invoices','invoice'),('countries_wide','countries','country'),('countries_amounts','countries','country'),('invoices_amounts','invoices','invoice')]:
        for region in result.pop(wide,[]):
            converted=[];totals=[]
            for r in region['rows']:
                if r.get(key) is None:continue
                is_total=norm(r.get(key)).startswith('total') or norm(r.get(key)) in ['grandtotal','totals']
                for cls in ['sale','credit','exception']:
                    amount=r.get(cls+'_amount');count=r.get(cls+'_count')
                    if not is_total and num(amount)==0 and num(count)==0:continue
                    d={key:r.get(key),'class':cls,'count':count,'amount':amount,'_loc':r['_loc'],'_cells':r['_cells']}
                    if count is None:d.update(_recover_count=True,_wide_total_count=r.get('count'))
                    (totals if is_total else converted).append(d)
            result[target].append({**region,'rows':converted})
            if wide=='invoices_wide' and totals:result['totals'].append({**region,'rows':totals})
    for region in result.pop('period_report',[]):
        converted=[];class_counts={c:[] for c in ['sale','credit','exception']};class_values={c:[] for c in ['sale','credit','exception']}
        aliases={'sales':'Gross sales','salesgbp':'Gross sales','ordinarysales':'Gross sales','grosssales':'Gross sales','salesvaluegbp':'Gross sales','salesvalue':'Gross sales','salesordinary':'Gross sales','credits':'Signed credits','signedcredits':'Signed credits','creditvalue':'Signed credits','creditsgbp':'Signed credits','creditssigned':'Signed credits','creditsgbpsigned':'Signed credits','netrecordedvalue':'Net recorded value','netrecordedgbp':'Net recorded value','netrecordedvaluegbp':'Net recorded value','net':'Net recorded value','calculableexceptionamountgbp':'Exception recorded amount','exceptioncalculatedamount':'Exception recorded amount','exceptionamountretainedgbp':'Exception recorded amount','exceptionvalueexcluded':'Exception recorded amount','exceptionsexcluded':'Exception recorded amount','exceptionamount':'Exception recorded amount','exceptionrecordedamount':'Exception recorded amount'}
        for r in region['rows']:
            count_class={'ordinarysaleslines':'sale','ordinarysalerows':'sale','salelines':'sale','salerows':'sale','saleslines':'sale','salesrows':'sale','creditlines':'credit','creditrows':'credit','exceptionlines':'exception','exceptionrows':'exception'}.get(norm(r.get('metric')))
            if count_class:class_counts[count_class].append(r)
            amount_class={'Gross sales':'sale','Signed credits':'credit','Exception recorded amount':'exception'}.get(aliases.get(norm(r.get('metric'))))
            if amount_class:class_values[amount_class].append(r)
            if norm(r.get('metric'))=='change' and num(r.get('oct_report')) is not None:
                result['bridge'].append({'sheet':region['sheet'],'header_row':region['header_row'],'rows':[{'component':'Net recorded value change','bridge_amount':r.get('oct_report'),'_loc':r['_loc'],'_cells':r['_cells']}]})
            if norm(r.get('metric')) in ['netrecordedvalue','netrecordedvaluegbp','netrecordedgbp','net']:
                result['bridge'].append({'sheet':region['sheet'],'header_row':region['header_row'],'rows':[{'component':label,'bridge_amount':r.get(role),'_loc':r['_loc'],'_cells':r['_cells']} for label,role in [('September net recorded value','sep_report'),('October net recorded value','oct_report')]]})
            if norm(r.get('metric')) in aliases:
                converted.append({**r,'metric':aliases[norm(r['metric'])],'value':r.get('oct_report'),'_period':'October'})
        result['report'].append({**region,'rows':converted})
        paired=[]
        for cl in class_counts:
            for c in class_counts[cl]:
                for v in class_values[cl]:paired.append({'class':cl,'count':c.get('oct_report'),'amount':v.get('oct_report'),'_loc':c['_loc']+' + '+v['_loc'],'_cells':v['_cells']})
        if paired:result['totals'].append({**region,'rows':paired})
    for region in result.pop('period_totals',[]):
        converted=[]
        for r in region['rows']:
            if 'outofperiod' in norm(r.get('period')):continue
            if not all(num(r.get(c+'_amount')) is not None for c in ['sale','credit','exception']):continue
            for cl,label in [('sale','Gross sales'),('credit','Signed credits'),('exception','Exception recorded amount')]:
                converted.append({'class':cl,'count':r.get(cl+'_count'),'amount':r.get(cl+'_amount'),'_loc':r['_loc'],'_cells':r['_cells']})
                result['report'].append({'sheet':region['sheet'],'header_row':region['header_row'],'rows':[{'metric':label,'value':r.get(cl+'_amount'),'_loc':r['_loc'],'_cells':r['_cells']}]})
        result['totals'].append({**region,'rows':converted})
    endpoint_sheets={}
    for region in result['bridge']:
        for r in region['rows']:
            label=component(r.get('component'))
            if label in ['September net recorded value','October net recorded value']:endpoint_sheets.setdefault(r['_loc'].rsplit('!',1)[0],set()).add(label)
    changes=[r for r in key_changes if len(endpoint_sheets.get(r['_loc'].rsplit('!',1)[0],set()))==2]
    if changes:result['bridge'].append({'sheet':'labelled changes adjacent to both net endpoints','header_row':None,'rows':changes})
    # A reconciliation table may label the three classes and split adjacent periods.
    for region in result.pop('reconciliation',[]):
        converted=[];outside=[]
        for r in region['rows']:
            label=norm(r.get('description'))
            cls={'sales':'sale','credits':'credit','exceptions':'exception'}.get(label)
            if cls:converted.append({**r,'class':cls})
            elif 'outofperiod' in label:outside.append(r)
        if outside and all(num(r.get('count')) is not None and num(r.get('amount')) is not None for r in outside):
            converted.append({'class':'outside_scope','count':sum(num(r['count']) for r in outside),'amount':sum(num(r['amount']) for r in outside),'_loc':region['sheet'],'_cells':{}})
        result['totals'].append({**region,'rows':converted})
    for region in result['queue']:
        for r in region['rows']:
            if 'issue' not in r and str(r.get('reason') or '').strip():r['issue']='missing_customer_attribution' if 'customer' in str(r['reason']).lower() else 'business_exception'
    # A legal formula whose value is absent must not become a business zero.
    with zipfile.ZipFile(path) as z:
        has_formulas=any(re.search(rb'<(?:\w+:)?f(?:\s|>)',z.read(n)) for n in z.namelist() if re.fullmatch(r'xl/worksheets/sheet\d+.xml',n))
    if not has_formulas:return result,text
    f=openpyxl.load_workbook(path,read_only=True,data_only=False)
    v=openpyxl.load_workbook(path,read_only=True,data_only=True)
    for sf,sv in zip(f,v):
        for rf,rv in zip(sf.iter_rows(),sv.iter_rows()):
            for cf,cv in zip(rf,rv):
                if cf.data_type=='f' and (cv.value is None or cv.data_type=='e'):
                    f.close();v.close();raise ParsePending(f'Uncached or unsupported formula at {sf.title}!{cf.coordinate}')
    f.close();v.close()
    return result,text

def num(v):
    if v is None or norm(v) in {'x','c','w','low','na','unavailable','suppressed','notavailable'}:return None
    try:return Decimal(str(v).replace(',','').replace('£','').strip())
    except (InvalidOperation,ValueError):return None
def eq(a,b,tol=Decimal('0.0000001')):
    if b is None:return num(a) is None
    aa=num(a);bb=num(b)
    return aa is not None and bb is not None and abs(aa-bb)<=tol
def population(got,expected):
    a,b=Counter(got),Counter(list(expected));n=sum(b.values())
    return max(0,(n-sum((b-a).values())-sum((a-b).values()))/max(1,n))
def consensus_rows(regions,key):
    """Multiple final representations stay distinct; no last-write-wins."""
    rows=[r for t in regions for r in t['rows'] if r.get(key) is not None]
    by={}
    for r in rows:by.setdefault(str(r[key]),[]).append(r)
    return rows,by
def mean(vals):return sum(vals)/len(vals) if vals else 0

def visible_numbers(report_rows,text,aliases):
    """Bind numeric claims from explicit labels or nearby visible prose only."""
    found=[];normalized=[norm(a) for a in aliases]
    for r in report_rows:
        label=norm(r.get('metric'))
        if any(label==a or label in [a+'gbp',a+'pounds',a+'count'] for a in normalized):found.append(r.get('value'))
    patterns=['[\\s_]*'.join(re.escape(w) for w in a.split()) for a in aliases]
    pat=re.compile(r'(?:'+ '|'.join(patterns)+r')\s*(?:\([^)]*\))?[\s:|=]*(?:(?:were|was|is|are|totalled|totaled|of|amounted to)\s*)?(?:GBP\s*|£\s*)?[\s:|=]*(\(?-?\d[\d,]*(?:\.\d+)?\)?)',re.I)
    for line in text:
        if line.startswith(('__BRIDGE__','__SEPTEMBER__','__UNBOUND__')):continue
        if '||' in line and sum(num(part.strip()) is not None for part in line.split('||'))>1:continue
        for m in pat.finditer(line):
            if re.match(r'\s*(?:%|percent|lines\b|transactions\b)',line[m.end():],re.I):continue
            # A separately labelled comparison-month endpoint is not an October report claim.
            if re.search(r'(?:september|sep|opening)\s*$',line[:m.start()],re.I):continue
            v=m[1];found.append('-'+v[1:-1] if v.startswith('(') else v)
    return found

def charts(path):
    ns={'c':'http://schemas.openxmlformats.org/drawingml/2006/chart','a':'http://schemas.openxmlformats.org/drawingml/2006/main'}
    w=openpyxl.load_workbook(path,data_only=True,read_only=True)
    def reference(f):
        m=re.fullmatch(r"(?:'((?:[^']|'')+)'|([^!]+))!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)",f or '')
        if not m:raise ParsePending('Unsupported legal chart reference: '+str(f))
        name=(m[1] or m[2]).replace("''", "'")
        if name not in w:raise ParsePending('Chart references unknown worksheet')
        a,b,c,d=range_boundaries(m[3]);return [w[name].cell(r,col).value for r in range(b,d+1) for col in range(a,c+1)]
    out=[]
    with zipfile.ZipFile(path) as z:
        for filename in z.namelist():
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',filename):
                tree=ET.fromstring(z.read(filename));series=[];titles=[' '.join(x.text or '' for x in e.findall('.//a:t',ns)) for e in tree.findall('.//c:title',ns)]
                for ser in tree.findall('.//c:ser',ns):
                    item={}
                    tx=ser.find('c:tx',ns)
                    txf=tx.find('.//c:f',ns) if tx is not None else None
                    txv=tx.find('.//c:v',ns) if tx is not None else None
                    item['name']=str(reference(txf.text)[0]) if txf is not None else txv.text if txv is not None else ''
                    for tag,label in [('cat','categories'),('val','values')]:
                        el=ser.find('c:'+tag,ns)
                        if el is None:raise ParsePending('Unsupported chart series shape')
                        f=el.find('.//c:f',ns)
                        cache=[p.find('c:v',ns).text for p in el.findall('.//c:pt',ns) if p.find('c:v',ns) is not None]
                        vals=reference(f.text) if f is not None else cache
                        if not vals:raise ParsePending('Unresolvable chart values')
                        item[label]=vals;item[label+'_cache']=cache;item[label+'_reference']=f.text if f is not None else None
                    series.append(item)
                out.append({'chart':filename,'titles':titles,'series':series})
    w.close();return out
