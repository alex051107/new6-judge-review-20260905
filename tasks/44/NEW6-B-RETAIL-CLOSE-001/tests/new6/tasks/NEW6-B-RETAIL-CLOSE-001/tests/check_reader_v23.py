"""One focused reader batch; no API and no legacy full-suite replay."""
from pathlib import Path
import sys,json,copy,xml.etree.ElementTree as ET
from collections import defaultdict
from fixture_xml import Fixture,N,NS
from openpyxl.utils.cell import get_column_letter
from evaluate import evaluate,kind
from chart_evidence import chart_facts,raster_charts
from read_candidate import ParsePending
T=Path(__file__).resolve().parents[1];OUT=T/'metadata/reader_v23';F=OUT/'fixtures';F.mkdir(exist_ok=True)
TRUTH=json.loads((T/'solution/oracle.json').read_text());BT=json.loads((T/'solution/bridge_oracle.json').read_text());checks=[]
def check(name,condition,detail):
 checks.append({'case':name,'passed':bool(condition),'detail':detail});print(name,bool(condition),flush=True)
 if not condition:raise AssertionError(name)
def write_rows(f,name,rows):
 root=f.root(name);sd=root.find('m:sheetData',NS);sd[:]=[]
 for rn,vals in enumerate(rows,1):
  row=ET.SubElement(sd,'{'+N+'}row',{'r':str(rn)})
  for cn,v in enumerate(vals,1):
   if v is None:continue
   cell=ET.SubElement(row,'{'+N+'}c',{'r':get_column_letter(cn)+str(rn)})
   if isinstance(v,(int,float)):ET.SubElement(cell,'{'+N+'}v').text=str(v)
   else:cell.attrib['t']='inlineStr';ET.SubElement(ET.SubElement(cell,'{'+N+'}is'),'{'+N+'}t').text=str(v)
 dim=root.find('m:dimension',NS)
 if dim is not None:dim.attrib['ref']=f'A1:{get_column_letter(max(map(len,rows)))}{len(rows)}'
 f.update(name,root)
def run(name,path):
 r=evaluate(path,T/'data/input_files');(OUT/(name+'.receipt.json')).write_text(json.dumps(r,ensure_ascii=False,indent=2));return r
if __name__=='__main__':
 selected=set(sys.argv[1:])
 if not selected or 'workbook' in selected:
  baseline=run('reference',T/'solution/reference.xlsx');check('reference_all_facts',baseline['evaluation_status']=='SCORED' and all(x==1 for x in baseline['criterion_scores'].values()),baseline.get('criterion_scores') or baseline['evidence'])
  f=Fixture(T/'solution/reference.xlsx');by=defaultdict(dict)
  for invoice,cl,count,amount,*rest in TRUTH['invoice_totals']:by[invoice][cl]=(count,float(amount))
  rows=[['InvoiceNo','Country','Sep Sales rows','Sep Sales GBP','Sep Credit rows','Sep Credits GBP','Sep Exception rows','Sep Exception GBP','Oct Sales rows','Oct Sales GBP','Oct Credit rows','Oct Credits GBP','Oct Exception rows','Oct Exception GBP']]
  for inv,cl in by.items():rows.append([inv,'United Kingdom',999,-9876,999,-9876,999,-9876,*[v for cls in ['sale','credit','exception'] for v in cl.get(cls,(0,0))]])
  write_rows(f,'Invoice totals',rows);legal=F/'legal_period_invoice_wide.xlsx';f.save(legal)
  r=run('legal_period_invoice_wide',legal);check('legal_period_invoice_wide',r.get('criterion_scores')==baseline.get('criterion_scores') and r['evaluation_status']=='SCORED',r.get('criterion_scores') or r['evidence'])
  wrong=F/'wrong_invoice_value.xlsx';Fixture(legal).cell('Invoice totals','J2',float(rows[1][9])+100).save(wrong)
  r=run('wrong_invoice_value',wrong);base=baseline['criterion_scores'];scores=r.get('criterion_scores') or {};check('single_invoice_value_only_R003',r['evaluation_status']=='SCORED' and scores['R003']<base['R003'] and all(scores[k]==v for k,v in base.items() if k!='R003'),scores)
 if not selected or 'charts' in selected:
  reg=json.loads((OUT/'visual_claims.json').read_text());cs=[dict(ch,chart='retained labelled PNG') for e in reg.values() for ch in e.get('charts',[])]
  facts,_,bindings=chart_facts(cs,TRUTH,BT,kind);check('labelled_country_and_bridge_PNG',all(facts),bindings)
  bad=copy.deepcopy(cs);bad[0]['series'][0]['values'][0]+=1000;bf,_,_=chart_facts(bad,TRUTH,BT,kind);check('single_chart_value_preserves_other_facts',not bf[1] and all(a==b for i,(a,b) in enumerate(zip(facts,bf)) if i!=1),bf)
  bad=copy.deepcopy(cs);bad[0]['series'][0]['categories'][0]='Unspecified territory'
  try:chart_facts(bad,TRUTH,BT,kind);pending=False
  except ParsePending:pending=True
  check('unknown_chart_role_pending',pending,{'expected':'JUDGE_ERROR/no numeric score'})
  f=Fixture(T/'solution/reference.xlsx');f.parts['xl/media/unregistered.png']=b'unknown-image-byte-sequence';p=F/'unknown_raster.xlsx';f.save(p)
  try:raster_charts(p,T);pending=False
  except ParsePending:pending=True
  check('unknown_image_pending',pending,{'expected':'JUDGE_ERROR/no numeric score'})
 (OUT/('calibration_'+('_'.join(sorted(selected)) or 'all')+'.json')).write_text(json.dumps({'status':'CALIBRATION_PASSED','checks':checks,'api_calls':0,'fixed_facts':'Same R001–R008 and three unchanged profiles; PNG observations hash-bound, no sample-name score.'},ensure_ascii=False,indent=2))
