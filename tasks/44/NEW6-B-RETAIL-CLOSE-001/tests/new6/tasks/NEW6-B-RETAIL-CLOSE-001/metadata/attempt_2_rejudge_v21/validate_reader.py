"""Fixed-scope v2.1 rejudge calibration; no candidate edits or new generation."""
from pathlib import Path
import sys,json,re,xml.etree.ElementTree as ET
from decimal import Decimal
import openpyxl
from openpyxl.utils.cell import get_column_letter,column_index_from_string,range_boundaries
P=Path(__file__).resolve().parent;T=P.parents[1];sys.path.insert(0,str(T/'tests'))
from fixture_xml import Fixture,NS
from evaluate import evaluate
R=json.loads((P/'recalc.json').read_text());ORIG=Path(R['source']);ACT=Path(R['output']);INP=ORIG.parents[1]/'input';F=P/'fixtures';F.mkdir(exist_ok=True)
C={f'R{i:03}' for i in range(1,9)}
def shift(f):
 def address(v):
  m=re.fullmatch(r'([A-Z]+)(\d+)',v);return get_column_letter(column_index_from_string(m[1])+1)+str(int(m[2])+3)
 for name in f.sheets:
  root=f.root(name)
  for row in root.find('m:sheetData',NS):
   row.attrib['r']=str(int(row.attrib['r'])+3)
   for c in row:
    c.attrib['r']=address(c.attrib['r'])
    formula=c.find('m:f',NS)
    if formula is not None:formula.text=re.sub(r'[A-Z]+\d+',lambda m:address(m[0]),formula.text)
  dim=root.find('m:dimension',NS)
  if dim is not None:dim.attrib['ref']=':'.join(address(a) for a in dim.attrib['ref'].split(':'))
  f.update(name,root)
 return f
# Independently audit actual native SUM outputs against the unchanged original input ranges.
a=openpyxl.load_workbook(ORIG,read_only=True,data_only=False);b=openpyxl.load_workbook(ACT,read_only=True,data_only=True);sums=[]
for sa,sb in zip(a,b):
 rows=list(sa.iter_rows())
 for row in rows:
  for c in row:
   if c.data_type!='f':continue
   m=re.fullmatch(r'=SUM\(([A-Z]+\d+:[A-Z]+\d+)\)',c.value)
   if not m:raise ValueError('Unanticipated formula requires supported native audit')
   x,y,xx,yy=range_boundaries(m[1]);expected=sum((Decimal(str(rows[j-1][i-1].value)) for j in range(y,yy+1) for i in range(x,xx+1) if isinstance(rows[j-1][i-1].value,(int,float))),Decimal(0));got=Decimal(str(sb[c.coordinate].value));sums.append({'sheet':sa.title,'cell':c.coordinate,'original_formula':c.value,'independent_range_sum':str(expected),'native_value':str(got),'passed':abs(expected-got)<Decimal('.000001')})
a.close();b.close();assert len(sums)==13 and all(x['passed'] for x in sums)
(P/'native_formula_audit.json').write_text(json.dumps({'original_unchanged':R['original_unchanged'],'native_recalculation_receipt':'recalc.json','formula_count':len(sums),'checks':sums,'new_recalculations':0},indent=2))
base=evaluate(ACT,INP);(P/'same_original_receipt.json').write_text(json.dumps(base,indent=2));print('ACTUAL',base['evaluation_status'],base.get('criterion_scores'),{k:v.get('score_decimal') for k,v in base['profiles'].items()},flush=True);assert base['evaluation_status']=='SCORED'
bf=base['criterion_scores'];cases=[('reference',T/'solution/reference.xlsx',None,[],True),('equivalent_actual_layout',ACT,shift,[],False),('wrong_invoice_amount',ACT,lambda f:f.cell('Invoice Analysis','B4',22306),['R003'],False),('wrong_bridge_quantity',ACT,lambda f:f.cell('Bridge Analysis','B7',45947.3182539217),['R007'],False),('duplicate_sku',ACT,lambda f:f.duplicate_row('SKU Schedule',4),['R008'],False),('mixed_bridge_claim',ACT,lambda f:f.duplicate_row('Bridge Analysis',7,{'B':999}),['R007'],False),('unbound_bridge_amount',ACT,lambda f:f.cell('Bridge Analysis','B3','Unresolved numeric field'),None,False)]
out=[]
for name,source,fn,loss,isref in cases:
 path=source
 if fn:path=F/(name+'.xlsx');fn(Fixture(source)).save(path)
 r=evaluate(path,T/'data/input_files' if isref else INP)
 if loss is None:ok=r['evaluation_status']=='JUDGE_ERROR' and r['normalized_score'] is None
 elif r['evaluation_status']!='SCORED':ok=False
 else:
  expected={k:1 for k in C} if isref else bf;facts=r['criterion_scores'];ok=all(facts[k]<expected[k] for k in loss) and all(facts[k]==expected[k] for k in C-set(loss)) and all(v['criterion_scores']==facts for v in r['profiles'].values())
 rec={'case':name,'passed':ok,'loss_only':loss,'result':r};(F/(name+'.receipt.json')).write_text(json.dumps(rec,indent=2));out.append({'case':name,'passed':ok});print(name,ok,r['evaluation_status'],r.get('criterion_scores'),r['evidence'].get('reason'),flush=True)
receipt={'task_version':'new6-b1-v2','judge_version':'new6-b1-reader-v2.1','passed':all(x['passed'] for x in out),'checks':out,'native_formula_checks':13,'native_formula_all_passed':True,'agent_generation_calls':0,'original_unchanged':R['original_unchanged'],'validation_budget':{'categories':'Actual native formula receipt plus independent range-sum audit; reference, legitimate relocated actual layout, two isolated business errors, duplicates/mixed claims, parser uncertainty.','runner_count':1,'independent_review_count':0,'hash_count':0,'early_stop':'Any unsupported layout retains null score; any failed fixture prevents completion.','actual':'One combined focused runner. Existing v2 calibration not rerun; no duplicate native recalculation.'},'semantics':'Weights and source populations unchanged. Missing category counts recover solely from candidate detail. Blank quantities/values for the explicitly inactive month of a new/exited SKU are zero, not an omission. Zero-population wide categories do not invent source groups.'}
(P/'validation_receipt.json').write_text(json.dumps(receipt,indent=2));assert receipt['passed']
