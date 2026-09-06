"""UCI352 October oracle. Pure Decimal business arithmetic, no candidate reads.

Full official workbook is streamed once; no rows are selected for model failure.
An independent reversed, document-first grouping verifies all classification and
country totals. Physical row identity is never deduplicated.
"""
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from collections import defaultdict, Counter
import csv, json, argparse
import openpyxl

TASK=Path(__file__).resolve().parents[1]
START=datetime(2011,10,1);END=datetime(2011,11,1)
RAW_HEADERS=['source_row_id','InvoiceNo','StockCode','Description','Quantity','InvoiceDate','UnitPrice','CustomerID','Country']
def text(v):
    if v is None:return ''
    if isinstance(v,float) and v.is_integer():return str(int(v))
    return str(v)
def classify(r):
    day=datetime.fromisoformat(r['InvoiceDate'])
    q=Decimal(r['Quantity']) if r['Quantity'] else None
    p=Decimal(r['UnitPrice']) if r['UnitPrice'] else None
    amount=q*p if q is not None and p is not None else None
    missing=not r['CustomerID']
    if not START<=day<END:return 'outside_scope',amount,missing,'outside October reporting window'
    essential=all(r[k] for k in ['InvoiceNo','StockCode','Description','Quantity','InvoiceDate','UnitPrice','Country'])
    credit=r['InvoiceNo'].upper().startswith('C')
    if essential and q>0 and p>0 and not credit:return 'sale',amount,missing,''
    if essential and q<0 and p>0 and credit:return 'credit',amount,missing,''
    reasons=[]
    if not essential:reasons.append('missing essential field')
    if p is not None and p<=0:reasons.append('nonpositive unit price')
    if q==0:reasons.append('zero quantity')
    if q is not None and ((q<0 and not credit) or (q>0 and credit)):reasons.append('quantity and invoice identity disagree')
    return 'exception',amount,missing,'; '.join(reasons) or 'invalid sign or identity combination'

def records(original):
    w=openpyxl.load_workbook(original,data_only=True,read_only=True)
    for s in w:
        headers=[c.value for c in next(s.iter_rows())]
        assert headers==RAW_HEADERS[1:],headers
        for num,row in enumerate(s.iter_rows(min_row=2,values_only=True),2):
            day=row[4]
            if not isinstance(day,datetime):raise ValueError(f'Unusable timestamp at {s.title}:{num}')
            if datetime(2011,9,30)<=day<datetime(2011,11,2):
                vals=[f'{s.title}:{num}',*[day.isoformat(sep=' ') if i==4 else text(v) for i,v in enumerate(row)]]
                yield dict(zip(RAW_HEADERS,vals))

def aggregate(rows):
    totals=defaultdict(lambda:[0,Decimal(0)])
    invoices=defaultdict(lambda:[0,Decimal(0)])
    countries=defaultdict(lambda:[0,Decimal(0)])
    classified=[];exceptions=[];invoice_countries=defaultdict(set)
    for r in rows:
        status,amount,missing,reason=classify(r)
        classified.append([*[r[k] for k in RAW_HEADERS],status,str(amount) if amount is not None else None,missing,reason])
        totals[status][0]+=1;totals[status][1]+=amount or Decimal(0)
        if status=='outside_scope':continue
        key=(r['InvoiceNo'],status);invoices[key][0]+=1;invoices[key][1]+=amount or Decimal(0)
        ckey=(r['Country'],status);countries[ckey][0]+=1;countries[ckey][1]+=amount or Decimal(0)
        invoice_countries[r['InvoiceNo']].add(r['Country'])
        if status=='exception' or missing:
            exceptions.append([r['source_row_id'],r['InvoiceNo'],status,'business_exception' if status=='exception' else 'missing_customer_attribution',reason or 'missing CustomerID',str(amount) if amount is not None else None])
    # Independent document-first calculation, reverse occurrence order, independent
    # reexpression of policy and then regroup by country and status.
    checks=defaultdict(lambda:[0,Decimal(0)]);docs=defaultdict(list)
    for r in reversed(rows):docs[r['InvoiceNo']].append(r)
    for doc,occurrences in docs.items():
        for r in occurrences:
            day=datetime.strptime(r['InvoiceDate'],'%Y-%m-%d %H:%M:%S')
            q,p=Decimal(r['Quantity']),Decimal(r['UnitPrice']);v=q*p
            ordinary=(q>0 and p>0 and doc[:1].upper()!='C')
            cancellation=(q<0 and p>0 and doc[:1].upper()=='C')
            valid=all(r[k]!='' for k in RAW_HEADERS if k not in ['source_row_id','CustomerID'])
            status='outside_scope' if day.month!=10 else 'sale' if ordinary and valid else 'credit' if cancellation and valid else 'exception'
            for key in [('total',status),('invoice',doc,status),('country',r['Country'],status)]:
                if key[0]!='total' and status=='outside_scope':continue
                checks[key][0]+=1;checks[key][1]+=v
    assert all(checks[('total',k)]==v for k,v in totals.items())
    assert all(checks[('invoice',*k)]==v for k,v in invoices.items())
    assert all(checks[('country',*k)]==v for k,v in countries.items())
    totals_out=[[k,v[0],str(v[1])] for k,v in sorted(totals.items())]
    business_occurrences=Counter(tuple(r[k] for k in RAW_HEADERS[1:]) for r in rows)
    return {'raw_headers':RAW_HEADERS,'classified_headers':RAW_HEADERS+['Classification','RecordedAmountGBP','MissingCustomer','TreatmentReason'],
      'classified':classified,'invoice_totals':[[*k,v[0],str(v[1]),'; '.join(sorted(invoice_countries[k[0]]))] for k,v in sorted(invoices.items())],
      'country_totals':[[*k,v[0],str(v[1])] for k,v in sorted(countries.items())],
      'exceptions':exceptions,'totals':totals_out,
      'net_recorded_value':str(totals['sale'][1]+totals['credit'][1]),
      'source_occurrences':len(rows),'repeated_looking_extra_occurrences':sum(n-1 for n in business_occurrences.values() if n>1),
      'missing_customer_in_scope':sum(bool(r[11]) for r in classified if r[9]!='outside_scope'),
      'invoices_with_country_conflict':{k:sorted(v) for k,v in invoice_countries.items() if len(v)>1},
      'independent_verification':'Forward occurrence Decimal totals matched independent reversed document-first policy/grouping for every invoice, country and status.'}

if __name__=='__main__':
    a=argparse.ArgumentParser();a.add_argument('--source',required=True);args=a.parse_args()
    rows=list(records(args.source));out=aggregate(rows)
    (TASK/'solution/oracle.json').write_text(json.dumps(out,ensure_ascii=False,separators=(',',':')))
    # Private source extraction is released only after the oracle/reference step.
    with (TASK/'metadata/source_extract_private.csv').open('w',newline='') as f:
        writer=csv.DictWriter(f,fieldnames=RAW_HEADERS);writer.writeheader();writer.writerows(rows)
    print(json.dumps({k:v for k,v in out.items() if k in ['totals','source_occurrences','net_recorded_value','repeated_looking_extra_occurrences','missing_customer_in_scope','invoices_with_country_conflict','independent_verification']},indent=2))
