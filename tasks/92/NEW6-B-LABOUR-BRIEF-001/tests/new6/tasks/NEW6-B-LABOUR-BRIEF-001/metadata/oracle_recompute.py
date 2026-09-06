"""Independent source-only truth for the three-release briefing update."""
from pathlib import Path
from decimal import Decimal,ROUND_HALF_UP,InvalidOperation
import json,re,zipfile,posixpath,xml.etree.ElementTree as ET
import openpyxl
TASK=Path(__file__).resolve().parents[1]
PREFIX=('E06','E07','E08','E09')
FILES={'2023':'ons_li01_january2023_corrected.xlsx','2024':'ons_li01_january2024.xlsx','2025':'ons_li01_january2025.xlsx'}
NS={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
def printed(v):
    if not isinstance(v,(int,float,Decimal)) or isinstance(v,bool):return None
    return str(Decimal(str(v)).quantize(Decimal('.1'),rounding=ROUND_HALF_UP))
def short(v):return str(v or '').strip().split('\n')[0]
def xml_table(path):
    with zipfile.ZipFile(path) as z:
        strings=[]
        if 'xl/sharedStrings.xml' in z.namelist():
            strings=[''.join(t.text or '' for t in si.iter('{'+NS['m']+'}t')) for si in ET.fromstring(z.read('xl/sharedStrings.xml'))]
        book=ET.fromstring(z.read('xl/workbook.xml'))
        sheet=next(s for s in book.find('m:sheets',NS) if s.attrib['name']=='LI01')
        rels={r.attrib['Id']:r.attrib['Target'] for r in ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))}
        target=rels[sheet.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']]
        target=target.lstrip('/') if target.startswith('/') else posixpath.normpath('xl/'+target)
        output=[]
        for row in ET.fromstring(z.read(target)).findall('m:sheetData/m:row',NS):
            values={}
            for cell in row:
                value=cell.find('m:v',NS);kind=cell.get('t');raw=None
                if value is not None and value.text is not None:
                    raw=strings[int(value.text)] if kind=='s' else value.text
                    if kind not in ['s','str','e','b']:
                        try:raw=Decimal(raw)
                        except InvalidOperation:pass
                elif kind=='inlineStr':raw=''.join(t.text or '' for t in cell.iter('{'+NS['m']+'}t'))
                values[re.sub(r'\d','',cell.attrib['r'])]=raw
            output.append(values)
    headers=next({short(v):c for c,v in row.items()} for row in output if 'Geography code' in [short(v) for v in row.values()])
    fields={k:headers[v] for k,v in [('code','Geography code'),('name','Geography'),('employment','Employment rate'),('unemployment','Unemployment rate')]}
    return {str(row[fields['code']]):{k:row.get(col) for k,col in fields.items()} for row in output if str(row.get(fields['code'],'')).startswith(PREFIX)}
def read_source(year):
    path=TASK/'data/input_files'/FILES[year];w=openpyxl.load_workbook(path,read_only=True,data_only=True);s=w['LI01']
    row=next(row for row in s.iter_rows(max_row=15) if any(short(c.value)=='Geography code' for c in row))
    header={short(c.value):c for c in row if c.value is not None}
    fields={'code':'Geography code','name':'Geography','employment':'Employment rate','unemployment':'Unemployment rate'}
    columns={k:header[label].column for k,label in fields.items()}
    records={}
    for row in s.iter_rows(min_row=row[0].row+1):
        code=row[columns['code']-1].value
        if not isinstance(code,str) or not code.startswith(PREFIX):continue
        assert code not in records,('duplicate source code',year,code)
        record={'code':code,'name':row[columns['name']-1].value}
        for metric in ['employment','unemployment']:
            cell=row[columns[metric]-1];record[metric]=printed(cell.value);record[metric+'_raw']=str(cell.value)
            record[metric+'_source']=FILES[year]+'#LI01!'+cell.coordinate
        records[code]=record
    independent=xml_table(path);assert set(independent)==set(records)
    for code,record in records.items():
        assert independent[code]['name']==record['name']
        for metric in ['employment','unemployment']:assert printed(independent[code][metric])==record[metric],(year,code,metric)
    context={'employment_header':header['Employment rate'].value,'unemployment_header':header['Unemployment rate'].value,
             'cover':[str(c.value) for row in w['Cover_sheet'] for c in row if c.value is not None],
             'notes':[str(c.value) for row in w['Notes'] for c in row if c.value is not None],
             'correction':[str(c.value) for row in w['Note'] for c in row if c.value is not None] if 'Note' in w.sheetnames else []}
    assert 'age 16 to 64' in context['employment_header']
    assert 'age 16 and older' in context['unemployment_header']
    assert any('economically active' in x.lower() for x in context['notes'])
    w.close();return records,context
def delta(a,b):return str(Decimal(b)-Decimal(a)) if a is not None and b is not None else None
def recompute():
    sources={};contexts={}
    for year in FILES:sources[year],contexts[year]=read_source(year)
    codes=sorted(set().union(*(set(v) for v in sources.values())));panel=[]
    for code in codes:
        names=[sources[y][code]['name'] for y in FILES if code in sources[y]]
        row={'code':code,'name':names[-1],'name_aliases':list(dict.fromkeys(names)),'missing':[]}
        for year in FILES:
            for metric in ['employment','unemployment']:
                value=sources[year].get(code,{}).get(metric);row[metric+'_'+year]=value
                if value is None:row['missing'].append({'year':year,'metric':metric,'reason':'geography absent' if code not in sources[year] else 'source value unavailable or suppressed','source_token':sources[year].get(code,{}).get(metric+'_raw')})
        for metric in ['employment','unemployment']:
            row[metric+'_change_23_24']=delta(row[metric+'_2023'],row[metric+'_2024'])
            row[metric+'_change_24_25']=delta(row[metric+'_2024'],row[metric+'_2025'])
            row[metric+'_change_23_25']=delta(row[metric+'_2023'],row[metric+'_2025'])
        previous_available=all(row[m+'_'+y] is not None for m in ['employment','unemployment'] for y in ['2023','2024'])
        row['previous_eligible']=previous_available and Decimal(row['unemployment_change_23_24'])>0 and Decimal(row['employment_change_23_24'])<0
        row['comparable']=not row['missing']
        row['eligible']=row['comparable'] and all(Decimal(row['unemployment_change_'+period])>0 and Decimal(row['employment_change_'+period])<0 for period in ['23_24','24_25'])
        panel.append(row)
    previous=sorted([r for r in panel if r['previous_eligible']],key=lambda r:(-Decimal(r['unemployment_change_23_24']),r['code']))[:5]
    shortlist=sorted([r for r in panel if r['eligible']],key=lambda r:(-Decimal(r['unemployment_change_23_25']),r['code']))[:5]
    groups={}
    for row in panel:
        if row['eligible']:groups.setdefault(Decimal(row['unemployment_change_23_25']),[]).append(row['code'])
    assert [r['code'] for r in shortlist]==[code for amount in sorted(groups,reverse=True) for code in sorted(groups[amount])][:5]
    old={r['code'] for r in previous};new={r['code'] for r in shortlist}
    out={'source_records':sources,'source_context':contexts,'panel':panel,
         'previous_shortlist':[r['code'] for r in previous],'shortlist':[r['code'] for r in shortlist],
         'eligible_codes':[r['code'] for r in panel if r['eligible']],
         'movements':{'retained':sorted(old&new),'entered':sorted(new-old),'left':sorted(old-new)},
         'counts':{'union':len(panel),'three_period_comparable':sum(r['comparable'] for r in panel),'eligible':sum(r['eligible'] for r in panel),'previous_eligible':sum(r['previous_eligible'] for r in panel),'selected':len(shortlist)},
         'independent_verification':'Every eligible-source geography/name/published rate matched a second OOXML reader; shortlist matched independent grouped sort.'}
    return out
if __name__=='__main__':
    result=recompute();(TASK/'solution/oracle.json').write_text(json.dumps(result,ensure_ascii=False,indent=2)+'\n')
    (TASK/'metadata/source_records.json').write_text(json.dumps({'releases':result['source_records'],'context':result['source_context']},ensure_ascii=False,indent=2)+'\n')
    print(json.dumps({k:result[k] for k in ['counts','previous_shortlist','shortlist','movements','independent_verification']},ensure_ascii=False,indent=2))
