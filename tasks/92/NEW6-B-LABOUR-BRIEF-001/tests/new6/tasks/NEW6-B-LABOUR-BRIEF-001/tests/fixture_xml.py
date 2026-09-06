"""Deterministic task-fixture mutations preserving unrelated OOXML bytes.

Artifact-tool authored the reference. These deliberately controlled binary test
mutations are not a candidate formula interpreter and never edit the reference.
"""
from pathlib import Path
from copy import deepcopy
import zipfile,re,posixpath,xml.etree.ElementTree as ET
from openpyxl.utils.cell import get_column_letter,column_index_from_string
N='http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS={'m':N,'c':'http://schemas.openxmlformats.org/drawingml/2006/chart'}
ET.register_namespace('',N)
class Fixture:
    def __init__(self,path):
        with zipfile.ZipFile(path) as z:self.parts={n:z.read(n) for n in z.namelist()}
        wb=ET.fromstring(self.parts['xl/workbook.xml']);rels=ET.fromstring(self.parts['xl/_rels/workbook.xml.rels'])
        links={r.attrib['Id']:posixpath.normpath(posixpath.join('xl',r.attrib['Target'])).lstrip('/') for r in rels}
        self.sheets={s.attrib['name']:links[s.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']] for s in wb.find('m:sheets',NS)}
    def root(self,name):return ET.fromstring(self.parts[self.sheets[name]])
    def update(self,name,root):self.parts[self.sheets[name]]=ET.tostring(root,encoding='utf-8',xml_declaration=True)
    def cell(self,name,ref,value=None,formula=None,uncached=False):
        root=self.root(name);sd=root.find('m:sheetData',NS);rn=int(re.search(r'\d+',ref)[0])
        row=next((r for r in sd if int(r.attrib['r'])==rn),None)
        if row is None:row=ET.SubElement(sd,'{'+N+'}row',{'r':str(rn)})
        c=next((c for c in row if c.attrib['r']==ref),None)
        if c is None:c=ET.SubElement(row,'{'+N+'}c',{'r':ref})
        for child in list(c):c.remove(child)
        c.attrib.pop('t',None)
        if formula is not None:
            ET.SubElement(c,'{'+N+'}f').text=formula.lstrip('=')
            if not uncached:ET.SubElement(c,'{'+N+'}v').text=str(value)
        elif value is None:pass
        elif isinstance(value,(int,float)):
            ET.SubElement(c,'{'+N+'}v').text=str(value)
        else:
            c.attrib['t']='inlineStr';si=ET.SubElement(c,'{'+N+'}is');ET.SubElement(si,'{'+N+'}t').text=str(value)
        dim=root.find('m:dimension',NS)
        if dim is not None:
            old=dim.attrib['ref'].split(':')[-1];oc=column_index_from_string(re.sub(r'\d','',old));orr=int(re.search(r'\d+',old)[0]);nc=column_index_from_string(re.sub(r'\d','',ref))
            dim.attrib['ref']='A1:'+get_column_letter(max(oc,nc))+str(max(orr,rn))
        self.update(name,root);return self
    def clear_sheet(self,name):
        root=self.root(name);sd=root.find('m:sheetData',NS)
        for row in list(sd):sd.remove(row)
        dim=root.find('m:dimension',NS)
        if dim is not None:dim.attrib['ref']='A1'
        self.update(name,root);return self
    def shared_categories(self):
        for name,data in list(self.parts.items()):
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',name):
                root=ET.fromstring(data);sers=root.findall('.//c:ser',NS)
                assert len(sers)>1 and sers[0].find('c:cat',NS) is not None
                for ser in sers[1:]:
                    el=ser.find('c:cat',NS)
                    if el is not None:ser.remove(el)
                self.parts[name]=ET.tostring(root,encoding='utf-8',xml_declaration=True);return self
        raise ValueError('No compatible chart for shared-category fixture')
    def remove_row(self,name,rn):
        root=self.root(name);sd=root.find('m:sheetData',NS);row=next(r for r in sd if int(r.attrib['r'])==rn);sd.remove(row);self.update(name,root);return self
    def duplicate_row(self,name,rn,changes=None):
        root=self.root(name);sd=root.find('m:sheetData',NS);row=deepcopy(next(r for r in sd if int(r.attrib['r'])==rn));nr=max(int(r.attrib['r']) for r in sd)+1
        row.attrib['r']=str(nr)
        for c in row:c.attrib['r']=re.sub(r'\d+',str(nr),c.attrib['r'])
        sd.append(row);dim=root.find('m:dimension',NS)
        if dim is not None:dim.attrib['ref']=re.sub(r'\d+$',str(nr),dim.attrib['ref'])
        self.update(name,root)
        for col,v in (changes or {}).items():self.cell(name,f'{col}{nr}',v)
        return self
    def stale_cache(self):
        for name,data in self.parts.items():
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',name):
                root=ET.fromstring(data);ref=root.find('.//c:ser/c:val/c:numRef',NS);cache=ref.find('c:numCache',NS)
                if cache is None:cache=ET.SubElement(ref,'{'+NS['c']+'}numCache')
                for child in list(cache):cache.remove(child)
                ET.SubElement(cache,'{'+NS['c']+'}ptCount',{'val':'1'})
                p=ET.SubElement(cache,'{'+NS['c']+'}pt',{'idx':'0'});ET.SubElement(p,'{'+NS['c']+'}v').text='99999'
                self.parts[name]=ET.tostring(root,encoding='utf-8',xml_declaration=True);return self
        raise ValueError('Fixture requested stale chart but reference has no chart')
    def layout(self):
        maps={}
        for name,part in self.sheets.items():
            root=self.root(name);sd=root.find('m:sheetData',NS);rows=list(sd);nr=max(int(r.attrib['r']) for r in rows)
            nc=max(column_index_from_string(re.sub(r'\d','',c.attrib['r'])) for r in rows for c in r)
            cm={i:(i%nc)+2 for i in range(1,nc+1)}
            def rm(r):return 4 if r==1 else nr-r+5
            for r in rows:
                old=int(r.attrib['r']);r.attrib['r']=str(rm(old))
                for c in r:
                    oldc=column_index_from_string(re.sub(r'\d','',c.attrib['r']));c.attrib['r']=get_column_letter(cm[oldc])+str(rm(old))
                r[:]=sorted(r,key=lambda c:column_index_from_string(re.sub(r'\d','',c.attrib['r'])))
            sd[:]=sorted(rows,key=lambda r:int(r.attrib['r']))
            dim=root.find('m:dimension',NS)
            if dim is not None:dim.attrib['ref']=f'B4:{get_column_letter(nc+1)}{nr+3}'
            self.update(name,root);maps[name]=(cm,nr)
        for name,data in list(self.parts.items()):
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',name):
                root=ET.fromstring(data)
                for f in root.findall('.//c:f',NS):
                    m=re.fullmatch(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?",f.text)
                    if not m:raise ValueError('Fixture chart relocation is unsupported')
                    cm,nr=maps[m[1]];r1=int(m[3]);r2=int(m[5] or m[3]);a=4 if r1==1 else nr-r1+5;b=4 if r2==1 else nr-r2+5
                    col1=get_column_letter(cm[column_index_from_string(m[2])]);col2=get_column_letter(cm[column_index_from_string(m[4] or m[2])]);lo,hi=sorted([a,b])
                    f.text=f"'{m[1]}'!${col1}${lo}"+(f':${col2}${hi}' if m[5] else '')
                self.parts[name]=ET.tostring(root,encoding='utf-8',xml_declaration=True)
        return self
    def save(self,path):
        Path(path).parent.mkdir(parents=True,exist_ok=True)
        with zipfile.ZipFile(path,'w',zipfile.ZIP_DEFLATED,compresslevel=1) as z:
            for n,data in self.parts.items():z.writestr(n,data)
        return str(path)
