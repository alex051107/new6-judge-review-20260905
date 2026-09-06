"""Evidence-backed USPS evaluator. Candidate regions discovered from labels.

Supported: labelled long rate tables, any sheet/row/column order, aliases below,
literal extraction and any LibreOffice-supported formula implementation.
Unrecognized legitimate layout/engine is JUDGE_ERROR with no score.
"""
from __future__ import annotations
from pathlib import Path
from collections import Counter,defaultdict
from decimal import Decimal as D
import argparse,csv,json,re,sys,zipfile,xml.etree.ElementTree as ET
from datetime import datetime,date
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT=Path(__file__).resolve().parents[1]
sys.path[:0]=[str(ROOT.parents[1]/'common'),str(ROOT/'metadata')]
from runtime import output_status,score_profiles,recalculate_xlsx,RecalcUnavailable,sha256
from oracle_recompute import compute
RATE_FIELDS={'service','upper_bound','weight_unit','zone','usd'}
QUOTE_FIELDS={'request_id','weight','weight_unit','zone','priority_usd','ground_usd','selected','selected_usd'}
ALIASES={
 'service':['service','product','mail service'], 'upper_bound':['upper bound','weight not over','maximum weight','band upper bound'],
 'weight_unit':['weight unit','unit','units'], 'zone':['zone','destination zone'],
 'usd':['usd','rate usd','rate','price usd','retail price','price'], 'effective_date':['effective date','effective'],
 'source_page':['source page','page','pdf page','printed page','source reference','source citation','citation'], 'request_id':['request id','request','quote id','shipment id'],
 'weight':['weight','request weight','parcel weight'], 'priority_usd':['priority usd','priority mail','priority mail price','priority mail usd','priority price'],
 'ground_usd':['ground usd','ground advantage','ground advantage price','ground advantage usd','ground price'],
 'selected':['selected','selected service','chosen service','lower price option','cheapest service'],
 'selected_usd':['selected usd','chosen price','selected price','chosen usd','lowest price','lower price usd']}
def norm(v):return re.sub(r'[^a-z0-9]+',' ',str(v or '').lower()).strip()
ALIAS={norm(a):k for k,vs in ALIASES.items() for a in vs}
def dec(v):
    try:
        d=D(str(v).replace('$','').replace(',',''))
        return d if d.is_finite() else None
    except Exception:return None
def same(a,b):return dec(a) is not None and dec(b) is not None and abs(dec(a)-dec(b))<=D('0.0000001')
def service(v):
    n=norm(v)
    if n in ('priority','priority mail','priority mail retail','usps priority mail'):return 'priority'
    if n in ('ground','ground advantage','usps ground advantage','usps ground advantage retail','ground advantage retail'):return 'ground'
    return n
def unit(v):
    n=norm(v)
    return 'lb' if n in ('lb','lbs','pound','pounds') else 'oz' if n in ('oz','ounce','ounces') else n

def effective_date(v):
    if isinstance(v,datetime):return v.date()==date(2026,7,12)
    if isinstance(v,date):return v==date(2026,7,12)
    if isinstance(v,(int,float)):
        try:return from_excel(v).date()==date(2026,7,12)
        except Exception:return False
    for fmt in ('%Y-%m-%d','%m/%d/%Y','%d %B %Y','%B %d, %Y','%d-%b-%Y','%d %b %Y'):
        try:
            if datetime.strptime(str(v),fmt).date()==date(2026,7,12):return True
        except ValueError:pass
    return False

def page_ref(v,page):
    return page in [int(x) for x in re.findall(r'(?<!\d)\d+(?!\d)',str(v))]

class ParsePending(RuntimeError):
    def __init__(self,message,details=None):
        super().__init__(message);self.details=details or {}

def require_native_results(path):
    computed=load_workbook(path,data_only=True)
    unresolved=[f'{s.title}!{c.coordinate}: {c.value}' for s in computed for row in s for c in row if c.data_type=='e' and c.value in ('#NAME?','#VALUE!','#N/A')]
    if unresolved:raise ParsePending('Formula engine or legal implementation could not be resolved; manual/native distinction required: '+', '.join(unresolved[:8]))

def wide_rate_table(sheet, header, zone_columns):
    """Bind a labelled weight-by-zone matrix without inspecting expected prices."""
    weight_cells=[c for c in sheet[header] if re.search(r'weight|upper.?bound',str(c.value or ''),re.I)]
    if len(weight_cells)!=1:return None
    weight_cell=weight_cells[0];heading=str(weight_cell.value)
    scope=[];svc=None
    for row in reversed(list(sheet.iter_rows(min_row=max(1,header-8),max_row=header-1))):
        text=' | '.join(str(c.value) for c in row if c.value is not None)
        if text:scope.insert(0,text)
        names=cited_services(text)
        if len(names)==1:svc=next(iter(names));break
        if len(names)>1:return None
    if not svc:return None
    context=' | '.join(scope+[heading,'zone table'])
    zones={c:int(re.search(r'[1-8]$',norm(sheet.cell(header,c).value)).group()) for c in zone_columns}
    table={'sheet':sheet.title,'header_row':header,'kind':'wide_rate','fields':{'upper_bound':weight_cell.column},'record_count':0,'nonempty_body_rows':0,'service_binding':scope,'oracle_values_used_for_discovery':False}
    records=[]
    for rid in range(header+1,sheet.max_row+1):
        value=sheet.cell(rid,weight_cell.column).value
        prices=[sheet.cell(rid,c).value for c in zone_columns]
        if value is None and all(v is None for v in prices):
            if records:break
            continue
        if cited_services(str(value)) or re.search(r'weight|upper.?bound',str(value or ''),re.I):break
        if all(v is None for v in prices):
            if records:break
            return None
        text=str(value or '').strip()
        label=re.fullmatch(r'(?:([0-9.]+)\s*[-–]\s*)?([0-9.]+)\s*(oz|ounces?|lbs?|pounds?)?',text,re.I)
        if not label:return None
        unit_text=label.group(3)
        if not unit_text:
            found=re.search(r'\b(oz|ounces?|lbs?|pounds?)\b',heading,re.I)
            unit_text=found.group(1) if found else None
        if not unit_text:return None
        for col,zone in zones.items():
            records.append({'service':svc,'upper_bound':label.group(2),'weight_unit':unit(unit_text),'zone':zone,'usd':sheet.cell(rid,col).value,'source_page':context if re.search(r'notice\s*123|\.pdf|source',context,re.I) else None,'_table':{'sheet':sheet.title,'header_row':header},'_cells':{'upper_bound':(sheet.title,sheet.cell(rid,weight_cell.column).coordinate),'usd':(sheet.title,sheet.cell(rid,col).coordinate)}})
        table['nonempty_body_rows']+=1
    table['record_count']=len(records)
    return (records,table) if records else None

def parse(path):
    w=load_workbook(path,data_only=True);rates=[];quotes=[];totals=[];texts=[];tables=[];unbound_grids=[];visible_rows=[]
    for s in w:
        if s.max_row>50000 or s.max_column>300:raise ParsePending('Workbook exceeds reviewed reader bound')
        headers=[]
        for row in s:
            vals=[c.value for c in row];texts += [str(v) for v in vals if isinstance(v,str)]
            if any(v is not None for v in vals):visible_rows.append({'sheet':s.title,'row':row[0].row,'values':vals,'text':' | '.join(str(v) for v in vals if v is not None)})
            mapping={ALIAS[norm(c.value)]:c.column for c in row if norm(c.value) in ALIAS}
            kind='rate' if RATE_FIELDS<=set(mapping) else 'quote' if QUOTE_FIELDS<=set(mapping) else None
            if kind:headers.append((row[0].row,kind,mapping))
            # A recognized quote table is not evidence that a separate rate grid
            # was parsed. Identify unbound wide structures from public labels
            # and numeric shape only; never compare values to the Oracle.
            zone_columns=[c.column for c in row if re.fullmatch(r'(?:zone|z)\s*[1-8]',norm(c.value))]
            if len(set(zone_columns))>=2:
                nearby=' '.join(str(c.value or '') for rr in s.iter_rows(min_row=max(1,row[0].row-2),max_row=row[0].row) for c in rr)
                weight_heading=bool(re.search(r'weight|upper.?bound|\boz\b|\blbs?\b',nearby,re.I))
                numeric_body=any(sum(dec(s.cell(rid,c).value) is not None for c in zone_columns)>=2 for rid in range(row[0].row+1,min(s.max_row,row[0].row+3)+1))
                if kind is None and weight_heading and numeric_body:
                    bound=wide_rate_table(s,row[0].row,zone_columns)
                    if bound:
                        found,table=bound;rates.extend(found);tables.append(table)
                    else:unbound_grids.append({'sheet':s.title,'header_row':row[0].row,'zone_columns':zone_columns,'reason':'wide weight/zone grid has not been semantically bound'})
            for c in row:
                if norm(c.value) in ('batch total','batch total usd','total batch usd','total postage','batch postage'):
                    candidates=[x for x in row[c.column:] if x.value is not None]
                    if candidates:totals.append({'value':candidates[0].value,'cell':f'{s.title}!{candidates[0].coordinate}'})
        for hi,(start,kind,m) in enumerate(headers):
            end=headers[hi+1][0]-1 if hi+1<len(headers) else s.max_row
            table={'sheet':s.title,'header_row':start,'kind':kind,'fields':m,'record_count':0,'nonempty_body_rows':0};tables.append(table)
            for rid in range(start+1,end+1):
                row={f:s.cell(rid,c).value for f,c in m.items()}
                if any(v is not None for v in row.values()):table['nonempty_body_rows']+=1
                if kind=='rate':
                    if not row.get('service') or dec(row.get('upper_bound')) is None:continue
                    row['service']=service(row['service']);row['weight_unit']=unit(row['weight_unit']);rates.append(row)
                else:
                    if not row.get('request_id') or not re.match(r'^Q\d+$',str(row['request_id'])):continue
                    row['weight_unit']=unit(row['weight_unit']);row['selected']=service(row['selected']);quotes.append(row)
                row['_table']={'sheet':s.title,'header_row':start}
                row['_cells']={f:(s.title,s.cell(rid,c).coordinate) for f,c in m.items()}
                table['record_count']+=1
    if not tables:raise ParsePending('No supported labelled tables; needs semantic review or layout adapter')
    if unbound_grids:
        raise ParsePending('Rate extraction is only partially parsed: unbound wide rate grid requires semantic review or a layout adapter',{'bound_tables':tables,'unbound_rate_grids':unbound_grids,'parsed_rate_rows':len(rates),'parsed_quote_rows':len(quotes),'oracle_values_used_for_discovery':False})
    delivery={'state':'BOUND','parsed_rate_rows':len(rates)}
    if not rates:
        empty_rate_tables=[t for t in tables if t['kind']=='rate' and t['nonempty_body_rows']==0]
        explicit_omission=bool(re.search(r'(?:rate (?:table|grid)|tariff (?:table|grid)|rates)\s+(?:was |were |is |are )?(?:omitted|not (?:provided|included|delivered)|missing)', '\n'.join(texts),re.I))
        if empty_rate_tables or explicit_omission:
            delivery={'state':'CONFIRMED_OMITTED','empty_labelled_rate_tables':empty_rate_tables,'explicit_candidate_omission_statement':explicit_omission}
        else:
            raise ParsePending('Quote data was parsed but rate delivery cannot be determined safely; no bound rate rows and no clear omission evidence',{'bound_tables':tables,'parsed_quote_rows':len(quotes),'rate_delivery':'UNRESOLVED','oracle_values_used_for_discovery':False})
    return {'rates':rates,'quotes':quotes,'totals':totals,'text':'\n'.join(texts),'tables':tables,'rate_delivery':delivery,'visible_rows':visible_rows}

# Bind provenance from candidate-visible labels and scope, never Oracle prices or
# reference positions. Explicit row citations take precedence over shared notes.
SERVICE_TEXT=re.compile(r'(?:USPS\s+)?Ground\s+Advantage|Priority\s+Mail',re.I)
def cited_services(text):
    return {'ground' if 'ground' in m.group().lower() else 'priority' for m in SERVICE_TEXT.finditer(text)}
def citation_value(value,svc):
    text=str(value or '').strip()
    if not text:return None
    # Numeric row references and explicitly labelled page references are precise.
    nums=re.findall(r'(?<!\d)\d+(?!\d)',text)
    if re.fullmatch(r'\d+(?:\.0)?',text) or re.search(r'\b(?:page|pages|p|pp)\.?\s*\d',text,re.I):
        return page_ref(text,5 if svc=='priority' else 7)
    # Equivalent locator names the unique Retail weight/zone service table.
    if svc in cited_services(text) and re.search(r'\bretail\b',text,re.I) and re.search(r'weight|zone|table|rate',text,re.I):return True
    return None

def rate_source_trace(row,candidate):
    svc=row['service'];explicit=row.get('source_page')
    if explicit is not None and str(explicit).strip():
        ok=citation_value(explicit,svc)
        if ok is None:raise ParsePending('A substantive row citation needs semantic binding',{'citation':str(explicit),'service':svc})
        return ok,{'binding':'row','citation':str(explicit)}
    table=row['_table'];same_table=[r for r in candidate['rates'] if r['_table']==table]
    scope={r['service'] for r in same_table};bound=[];unresolved=[]
    for entry in candidate['visible_rows']:
        text=entry['text']
        # Exclude the already parsed data and column-header rows themselves.
        if any(entry['sheet']==r['_table']['sheet'] and entry['row']==int(re.search(r'\d+$',next(iter(r['_cells'].values()))[1]).group()) for r in candidate['rates']):continue
        if entry['sheet']==table['sheet'] and entry['row']==table['header_row']:continue
        if not re.search(r'notice\s*123|source|\.pdf|\b(?:page|pages|p|pp)\.?\s*\d',text,re.I):continue
        for part in re.split(r';|\n',text):
            services=cited_services(part)
            local=entry['sheet']==table['sheet']
            if not services and local and len(scope)==1:services=scope
            if not services and local:services=cited_services(entry['sheet'])
            if svc not in services:
                if not services and re.search(r'\.pdf|\b(?:page|pages|p|pp)\.?\s*\d',part,re.I):unresolved.append(entry)
                continue
            if len(services)!=1:
                unresolved.append(entry);continue
            ok=citation_value(part,svc)
            if ok is None:
                # A service-specific source statement without a precise locator
                # may be legitimate; it is not evidence of a business omission.
                unresolved.append(entry)
            else:bound.append({'binding':'shared','sheet':entry['sheet'],'row':entry['row'],'citation':part,'correct':ok})
    if bound:return all(x['correct'] for x in bound),bound
    if unresolved:raise ParsePending('Shared source citation exists but cannot be bound to a Retail service/weight table',{'service':svc,'citations':unresolved,'oracle_values_used_for_discovery':False})
    return False,{'binding':'missing','service':svc}

def effective_date_fact(candidate):
    dates=[r.get('effective_date') for r in candidate['rates']]
    supplied=[v for v in dates if v is not None and str(v).strip()]
    if supplied and not all(effective_date(v) for v in supplied):return False,{'row_dates':list(map(str,supplied))}
    if dates and len(supplied)==len(dates):return True,{'binding':'row','dates':list(map(str,set(supplied)))}
    shared=[]
    for entry in candidate['visible_rows']:
        if not re.search(r'effective(?:\s+date)?',entry['text'],re.I):continue
        # A standard date may be in the adjacent cell, including a native Excel
        # date or serial, or embedded in the labelled global source sentence.
        values=[v for v in entry['values'] if isinstance(v,(date,datetime,int,float))]
        values += re.findall(r'\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4}|\d{1,2} [A-Za-z]+ \d{4}|[A-Za-z]+ \d{1,2}, \d{4}|\d{1,2}-[A-Za-z]{3}-\d{4})\b',entry['text'])
        if values:shared += [{'value':str(v),'correct':effective_date(v),'sheet':entry['sheet'],'row':entry['row']} for v in values]
    return bool(dates) and bool(shared) and all(x['correct'] for x in shared),{'binding':'shared','dates':shared}

def source_facts(candidate,bands,units):
    for b in sorted(bands):
        rows=[r for r in candidate['rates'] if (r['service'],dec(r['upper_bound']),r['weight_unit'])==b]
        traces=[rate_source_trace(r,candidate) for r in rows]
        add(units,'R006','source_trace '+str(b),bool(rows) and all(ok for ok,_ in traces),[e for _,e in traces],'Traceable Retail service/weight table; page number or equivalent locator')
    add(units,'R006','request_source_trace','quote_requests.csv' in candidate['text'] or ('project' in candidate['text'].lower() and 'request' in candidate['text'].lower()))
    add(units,'R006','pdf_identity_trace','Notice123' in candidate['text'] or 'Notice 123' in candidate['text'])

def rkey(r):return (service(r.get('service')),dec(r.get('upper_bound')),unit(r.get('weight_unit')),dec(r.get('zone')))
def qgroup(parsed):
    d=defaultdict(list)
    for q in parsed['quotes']:d[str(q['request_id'])].append(q)
    return d
def add(units,cid,label,ok,actual=None,expected=None):
    units[cid].append({'fact':label,'correct':bool(ok),'actual':actual,'expected':expected})
def fraction(units):return {k:str(D(sum(u['correct'] for u in us))/D(len(us))) for k,us in units.items()}

def quote_facts(parsed,expected,units,cid,prefix=''):
    qs=qgroup(parsed)
    for e in expected['quotes']:
        rows=qs[e['request_id']];q=rows[0] if len(rows)==1 else {}
        for field,key in [('priority_usd','priority'),('ground_usd','ground'),('selected','selected'),('selected_usd','selected_usd')]:
            ok=(q.get(field)==e[key]) if field=='selected' else same(q.get(field),e[key])
            add(units,cid,prefix+e['request_id']+'.'+field,ok,q.get(field),e[key])
    ts=parsed['totals']
    add(units,cid,prefix+'batch_total',bool(ts) and all(same(t['value'],expected['batch_total']) for t in ts),ts,expected['batch_total'])

def response_facts(before,after,expected_before,expected_after,units,prefix,changed_ids):
    """Delta response uses common USD meaning; baseline source accuracy stays R003."""
    bq,aq=qgroup(before),qgroup(after);eb={q['request_id']:q for q in expected_before['quotes']}
    for e in expected_after['quotes']:
        if e['request_id'] not in changed_ids:continue
        i=e['request_id'];b=bq[i][0] if len(bq[i])==1 else {};a=aq[i][0] if len(aq[i])==1 else {}
        for field,key in [('priority_usd','priority'),('ground_usd','ground'),('selected_usd','selected_usd')]:
            delta=dec(a.get(field))-dec(b.get(field)) if dec(a.get(field)) is not None and dec(b.get(field)) is not None else None
            target=D(e[key])-D(eb[i][key])
            # Active response and required invariance measure distinct abilities.
            # The Oracle fixes this partition before looking at any candidate.
            cid='R004' if target!=0 else 'R005'
            add(units,cid,prefix+i+'.'+field+'.delta',same(delta,target),str(delta),str(target))
        changed=e['selected']!=eb[i]['selected']
        ok=a.get('selected')==e['selected'] if changed else a.get('selected')==b.get('selected') and bool(a.get('selected'))
        add(units,'R004' if changed else 'R005',prefix+i+'.selected.response',ok,{'before':b.get('selected'),'after':a.get('selected')},{'changed':changed,'after':e['selected']})
    bt,at=before['totals'],after['totals'];target=D(expected_after['batch_total'])-D(expected_before['batch_total'])
    # Every displayed final total is paired by its candidate location, never by oracle value.
    bm={t['cell']:t['value'] for t in bt};am={t['cell']:t['value'] for t in at}
    deltas={cell:(str(dec(am[cell])-dec(value)) if dec(am.get(cell)) is not None and dec(value) is not None else None) for cell,value in bm.items()}
    ok=bool(bm) and set(bm)==set(am) and all(same(d,target) for d in deltas.values())
    add(units,'R004' if target!=0 else 'R005',prefix+'batch_total.delta',ok,deltas,str(target))

def patch_xlsx(source,target,changes):
    """Test-only isolated XML value mutations; leave every other member intact."""
    ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'};rid='{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    with zipfile.ZipFile(source) as zi:
        wb=ET.fromstring(zi.read('xl/workbook.xml'));rels=ET.fromstring(zi.read('xl/_rels/workbook.xml.rels'))
        relmap={r.attrib['Id']:r.attrib['Target'] for r in rels};paths={}
        for sheet in wb.find('m:sheets',ns):
            t=relmap[sheet.attrib[rid]];paths[sheet.attrib['name']]=t.lstrip('/') if t.startswith('/') else 'xl/'+t
        grouped=defaultdict(dict)
        for sheet,cell,value in changes:grouped[paths[sheet]][cell]=value
        with zipfile.ZipFile(target,'w',zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                b=zi.read(item.filename)
                if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                    tree=ET.fromstring(b);found=set()
                    for c in tree.findall('.//m:c',ns):
                        coord=c.attrib['r']
                        if c.find('m:f',ns) is not None:
                            for cache in c.findall('m:v',ns):c.remove(cache)
                        if coord not in grouped.get(item.filename,{}):continue
                        found.add(coord);c.attrib.pop('t',None)
                        for child in list(c):
                            if child.tag.rsplit('}',1)[-1] in ('f','v','is'):c.remove(child)
                        ET.SubElement(c,'{'+ns['m']+'}v').text=str(grouped[item.filename][coord])
                    if found!=set(grouped.get(item.filename,{})):raise ParsePending('Input control not a simple persisted cell')
                    b=ET.tostring(tree,encoding='utf-8',xml_declaration=True)
                if item.filename=='xl/workbook.xml':
                    tree=ET.fromstring(b);cp=tree.find('m:calcPr',ns)
                    if cp is None:cp=ET.SubElement(tree,'{'+ns['m']+'}calcPr')
                    cp.attrib.update(calcMode='auto',fullCalcOnLoad='1',forceFullCalc='1',calcId='0')
                    b=ET.tostring(tree,encoding='utf-8',xml_declaration=True)
                zo.writestr(item,b)

SCENARIOS=[('four_ounce_boundary',{'Q01':{'weight':'4.1'}}),('eight_ounce_boundary',{'Q03':{'weight':'8.1'}}),('pound_boundary',{'Q08':{'weight':'1.01'}}),('weight_zone_joint',{'Q01':{'weight':'4.1','zone':'8'}})]

def evaluate(answer,out,completed_run=False,dynamic=True,reuse_native=False):
    answer=Path(answer);out=Path(out);out.mkdir(parents=True,exist_ok=True);evidence={};state=output_status(answer)
    if state:
        if state=='OUTPUT_MISSING' and not completed_run:state='INFRA_ERROR'
        result=score_profiles(ROOT/'rubric.json',status=state,evidence={'completed_run':completed_run})
        (out/'result.json').write_text(json.dumps(result,indent=2));return result
    try:
        previous=json.loads((out/'result.json').read_text()) if reuse_native and (out/'result.json').exists() else {}
        previous_evidence=previous.get('evidence',{})
        if previous_evidence and previous_evidence.get('baseline_recalc',{}).get('source_sha256_before')!=sha256(answer):previous_evidence={}
        formulas=load_workbook(answer,data_only=False)
        unsupported=[f'{s.title}!{c.coordinate}' for s in formulas for row in s for c in row if c.data_type=='f' and re.search(r'(?:_xlfn\.)?PY\(',c.value,re.I)]
        if unsupported:raise ParsePending('Python in Excel formulas require native environment: '+', '.join(unsupported))
        if previous_evidence.get('baseline_recalc'):
            receipt=previous_evidence['baseline_recalc'];baseline_path=Path(receipt['output']);evidence['native_outputs_reused']=True
        else:baseline_path,receipt=recalculate_xlsx(answer,out/'baseline')
        evidence['baseline_recalc']=receipt
        require_native_results(baseline_path)
        candidate=parse(baseline_path);evidence['discovered_tables']=candidate['tables'];evidence['rate_delivery']=candidate['rate_delivery']
        rates=json.loads((ROOT/'metadata/rates.json').read_text());requests=list(csv.DictReader((ROOT/'data/input_files/quote_requests.csv').open()));expected=compute(rates,requests)
        units={f'R{i:03}':[] for i in range(1,7)};group=defaultdict(list)
        for r in candidate['rates']:group[rkey(r)].append(r)
        for r in rates:
            matches=group[rkey(r)]
            add(units,'R001',str(rkey(r)),len(matches)==1 and same(matches[0]['usd'],r['usd']),[x['usd'] for x in matches],r['usd'])
        bands={(r['service'],dec(r['upper_bound']),r['weight_unit']) for r in rates}
        candidate_bands=Counter((r['service'],dec(r['upper_bound']),r['weight_unit']) for r in candidate['rates'])
        for b in sorted(bands):add(units,'R002','band_identity '+str(b),candidate_bands[b]==8,candidate_bands[b],8)
        date_ok,date_evidence=effective_date_fact(candidate)
        add(units,'R002','effective_date',date_ok,date_evidence)
        quote_facts(candidate,expected,units,'R003')
        qs=qgroup(candidate)
        for req in requests:
            rows=qs[req['request_id']];q=rows[0] if len(rows)==1 else {}
            add(units,'R005','request_scope '+req['request_id'],same(q.get('weight'),req['weight']) and same(q.get('zone'),req['zone']) and q.get('weight_unit')==req['weight_unit'],{f:q.get(f) for f in ['weight','weight_unit','zone']},req)
        add(units,'R005','extra_request_identity',set(qs)=={r['request_id'] for r in requests},sorted(qs))
        source_facts(candidate,bands,units)
        if dynamic:
            before_rates=Counter((rkey(r),str(r['usd']),str(r.get('effective_date')),str(r.get('source_page'))) for r in candidate['rates'])
            evidence['dynamic']=[]
            for label,edits in SCENARIOS:
                changed_requests=[dict(r,**edits.get(r['request_id'],{})) for r in requests];changes=[]
                for req_id,fields in edits.items():
                    if len(qs[req_id])!=1:continue
                    q=qs[req_id][0]
                    for field,value in fields.items():changes.append((*q['_cells'][field],value))
                case=out/label;case.mkdir(exist_ok=True);mutated=case/'mutated.xlsx'
                cached=next((d for d in previous_evidence.get('dynamic',[]) if d['name']==label and d['declared_changes']==edits),None)
                if cached:rr=cached['recalc'];recalculated=Path(rr['output'])
                else:
                    patch_xlsx(baseline_path,mutated,changes)
                    recalculated,rr=recalculate_xlsx(mutated,case/'recalculated')
                require_native_results(recalculated);after=parse(recalculated)
                ex=compute(rates,changed_requests);response_facts(candidate,after,expected,ex,units,label+'.',set(edits))
                after_rates=Counter((rkey(r),str(r['usd']),str(r.get('effective_date')),str(r.get('source_page'))) for r in after['rates'])
                # Source preservation and unaffected outputs are separately recorded evidence.
                aq=qgroup(after);unaffected=[]
                for req in requests:
                    if req['request_id'] in edits:continue
                    i=req['request_id'];bq=qs[i][0] if len(qs[i])==1 else {};zq=aq[i][0] if len(aq[i])==1 else {}
                    unaffected.append({'request_id':i,'same':all(bq.get(f)==zq.get(f) for f in ['weight','weight_unit','zone','priority_usd','ground_usd','selected','selected_usd'])})
                # Retention is required, but cannot earn active-response credit.
                actual_inputs_ok=all(len(aq[i])==1 and all(same(aq[i][0].get(field),v) for field,v in fields.items()) for i,fields in edits.items())
                add(units,'R005',label+'.source_and_unaffected_preserved',before_rates==after_rates and all(x['same'] for x in unaffected) and actual_inputs_ok)
                evidence['dynamic'].append({'name':label,'declared_changes':edits,'actual_control_cells':changes,'changed_inputs_read_back_correct':actual_inputs_ok,'recalc':rr,'rate_facts_unchanged':before_rates==after_rates,'unaffected_quotes':unaffected,'oracle_total':ex['batch_total']})
        else:raise ParsePending('Dynamic verification explicitly skipped; final score unavailable')
        evidence['facts']=units;evidence['denominators']={k:len(v) for k,v in units.items()}
        evidence['candidate_self_consistency']={'quotes':[{ 'request_id':q['request_id'],'selected_price_matches_own_services':same(q['selected_usd'], min(x for x in [dec(q['priority_usd']),dec(q['ground_usd'])] if x is not None)) if dec(q['priority_usd']) is not None and dec(q['ground_usd']) is not None else False} for q in candidate['quotes']], 'all_batch_totals_match_own_detail':all(same(t['value'],sum((dec(q['selected_usd']) or D(0) for q in candidate['quotes']),D(0))) for t in candidate['totals']) if candidate['totals'] else False}
        result=score_profiles(ROOT/'rubric.json',fraction(units),evidence=evidence)
    except (ParsePending,RecalcUnavailable) as exc:
        evidence['pending_reason']=str(exc)
        if getattr(exc,'details',None):evidence['parse_evidence']=exc.details
        result=score_profiles(ROOT/'rubric.json',status='JUDGE_ERROR',evidence=evidence)
    except Exception as exc:
        evidence['judge_exception']=type(exc).__name__+': '+str(exc);result=score_profiles(ROOT/'rubric.json',status='JUDGE_ERROR',evidence=evidence)
    (out/'result.json').write_text(json.dumps(result,indent=2,ensure_ascii=False,default=str))
    return result

if __name__=='__main__':
    p=argparse.ArgumentParser();p.add_argument('answer',nargs='?',default='/app/output/answer.xlsx');p.add_argument('--out',default='/logs/verifier');p.add_argument('--completed-run',action='store_true');a=p.parse_args()
    result=evaluate(a.answer,a.out,a.completed_run);print(json.dumps({k:v for k,v in result.items() if k!='evidence'},indent=2))
    if result['evaluation_status']=='JUDGE_ERROR':sys.exit(2)
