"""Focused partial-rate parsing regression; one native baseline, no Agent call."""
import json
from decimal import Decimal as D
from pathlib import Path
from openpyxl import load_workbook
from evaluate import ROOT,parse,ParsePending,evaluate

def main():
    out=ROOT/'metadata/partial_parse_repair';out.mkdir(exist_ok=True)
    wide=ROOT/'fixtures/wide_rates_long_quotes.xlsx'
    result=evaluate(wide,out/'wide',completed_run=True)
    assert result['evaluation_status']=='JUDGE_ERROR'
    assert result['normalized_score'] is None and result['pass'] is None
    assert all(p['normalized_score'] is None and p['pass'] is None for p in result['profiles'].values())
    evidence=result['evidence']['parse_evidence']
    assert evidence['parsed_quote_rows']==12 and evidence['parsed_rate_rows']==0
    assert len(evidence['unbound_rate_grids'])==2 and not evidence['oracle_values_used_for_discovery']
    # Verify this fixture contains a real correct alternative rate representation,
    # rather than naming a malformed file "equivalent". These fixture-authored
    # addresses validate the test only and are never used by candidate discovery.
    fresh=Path(result['evidence']['baseline_recalc']['output']);w=load_workbook(fresh,data_only=True)
    oracle=json.loads((ROOT/'metadata/oracle_expected.json').read_text())
    for idx,expected in enumerate(oracle['quotes'],6):
        assert w['Quotes'].cell(idx,1).value==expected['request_id']
        for col,key in [(9,'priority'),(10,'ground'),(12,'selected_usd')]:assert D(str(w['Quotes'].cell(idx,col).value))==D(expected[key])
        assert w['Quotes'].cell(idx,11).value==expected['selected']
    assert D(str(w['Quotes']['B20'].value))==D(oracle['batch_total'])
    rates=json.loads((ROOT/'metadata/rates.json').read_text());price_count=0
    for service,header,start,end in [('priority',5,6,15),('ground',19,20,33)]:
        assert [w['Rates'].cell(header,c).value for c in range(3,11)]==[f'Zone {i}' for i in range(1,9)]
        for row in range(start,end+1):
            bound,unit=w['Rates'].cell(row,1).value,w['Rates'].cell(row,2).value
            for zone in range(1,9):
                matches=[r for r in rates if r['service']==service and D(r['upper_bound'])==D(str(bound)) and r['weight_unit']==unit and r['zone']==zone]
                assert len(matches)==1
                assert D(str(w['Rates'].cell(row,zone+2).value))==D(matches[0]['usd']);price_count+=1
    assert price_count==192
    empty=parse(ROOT/'fixtures/empty_rate_table.xlsx')
    assert len(empty['quotes'])==12 and not empty['rates']
    assert empty['rate_delivery']['state']=='CONFIRMED_OMITTED'
    try:parse(ROOT/'fixtures/unresolved_rate_delivery.xlsx')
    except ParsePending as e:assert e.details['rate_delivery']=='UNRESOLVED'
    else:raise AssertionError('Ambiguous missing rate binding must not become zero business credit')
    original=parse(ROOT/'solution/reference.xlsx')
    assert len(original['rates'])==192 and len(original['quotes'])==12 and original['rate_delivery']['state']=='BOUND'
    receipt={'judge_version':'new6-usps-facts-v1.1-partial-parse','passed':True,'category':'partial semantic parsing','native_baseline_runs':1,'dynamic_recalc_runs':0,'agent_calls':0,'paid_cost_usd':'0','assertions':{'legitimate_wide_grids_plus_long_quotes':'JUDGE_ERROR/null for all profiles','wide_fixture_source_prices_verified':192,'wide_fixture_quote_outputs_verified':49,'unbound_grid_detection':'Public field labels and row shape only; no Oracle value used to find candidate columns','clearly_empty_labelled_rate_table':'CONFIRMED_OMITTED with parsed rates empty, available for business omission grading','uncertain_rate_delivery':'JUDGE_ERROR/UNRESOLVED, not numeric zero','unchanged_long_reference':'192 bound rates and12 quotes remain readable'},'scope_preserved':['Frozen first-attempt wrapper unchanged','Agent-visible inputs unchanged','Original first-attempt result unchanged','No Agent retry or execution of failed Agent text'],'known_limit':'This repair detects partial parsing and returns pending; it does not implement a wide-grid semantic adapter.'}
    (out/'receipt.json').write_text(json.dumps(receipt,indent=2));print(json.dumps(receipt,indent=2))
if __name__=='__main__':main()
