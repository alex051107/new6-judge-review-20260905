"""Read retained input facts as context for a separately delivered workbook.

No source-derived output schedules are inserted; candidate output evidence stays
in its own labelled tables. The in-memory context is never saved to the answer.
"""
import re
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
INPUT_DIR=ROOT/'data/input_files'
SOURCE_CACHE={}
norm=lambda value:re.sub(r'[^a-z0-9]+','',str(value or '').lower())


def bind_controls(w, native):
    result=native(w)
    names={'scenariotargetinvestmentgdp':'target','transitioncompletebyyear':'transition'}
    for sh in w:
        for row in sh.iter_rows(max_row=min(sh.max_row,100),max_col=min(sh.max_column,15)):
            for c in row:
                key=names.get(norm(c.value))
                if not key:continue
                values=[x for x in row[c.column:c.column+3] if isinstance(x.value,(int,float))]
                if len(values)!=1:raise ValueError('Labelled scenario control is not unique')
                binding=(sh.title,values[0].coordinate)
                if key in result and result[key]!=binding:raise ValueError('Several scenario controls have the same meaning')
                result[key]=binding
    return result


def attach_retained_source(w):
    required=['InputDataA_GeneralAssumptions','data','Readme']
    if all(sn in w for sn in required):return
    # A partly preserved or renamed original still needs explicit semantic review.
    if any(sn in w for sn in required):raise ValueError('Partly retained source layout needs a supported binding')
    source=INPUT_DIR/'LTGM_Zambia_restore.xlsx'
    if not source.exists():raise ValueError('Actual retained input was not collected')
    if str(source) not in SOURCE_CACHE:
        from runtime import recalculate_xlsx
        calculated,receipt=recalculate_xlsx(source,ROOT/'metadata/retained_input_recalc')
        SOURCE_CACHE[str(source)]=(load_workbook(source,data_only=False),load_workbook(calculated,data_only=True),receipt)
    src,computed,receipt=SOURCE_CACHE[str(source)]
    if not all(sn in src for sn in required):raise ValueError('Retained source layout cannot be safely bound')
    overrides={}
    names={'depreciationrate':['D13','E13'],'laborshare':['D16','E16'],'initialcapitaloutputratio':['I13','J13']}
    for sh in w:
        for row in sh.iter_rows(max_row=min(sh.max_row,100),max_col=min(sh.max_column,15)):
            for c in row:
                keys=names.get(norm(c.value))
                if not keys:continue
                values=[x.value for x in row[c.column:c.column+2] if isinstance(x.value,(int,float))]
                if len(values)!=1:raise ValueError('Candidate source assumption has ambiguous value')
                for key in keys:
                    if key in overrides and overrides[key]!=values[0]:raise ValueError('Contradictory candidate source assumptions')
                    overrides[key]=values[0]
    for sn in required:
        target=w.create_sheet(sn)
        section=computed[sn] if sn=='InputDataA_GeneralAssumptions' else src[sn]
        for row in section:
            for cell in row:
                if cell.value is not None:target.cell(cell.row,cell.column,cell.value)
    for cell,value in overrides.items():w['InputDataA_GeneralAssumptions'][cell]=value
    w._new6_retained_source_context=str(source)
    w._new6_retained_source_recalc=receipt
