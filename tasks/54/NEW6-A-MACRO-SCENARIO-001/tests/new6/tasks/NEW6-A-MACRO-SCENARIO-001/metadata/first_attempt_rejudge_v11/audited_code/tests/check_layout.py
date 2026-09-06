import json
from pathlib import Path
from openpyxl import Workbook,load_workbook
from openpyxl.chart import LineChart,Reference
import evaluate as e

out=e.ROOT/'metadata/layout_validation';out.mkdir(exist_ok=True)
source=load_workbook(e.ROOT/'solution/reference.xlsx',data_only=True)
facts,_=e.discover(source)
labels={'investment_share':'Investment share','capital_output_ratio':'K/Y','gdp':'GDP','gdp_per_capita':'GDP per capita','gdp_growth':'GDP growth','pc_growth':'GDP pc Growth','investment':'Annual investment','capital':'Capital stock'}
wb=Workbook();p=wb.active;p.title='Parameters'
p.append(['Scenario Target Investment/GDP:',.24]);p.append(['Transition Complete by Year:',2025])
p.append(['Money units: 2010 US dollars']);p.append(['The lower investment scenario reduces GDP growth relative to the baseline.'])
s=wb.create_sheet('Results');pairs=[(case,key) for case in ('baseline','scenario') for key in labels]
s.append(['Year']+[case.title()+' '+labels[key] for case,key in pairs])
for year in e.YEARS:s.append([year]+[facts[case,year,key][0]['value'] for case,key in pairs])
chart=LineChart();chart.title='GDP per Capita Growth Rate: comparison'
for case in ('baseline','scenario'):
    col=pairs.index((case,'pc_growth'))+2
    chart.add_data(Reference(s,min_col=col,min_row=1,max_row=18),titles_from_data=True)
chart.set_categories(Reference(s,min_col=1,min_row=2,max_row=18));s.add_chart(chart,'T2')
path=out/'fresh_layout.xlsx';wb.save(path)
new=load_workbook(path,data_only=True);details,_,_=e.checks_for(new)
assert all(x['ok'] for cid,items in details.items() if cid!='R004' for x in items),[(cid,[x['id'] for x in items if not x['ok']]) for cid,items in details.items()]
assert e.controls(new)=={'target':('Parameters','B1'),'transition':('Parameters','B2')}
wrong=load_workbook(path,data_only=True);wrong['Parameters'].append(['Depreciation rate:',.09])
bad,_,_=e.checks_for(wrong)
assert {x['id'] for x in bad['R001'] if not x['ok']}=={'source_control:D13','source_control:E13'}
assert all(x['ok'] for cid in ('R002','R003','R005','R006') for x in bad[cid])
missing=load_workbook(path,data_only=True);missing['Results'].delete_cols(2)
loss,_,_=e.checks_for(missing)
assert sum(not x['ok'] for x in loss['R002'])==17
assert all(x['ok'] for x in loss['R003'])
receipt={'passed':True,'checks':['fresh labelled workbook retains all current reference facts through actual input context','candidate-defined scenario controls located without fixed addresses','wrong disclosed depreciation loses exactly two source facts; other categories retained','omitted annual metric loses its17 facts; remaining outputs retained'],'native_recalc_runs':0,'source_values_used_for_candidate_output_discovery':False,'no_output_schedules_injected':True,'weights':'unchanged frozen v1'}
(out/'receipt.json').write_text(json.dumps(receipt,indent=2));print(json.dumps(receipt))
