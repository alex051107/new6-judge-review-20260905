"""NEW6 LTGM fact evaluator. Labels/years bind candidate quantities; no gold search.
Source-native baseline and tables with named case/metric columns are supported.
Unrecognized material layouts or legal unsupported formulas return JUDGE_ERROR.
"""
import argparse,json,sys,re,math
from collections import defaultdict,Counter
from pathlib import Path
from decimal import Decimal
from functools import lru_cache
import openpyxl
ROOT=Path(__file__).resolve().parents[1]
sys.path[:0]=[str(ROOT/'metadata'),str(ROOT.parents[1]/'common')]
from oracle_recompute import compute
from ooxml_edit import edit
from runtime import output_status,recalculate_xlsx,score_profiles,sha256,RecalcUnavailable
norm=lambda v:re.sub(r'[^a-z0-9]+','',str(v).casefold())
YEARS=range(2019,2036)
ALIASES={
 'investment_share':['Investment to GDP Ratio','investment / GDP','investment share','I/Y'],
 'capital_output_ratio':['Capital / Output Ratio','capital / output','K/Y'],
 'gdp':['Real Total GDP','GDP','output'],
 'gdp_per_capita':['Real GDP per capita','GDP per capita','per capita GDP'],
 'gdp_growth':['Real Total GDP growth rate','GDP growth','GDP growth rate'],
 'pc_growth':['GDP per Capita Growth rate','GDP pc Growth','GDP per capita growth','per capita GDP growth'],
 'investment':['Investment','annual investment'],
 'capital':['Capital','capital stock']}
ALIASES={k:{norm(v) for v in vals} for k,vals in ALIASES.items()}

def equal(a,b):
 try:return math.isfinite(float(a)) and abs(float(a)-float(b))<=1e-7*max(1,abs(float(b)))
 except (ValueError,TypeError):return False

def source_equal(a,b):
 # Excel and LibreOffice normalize an explicit empty string literal to a blank cell.
 if a in (None,'') and b in (None,''):return True
 return equal(a,b) if isinstance(b,(int,float)) else a==b

def required_engine_errors(facts):
 errors=[dict(r,quantity=str(k)) for k,rs in facts.items() for r in rs if r['value'] in ('#NAME?','#VALUE!','#N/A')]
 if errors:raise RecalcUnavailable('Required bound quantity returned an engine/parse error; manual compatibility adjudication needed: '+str(errors[:3]))

def metric(value):
 n=norm(value)
 for k,names in ALIASES.items():
  if n in names:return k
 return None

def case_metric(value):
 n=norm(value)
 for case,tokens in [('baseline',['baseline','basecase','base']),('scenario',['scenario','investmenttransition','review'])]:
  for token in tokens:
   if n.startswith(token):return case,metric(n[len(token):])
 return None,None

@lru_cache(maxsize=1)
def original_sheet_names():
 wb=openpyxl.load_workbook(ROOT/'metadata/configured_source.xlsx',read_only=True);names=frozenset(wb.sheetnames);wb.close();return names

def discover(w):
 facts=defaultdict(list);tables=[]
 def add(case,year,key,s,c):facts[(case,year,key)].append({'value':c.value,'sheet':s.title,'cell':c.coordinate})
 # Source-preserving worksheets: recognize the source's named baseline/scenario model,
 # then discover every row and year by its label; source addresses are never answer locators.
 for s in w:
  native_case='baseline' if norm(s.title) in ['submodel1','model1'] else 'scenario' if norm(s.title) in ['submodel1s','model1s'] else None
  if native_case:
   yearrows=[row for row in s.iter_rows(max_row=min(15,s.max_row),max_col=min(100,s.max_column)) if any(norm(c.value)=='year' for c in row)]
   if not yearrows:continue
   yearcols={c.column:int(c.value) for c in yearrows[0] if isinstance(c.value,(float,int)) and int(c.value) in YEARS}
   seen=[]
   for row in s.iter_rows(max_col=min(100,s.max_column)):
    labels=[(c,metric(c.value)) for c in row[:4] if metric(c.value)]
    for label,key in labels:
     for col,year in yearcols.items():add(native_case,year,key,s,s.cell(label.row,col));seen.append((year,key))
   tables.append({'kind':'native','sheet':s.title,'case':native_case,'seen':seen})
  # Ordinary output tables: Year + named baseline/scenario metrics, any position/order.
  if s.max_row>10000 or s.max_column>1000:raise ValueError('Workbook region exceeds parser bound')
  if s.title in ['data','DataSummary']:continue
  for row in s.iter_rows(max_row=min(s.max_row,200),max_col=min(s.max_column,100)):
   yearcells=[c for c in row if norm(c.value) in ['year','calendaryear']]
   named=[(c,*case_metric(c.value)) for c in row]
   named=[(c,case,key) for c,case,key in named if case and key]
   if not yearcells or len(named)<2:continue
   yc=yearcells[0];seen=[]
   for rr in range(yc.row+1,min(s.max_row,yc.row+150)+1):
    y=s.cell(rr,yc.column).value
    if not isinstance(y,(int,float)) or int(y) not in YEARS:continue
    for c,case,key in named:add(case,int(y),key,s,s.cell(rr,c.column));seen.append((case,int(y),key))
   tables.append({'kind':'report','sheet':s.title,'header_row':yc.row,'columns':[(case,key) for _,case,key in named],'seen':seen})
 # A material output table whose headers cannot be bound is a parser limitation,
 # even while some preserved native outputs remain readable. Empty/absent outputs score missing facts.
 missing=any(not facts.get((case,y,k)) for case in ('baseline','scenario') for y in YEARS for k in ALIASES)
 if missing:
  parsed_sheets={t['sheet'] for t in tables if t['kind']=='report'}
  for sh in w:
   if sh.title in original_sheet_names() or sh.title in parsed_sheets:continue
   vals=[c.value for row in sh.iter_rows(max_row=min(sh.max_row,200),max_col=min(sh.max_column,100)) for c in row]
   year_count=sum(isinstance(x,(int,float)) and int(x) in YEARS for x in vals)
   numeric_count=sum(isinstance(x,(int,float)) for x in vals)
   if year_count>=5 and numeric_count>=20:raise ValueError('Material candidate result table exists but cannot be safely bound: '+sh.title)
 required_engine_errors(facts)
 return facts,tables

def unit(facts,key,expected):
 rows=facts.get(key,[])
 return bool(rows) and all(equal(r['value'],expected) for r in rows)

def controls(w):
 result={}
 for s in w:
  if s.title=='data':continue
  for row in s.iter_rows(max_row=min(s.max_row,100),max_col=min(s.max_column,15)):
   for c in row:
    token=norm(c.value)
    if token not in ['targetinvestmentratioiy','reachtargetiyratebyyear']:continue
    headers=[(r,col) for r in range(max(1,c.row-10),c.row) for col in range(c.column+1,min(s.max_column,15)+1) if norm(s.cell(r,col).value)=='scenario' and norm(s.cell(r,col-1).value)=='baseline']
    if len(headers)!=1:continue
    _,col=headers[0];target=s.cell(c.row,col)
    result['target' if token.startswith('target') else 'transition']=(s.title,target.coordinate)
 return result

def chart_facts(w):
 def values(formula):
  try:
   sh,rng=formula.rsplit('!',1);sh=sh.strip("'").replace("''", "'");bounds=openpyxl.utils.range_boundaries(rng)
   return [c.value for row in w[sh].iter_rows(min_col=bounds[0],min_row=bounds[1],max_col=bounds[2],max_row=bounds[3]) for c in row]
  except Exception:return None
 candidates=[]
 for s in w:
  for chart_index,chart in enumerate(s._charts):
   series=[]
   for ser in chart.series:
    val=getattr(ser,'val',None);cat=getattr(ser,'cat',None)
    vr=getattr(val,'numRef',None) if val else None;cr=(getattr(cat,'numRef',None) or getattr(cat,'strRef',None)) if cat else None
    if vr is None or cr is None:continue
    vs=values(vr.f);ys=values(cr.f)
    if ys and all(isinstance(x,(int,float)) for x in ys) and set(int(x) for x in ys).intersection(YEARS):series.append({'years':ys,'values':vs,'value_range':vr.f,'category_range':cr.f,'label':(values(ser.tx.strRef.f) or [None])[0] if ser.tx and ser.tx.strRef else ser.tx.v if ser.tx else None})
   # A dedicated 2019–2035 comparison must identify both cases; unrelated source charts are preserved.
   if len(series)==2 and any(len(x['years'])<=17 for x in series):
    title=' '.join(r.t for para in chart.title.tx.rich.p for r in para.r) if chart.title and chart.title.tx and chart.title.tx.rich else ''
    declared_metric=metric(title.split(':')[0])
    for z in series:
     case,key=case_metric(z['label'])
     if case and not key:key=declared_metric
     if not case:
      n=norm(z['label']);case='baseline' if n in ('baseline','base','basecase') else 'scenario' if n in ('scenario','investmenttransition') else None
      key=declared_metric
     if case is None or key not in ALIASES:raise ValueError('Comparison chart series case/metric cannot be bound from visible label/title: '+str(z['label']))
     z.update(case=case,metric=key)
    candidates.append({'sheet':s.title,'chart_index':chart_index,'title':title,'series':series})
 return candidates

@lru_cache(maxsize=1)
def prefilled_source_prose():
 # Only a text-origin inventory: no cached numerical values enter the Oracle or scores.
 src=openpyxl.load_workbook(ROOT/'metadata/source_recalculated/configured_source.xlsx',data_only=True)
 return Counter(norm(c.value) for sh in src for row in sh.iter_rows(max_row=min(sh.max_row,200),max_col=min(sh.max_column,50)) for c in row if isinstance(c.value,str))

def interpretation_evidence(w,ref):
 # Only new prose can satisfy the requested interpretation; prefilled source notes
 # and units/headings do not become a delivered comparison explanation.
 supplied=prefilled_source_prose().copy()
 candidates=[];unknown=[]
 for sh in w:
  if sh.title=='data':continue
  for row in sh.iter_rows(max_row=min(sh.max_row,200),max_col=min(sh.max_column,50)):
   for c in row:
    if not isinstance(c.value,str):continue
    text=c.value.strip();token=norm(text)
    if supplied[token]>0:supplied[token]-=1;continue
    if len(text)<25 or re.match(r'^(money|units?|source|note on units)\s*:',text,re.I):continue
    if len(text.split())<6 and not re.search('[\u4e00-\u9fff]',text):continue
    # A header without a predicate is not a prose explanation.
    if not re.search(r'[.!。]|\b(is|are|has|have|grows|grow|produces?|reduces?|decreases?|increases?|slows?|reflects?|compared|relative)\b',text,re.I):continue
    relation=bool(re.search(r'baseline|scenario|investment|基准|情景|投资',text,re.I))
    result=bool(re.search(r'growth|grows|output|gdp|capital|增长|产出|资本',text,re.I))
    comparison=bool(re.search(r'lower|higher|less|more|reduc|decreas|increas|slow|fast|differ|compar|relative|gap|same|unchanged|低|高|降|增|差|相同|不变',text,re.I))
    record={'sheet':sh.title,'cell':c.coordinate,'text':text}
    if relation and result and comparison:candidates.append(record)
    else:unknown.append(record)
 if candidates:return True,candidates
 if unknown:raise ValueError('Candidate prose is present but this bounded interpretation check cannot reliably determine its comparison meaning: '+str(unknown[:2]))
 return False,[]

def require_bound_source_layout(w,ref):
 # Source-native coordinates are used only after their visible labels and record
 # identities confirm the layout. A legal relocation needs an adapter extension.
 for sn in ['InputDataA_GeneralAssumptions','data','Readme']:
  if sn not in w:raise ValueError('Preserved source section may have been relocated/renamed; source adapter cannot bind: '+sn)
 for co in ['C3','B7','B11','B13','B14','B16','B20','B22','B29','B31']:
  if norm(w['InputDataA_GeneralAssumptions'][co].value)!=norm(ref['InputDataA_GeneralAssumptions'][co].value):raise ValueError('Source assumption labels relocated or rewritten; do not score old coordinates: '+co)
 for co in ['A23','B23','C23','D23','E23']:
  if norm(w['data'][co].value)!=norm(ref['data'][co].value):raise ValueError('Protected source data columns/header relocated; adapter cannot bind original coordinates')
 def keys(sh):return [(r,sh.cell(r,3).value) for r in range(24,sh.max_row+1) if isinstance(sh.cell(r,3).value,str) and re.fullmatch('[A-Z]{3}',sh.cell(r,3).value)]
 old,new=keys(ref['data']),keys(w['data'])
 if Counter(v for r,v in old)==Counter(v for r,v in new) and old!=new:raise ValueError('Same source country records have been reordered; source protection requires key-based adapter extension, not a zero')

def checks_for(w,target='.24',transition=2025):
 attach_retained_source(w);ref=openpyxl.load_workbook(ROOT/'metadata/configured_source.xlsx',data_only=False);require_bound_source_layout(w,ref)
 facts,tables=discover(w);oracles={'baseline':compute(baseline=True),'scenario':compute(target,transition)};details={k:[] for k in ['R001','R002','R003','R004','R005','R006']}
 def add(cid,id,ok,actual=None,expected=None):details[cid].append({'id':id,'ok':bool(ok),'actual':actual,'expected':expected})
 for case,rows in oracles.items():
  for row in rows:
   for k in ALIASES:
    cid='R002' if k in ['investment_share','capital_output_ratio','capital','investment'] else 'R003';key=(case,row['year'],k)
    add(cid,f'{case}:{row["year"]}:{k}',unit(facts,key,row[k]),facts.get(key,[]),str(row[k]))
 # Source assumptions are located by documented source label/meaning, not candidate answer values.
 expected_controls=controls(w);add('R001','identified_scenario_controls',set(expected_controls)=={'target','transition'},expected_controls)
 for key,expected in [('target',float(target)),('transition',transition)]:
  sc=expected_controls.get(key);add('R001','initial_'+key,bool(sc) and equal(w[sc[0]][sc[1]].value,expected),w[sc[0]][sc[1]].value if sc else None,expected)
 for sn in ['InputDataA_GeneralAssumptions']:
  if sn not in w:raise ValueError('Cannot locate required preserved source assumption sheet')
  s=w[sn]
  # Original source control coordinates only bind immutable source facts, never submitted output schedules.
  for co,expected in {'B3':'Zambia','D7':2019,'D13':.049,'E13':.049,'D16':.618,'E16':.618,'I13':2.42,'J13':2.42,'D22':.006,'E22':.006,'D31':.01,'E31':.01}.items():
   v=s[co].value;add('R001','source_control:'+co,v==expected if isinstance(expected,str) else equal(v,expected),v,expected)
 text=' '.join(str(c.value) for s in w if s.title!='data' for row in s.iter_rows(max_row=min(s.max_row,200),max_col=min(s.max_column,50)) for c in row if isinstance(c.value,str))
 add('R001','money_units',bool(re.search(r'(2010).*(US|dollar)|US.*2010',text,re.I)))
 for case in ['baseline','scenario']:
  for key in ALIASES:
   declarations=[]
   for table in tables:
    if table['kind']=='report' and (case,key) in table['columns']:
     count=Counter(tuple(v) for v in table['seen']);declarations.append(all(count[(case,y,key)]==1 for y in YEARS))
    elif table['kind']=='native' and table['case']==case and any(k==key for y,k in table['seen']):
     count=Counter(tuple(v) for v in table['seen']);declarations.append(all(count[(y,key)]==1 for y in YEARS))
   add('R006',f'record_unique_complete:{case}:{key}',bool(declarations) and all(declarations),declarations)
 charts=chart_facts(w);chart_ok=False;chart_values_ok=False; chart_checks=[]
 for chart in charts:
  valid_chart=False;series=chart['series'];yearok=all([int(y) for y in z['years']]==list(YEARS) for z in series)
  if yearok:
   chart_ok=True
   if {z['case'] for z in series}=={'baseline','scenario'} and len({z['metric'] for z in series})==1:
    if all(z['values'] and len(z['values'])==17 and all(equal(v,r[z['metric']]) for v,r in zip(z['values'],oracles[z['case']])) for z in series):valid_chart=True
  chart_checks.append(valid_chart)
 chart_values_ok=bool(chart_checks) and all(chart_checks)
 add('R006','comparison_chart_all_years',chart_ok,charts);add('R006','comparison_chart_matches_series',chart_values_ok)
 interpretation_ok,interpretation=interpretation_evidence(w,ref);add('R006','comparison_interpretation_delivered',interpretation_ok,interpretation)
 # Input snapshot source records and the original baseline values remain a separate fact set.
 for sn in ['data','Readme']:
  if sn not in w:add('R005','source_sheet:'+sn,False);continue
  original=Counter((str(c.value),c.data_type) for row in ref[sn] for c in row if c.value is not None and c.data_type!='f')
  actual=Counter((str(c.value),c.data_type) for row in w[sn] for c in row if c.value is not None and c.data_type!='f')
  # Float presentation after native XML conversion may change string last bits; compare literal numeric payloads with tolerance by location.
  if sn=='data':
   ok=all(source_equal(w[sn][c.coordinate].value,c.value) for row in ref[sn] for c in row if c.value is not None and c.data_type!='f')
  else:ok=original==actual
  add('R005','protected_source_'+sn,ok)
 for row in oracles['baseline']:
  for k in ['investment_share','capital_output_ratio','gdp','gdp_per_capita','gdp_growth','pc_growth']:add('R005',f'baseline:{row["year"]}:{k}',unit(facts,('baseline',row['year'],k),row[k]))
 return details,facts,tables

def evaluate(path,evidence_dir,completed_run=True):
 evidence_dir=Path(evidence_dir);evidence_dir.mkdir(parents=True,exist_ok=True)
 status=output_status(path)
 if status:
  result=score_profiles(ROOT/'rubric.json',status=status if status!='OUTPUT_MISSING' or completed_run else 'INFRA_ERROR',evidence={'completed_run':completed_run})
  (evidence_dir/'evaluation.json').write_text(json.dumps(result,ensure_ascii=False,indent=2));return result
 try:
  raw=openpyxl.load_workbook(path,data_only=False); attach_retained_source(raw)
  unsupported=[f'{s.title}!{c.coordinate}: {c.value}' for s in raw for row in s for c in row if c.data_type=='f' and re.search(r'(_xlfn\.)?(LAMBDA|PY)\(',str(c.value),re.I)]
  if unsupported:raise RecalcUnavailable('Legal formula feature not supported by this recalculation adapter: '+unsupported[0])
  fresh,receipt=recalculate_xlsx(path,evidence_dir/'base');w=openpyxl.load_workbook(fresh,data_only=True);details,before,_=checks_for(w);ctrl=controls(w);probes=[]
  if set(ctrl)!={'target','transition'}:raise ValueError('Dynamic controls could not be bound uniquely by visible labels')
  for name,target,year in [('target_026','.26',2025),('transition_2027','.24',2027),('joint','.26',2027)]:
   mutated=evidence_dir/name/'mutated.xlsx';patch=defaultdict(dict)
   for key,value in [('target',float(target)),('transition',year)]:sn,co=ctrl[key];patch[sn][co]=value
   edit(path,mutated,patches=patch,clear_caches=True);afterpath,rec=recalculate_xlsx(mutated,evidence_dir/name/'recalc');after=openpyxl.load_workbook(afterpath,data_only=True);attach_retained_source(after);af,at=discover(after);o={'baseline':compute(baseline=True),'scenario':compute(target,year)}
   units=[]
   for case,rows in o.items():
    for row in rows:
     orig=compute(baseline=case=='baseline')[row['year']-2019]
     for k in ALIASES:
      key=(case,row['year'],k);prior=before.get(key,[]);current=af.get(key,[]);accurate=unit(af,key,row[k]);response=False
      if prior and current and len(prior)==len(current):response=all(equal(float(b['value'])-float(a['value']),row[k]-orig[k]) for a,b in zip(prior,current) if isinstance(a['value'],(int,float)) and isinstance(b['value'],(int,float))) and all(isinstance(a['value'],(int,float)) and isinstance(b['value'],(int,float)) for a,b in zip(prior,current))
      ok=response;units.append({'id':f'{name}:{case}:{row["year"]}:{k}','ok':ok,'absolute_accuracy':accurate,'delta_accuracy':response,'before':prior,'after':current,'expected_delta':str(row[k]-orig[k])})
   # Baseline outputs above have exact zero-delta assertions. All source literal controls except specified changes must also be invariant.
   history_ok=True
   for sn in ['data','InputDataA_GeneralAssumptions']:
    if sn not in after:history_ok=False;continue
    for row in raw[sn]:
     for c in row:
      if c.value is None or c.data_type=='f':continue
      a=after[sn][c.coordinate].value
      if not source_equal(a,c.value):history_ok=False
   units.append({'id':name+':unchanged_history_and_general_controls','ok':history_ok})
   chart_updated=False
   base_charts={(c['sheet'],c['chart_index']):c for c in chart_facts(w)}
   for chart in chart_facts(after):
    prior_chart=base_charts.get((chart['sheet'],chart['chart_index']))
    if not prior_chart:continue
    zs=chart['series'];ps=prior_chart['series']
    if len(zs)!=len(ps) or not all([int(y) for y in z['years']]==list(YEARS) for z in zs+ps):continue
    if {z['case'] for z in zs}!={'baseline','scenario'} or len({z['metric'] for z in zs})!=1:continue
    good=True
    for z,pz in zip(zs,ps):
     case,mk=z['case'],z['metric']
     if (case,mk)!=(pz['case'],pz['metric']):good=False;continue
     original_rows=compute(baseline=case=='baseline')
     if not z['values'] or not pz['values'] or len(z['values'])!=17 or len(pz['values'])!=17:good=False;continue
     for bv,av,er,old in zip(pz['values'],z['values'],o[case],original_rows):
      if not isinstance(bv,(int,float)) or not isinstance(av,(int,float)) or not equal(av-bv,er[mk]-old[mk]):good=False
    chart_updated=chart_updated or good
   units.append({'id':name+':chart_response_delta_and_declared_series','ok':chart_updated})
   details['R004']+=units;probes.append({'name':name,'changes':dict(patch),'receipt':rec})
  scores={cid:str(Decimal(sum(u['ok'] for u in us))/Decimal(len(us))) for cid,us in details.items()}
  evidence={'base_native_recalc':receipt,'dynamic_probes':probes,'fact_units':details,'fixed_scope':'2 cases x 17 years; per metric denominator fixed; report claims all checked; no result chosen by gold closeness','adapter_limit':'English source labels or English case/metric table headers; unrecognized material layout needs parser extension and JUDGE_ERROR, not capability failure'}
  result=score_profiles(ROOT/'rubric.json',scores,evidence=evidence)
 except (RecalcUnavailable,ValueError,KeyError,TypeError,openpyxl.utils.exceptions.InvalidFileException) as e:result=score_profiles(ROOT/'rubric.json',status='JUDGE_ERROR',evidence={'error_type':type(e).__name__,'error':str(e)})
 (evidence_dir/'evaluation.json').write_text(json.dumps(result,ensure_ascii=False,indent=2,default=str));return result

from layout_support import attach_retained_source, bind_controls
_original_controls=controls
def controls(w):return bind_controls(w,_original_controls)

if __name__=='__main__':
 p=argparse.ArgumentParser();p.add_argument('answer',nargs='?',default='/app/output/answer.xlsx');p.add_argument('--evidence-dir',default='/tmp/new6-a2-evidence');p.add_argument('--input-dir');p.add_argument('--completed-run',action='store_true');a=p.parse_args();import layout_support;layout_support.INPUT_DIR=Path(a.input_dir) if a.input_dir else ROOT/'data/input_files';r=evaluate(a.answer,a.evidence_dir,completed_run=a.completed_run);print(json.dumps(r,ensure_ascii=False,default=str))
