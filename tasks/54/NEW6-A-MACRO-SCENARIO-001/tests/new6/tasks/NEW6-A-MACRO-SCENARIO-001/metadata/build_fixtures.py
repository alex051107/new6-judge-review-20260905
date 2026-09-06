"""Reproducible OOXML fixture transformations; no numerical oracle injection."""
from pathlib import Path
import json,sys,zipfile,posixpath,re
from lxml import etree as E
import openpyxl
from openpyxl.formula.translate import Translator
from ooxml_edit import edit,T,RNS
ROOT=Path(__file__).resolve().parents[1];REF=ROOT/'solution/reference.xlsx';OUT=ROOT/'fixtures';OUT.mkdir(exist_ok=True)
w=openpyxl.load_workbook(REF,data_only=False);v=openpyxl.load_workbook(REF,data_only=True)
CASES=[]
def register(name,loss=(),keep=(),status='SCORED',description=''):
 CASES.append(dict(name=name,file=f'{name}.xlsx',expected_status=status,lose=list(loss),preserve=list(keep),description=description))
def patch(name,patches,**expect):
 edit(REF,OUT/(name+'.xlsx'),patches=patches,clear_caches=True);register(name,**expect)
allc=['R001','R002','R003','R004','R005','R006']
register('reference',keep=allc,description='Validated independent source-model reference.');CASES[-1]['file']='../solution/reference.xlsx'
patch('equivalent_formula',{'Submodel1s':{f'{openpyxl.utils.get_column_letter(c)}21':'=SUM('+w['Submodel1s'].cell(21,c).value[1:]+',0)' for c in range(6,22)}},keep=allc,description='Rewrites full capital growth recurrence using SUM identity; original engine computes all outputs.')
patch('equivalent_static_initialization',{'Submodel1s':{'E33':v['Submodel1s']['E33'].value,'E52':v['Submodel1s']['E52'].value}},keep=allc,description='Only immutable observed 2019 initialization is static; all restored future series remain dynamic.')
patch('wrong_timing',{'Submodel1s':{'F21':w['Submodel1s']['F21'].value.replace('E10','F10')}},loss=['R002','R003','R004','R006'],keep=['R001','R005'],description='One capital recurrence uses current investment instead of prior-year investment.')
patch('wrong_transition_year',{'InputDataB_ModelSpecAssumptions':{'E10':2024}},loss=['R001','R002','R003','R004','R006'],keep=['R005'],description='One editable input reaches target a year early.')
patch('constant_offset',{'Scenario comparison':{f'I{r}':w['Scenario comparison'][f'I{r}'].value+'+1000' for r in range(4,21)}},loss=['R003'],keep=['R001','R002','R004','R005','R006'],description='Scenario per-capita GDP report has constant 1000 level offset in every year; correct response differences must retain R004.')
patch('mixed_final',{'Scenario comparison':{'I20':w['Scenario comparison']['I20'].value+'+1000'}},loss=['R003'],keep=['R001','R002','R004','R005','R006'],description='Correct native final per-capita GDP conflicts with one final report cell; all stated representations checked.')
patch('duplicate_omission',{'Scenario comparison':{'A20':2034}},loss=['R002','R003','R004','R006'],keep=['R001'],description='Report repeats 2034 while omitting 2035; multiset assertions detect both.')
patch('legal_formula_limit',{'Submodel1s':{'F21':'=_xlfn.LAMBDA(x,x)('+w['Submodel1s']['F21'].value[1:]+')'}},status='JUDGE_ERROR',description='Legal identity LAMBDA unsupported by selected recalculation adapter is pending adjudication, never zero.')
patch('required_engine_error',{'Submodel1s':{'F21':'=NA()'}},status='JUDGE_ERROR',description='Required bound recalc error is separated from a scored business failure; unrelated existing memo #REF is ignored.')
patch('partial_parse',{'Scenario comparison':{openpyxl.utils.get_column_letter(c)+'3':'View '+str(c) for c in range(2,18)}},status='JUDGE_ERROR',description='Material results use unbound case labels while native source sheets still parse; no partial total.')
# Byte-level structural fixture: move complete named comparison table, preserving formula meaning.
with zipfile.ZipFile(REF) as z:files={n:z.read(n) for n in z.namelist()}
b=E.fromstring(files['xl/workbook.xml']);rels=E.fromstring(files['xl/_rels/workbook.xml.rels']);targets={r.get('Id'):r.get('Target') for r in rels};sheet=next(sh for sh in b.find(T('sheets')) if sh.get('name')=='Scenario comparison');sheetpath=posixpath.normpath(posixpath.join('xl',targets[sheet.get('{'+RNS+'}id')]));rt=E.fromstring(files[sheetpath])
for row in rt.find(T('sheetData')):
 row.set('r',str(int(row.get('r'))+4))
 for cell in row:
  old=cell.get('r');cr,cc=openpyxl.utils.cell.coordinate_to_tuple(old);new=f'{openpyxl.utils.get_column_letter(cc+2)}{cr+4}';cell.set('r',new);formula=cell.find(T('f'))
  if formula is not None and formula.text and '!' not in formula.text:formula.text=Translator('='+formula.text,origin=old).translate_formula(new)[1:]
files[sheetpath]=E.tostring(rt)
def shift(m):
 col,row=m.groups();return '$'+openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(col)+2)+'$'+str(int(row)+4)
for name in files:
 if name.startswith('xl/charts/') and name.endswith('.xml'):
  r=E.fromstring(files[name])
  for f in r.xpath('//*[local-name()="f"]'):
   if f.text and 'Scenario comparison' in f.text:f.text=re.sub(r'\$?([A-Z]+)\$?(\d+)',shift,f.text)
  files[name]=E.tostring(r)
with zipfile.ZipFile(OUT/'equivalent_layout.xlsx','w',zipfile.ZIP_DEFLATED) as z:
 for n,data in files.items():z.writestr(n,data)
register('equivalent_layout',keep=allc,description='Whole comparison table moves four rows/two columns; self-formulas and chart references move, source references preserve meaning.')
# Swap only visible series names while retaining their data. Semantic labels, not closeness, bind cases.
with zipfile.ZipFile(REF) as z:files={n:z.read(n) for n in z.namelist()}
for name in files:
 if name.startswith('xl/charts/') and name.endswith('.xml'):
  r=E.fromstring(files[name]);fs=r.xpath('//*[local-name()="tx"]/*[local-name()="strRef"]/*[local-name()="f"]')
  if len(fs)==2 and all('Scenario comparison' in (f.text or '') for f in fs):fs[0].text,fs[1].text=fs[1].text,fs[0].text;files[name]=E.tostring(r)
with zipfile.ZipFile(OUT/'swapped_chart_labels.xlsx','w',zipfile.ZIP_DEFLATED) as z:
 for n,data in files.items():z.writestr(n,data)
register('swapped_chart_labels',loss=['R004','R006'],keep=['R001','R002','R003','R005'],description='Chart baseline/scenario names swapped without changing series data.')
(OUT/'malformed.xlsx').write_bytes(b'This is not an XLSX file.');register('malformed',status='MALFORMED_OUTPUT',description='Completed run produced a corrupt workbook.')
register('missing',status='OUTPUT_MISSING',description='Completed run produced no workbook.')
register('unfilled_template',loss=['R002','R003','R004','R006'],keep=['R001','R005'],description='Incomplete template with no material unbound results scores missing business facts.');CASES[-1]['file']='../data/input_files/LTGM_Zambia_restore.xlsx'
# Business-equivalent source-country row reorder; current bounded protection parser is pending.
with zipfile.ZipFile(REF) as z:files={n:z.read(n) for n in z.namelist()}
b=E.fromstring(files['xl/workbook.xml']);rels=E.fromstring(files['xl/_rels/workbook.xml.rels']);targets={r.get('Id'):r.get('Target') for r in rels};sh=next(x for x in b.find(T('sheets')) if x.get('name')=='data');sp=posixpath.normpath(posixpath.join('xl',targets[sh.get('{'+RNS+'}id')]));rt=E.fromstring(files[sp]);sd=rt.find(T('sheetData'))
for rr in sd:
 old=int(rr.get('r'))
 if old not in (240,241):continue
 new=481-old;rr.set('r',str(new))
 for cell in rr:cell.set('r',re.sub(r'\d+$',str(new),cell.get('r')))
sd[:]=sorted(sd,key=lambda x:int(x.get('r')));files[sp]=E.tostring(rt)
with zipfile.ZipFile(OUT/'source_rows_reordered.xlsx','w',zipfile.ZIP_DEFLATED) as z:
 for n,bs in files.items():z.writestr(n,bs)
register('source_rows_reordered',status='JUDGE_ERROR',description='Equivalent complete source records reordered; native model parity checked, protection adapter cannot bind row relocation.')
(OUT/'manifest.json').write_text(json.dumps({'schema':'new6-calibration-v1','cases':CASES,'assertions':'Exact status, expected criterion losses and full-credit invariants; no pass/fail inferred from filename.'},indent=2))
print('Built',len(CASES),'fixture cases')
