"""Amazon source-method facts with native what-if recalculation.
Candidate tables are located by labels/year headers, never oracle values.
"""
from pathlib import Path
from collections import defaultdict,Counter
from decimal import Decimal
import argparse,json,re,sys,zipfile
import openpyxl
from openpyxl.utils import get_column_letter
TASK=Path(__file__).resolve().parents[1]
sys.path.insert(0,str(TASK.parents[1]/'common'))
sys.path.insert(0,str(TASK/'metadata'))
from runtime import recalculate_xlsx,score_profiles,output_status,sha256
from oracle_recompute import compute,raw_inputs
from ooxml_edit import edit

class Pending(Exception):pass
def norm(v):return re.sub(r'[^a-z0-9]','',str(v or '').lower())
def num(v):
 if isinstance(v,bool) or v is None:return None
 try:
  text=str(v).replace(',','').replace('£','').strip()
  return Decimal(text[:-1])/100 if text.endswith('%') else Decimal(text)
 except Exception:return None
def equal(a,b,tolerance=Decimal('.01')):
 a,b=num(a),num(b)
 return a is not None and b is not None and abs(a-b)<=max(tolerance,abs(b)*Decimal('0.000000001'))
RATE_KEYS={'growth','margin','tax','marginal_tax','riskfree','discount','discount_factor','terminal_discount','terminal_roc','price_ratio'}
def fact_equal(key,a,b):return equal(a,b,Decimal('0.00000001') if key[2] in RATE_KEYS else Decimal('.01'))
def input_equal(key,a,b):return equal(a,b,Decimal('0.00000001') if key in RATE_KEYS or key in ['shares','convergence'] else Decimal('.01'))
def mean(v):return sum(v)/len(v) if v else 0
ANNUAL={
 'growth':['Revenue growth rate','Revenue growth'], 'revenue':['Revenues','Revenue'],
 'margin':['EBIT (Operating) margin','Operating margin','EBIT margin'],
 'ebit':['EBIT (Operating income)','Operating income','EBIT'],
 'tax':['Tax rate','Effective tax rate'], 'nopat':['EBIT(1-t)','NOPAT','After tax operating income'],
 'reinvestment':['Reinvestment'],'fcff':['FCFF','Free cash flow to firm'],
 'discount':['Cost of capital','Discount rate'],
 'discount_factor':['Cumulated discount factor','Cumulative discount factor'], 'pv':['PV(FCFF)','PV of FCFF']}
BRIDGE={
 'terminal_cashflow':['Terminal cash flow'],'terminal_discount':['Terminal cost of capital'],
 'terminal_value':['Terminal value'],'terminal_pv':['PV(Terminal value)','PV of terminal value'],
 'forecast_pv':['PV (CF over next 10 years)','PV of forecast cash flows'],
 'total_pv':['Sum of PV'],'operating_assets':['Value of operating assets'],
 'debt':['Debt'],'minority':['Minority interests'],'cash':['Cash'],
 'nonoperating':['Non-operating assets'],'equity':['Value of equity','Equity value'],
 'options':['Value of options'],'common_equity':['Value of equity in common stock','Common equity value'],
 'shares':['Number of shares'],'value_per_share':['Estimated value /share','Value per share'],
 'price':['Price'],'price_ratio':['Price as % of value']}
INPUTS={
 'revenue':['Revenues'], 'ebit':['Operating income or EBIT'], 'interest':['Interest expense'],
 'equity_book':['Book value of equity'],'debt':['Book value of debt'],
 'cash':['Cash and Marketable Securities'],'nonoperating':['Cross holdings and other non-operating assets'],
 'minority':['Minority interests'],'shares':['Number of shares outstanding'], 'price':['Current stock price'],
 'tax':['Effective tax rate'],'marginal_tax':['Marginal tax rate'],
 'growth':['Compounded annual revenue growth rate over next 5 years','Five year revenue growth','Five year growth rate'],
 'margin':['Target pre-tax operating margin (EBIT as % of sales in year 10)','Target operating margin'],
 'convergence':['Year of convergence'],'riskfree':['Riskfree rate'],
 'discount':['Initial cost of capital','Initial discount rate'],
 'terminal_discount':['If yes, enter the cost of capital after year 10','Terminal discount rate'],
 'terminal_roc':['If yes, enter the return on capital you expect after year 10','Terminal return on capital']}
SUMMARY={
 'base_equity':['Base equity value','Base case equity value'],
 'review_equity':['Review equity value','Review case equity value'],
 'difference':['Review minus base equity value','Equity value difference','Difference in equity value'],
 'base_share':['Base value per share','Base case value per share'],
 'review_share':['Review value per share','Review case value per share']}
AL={norm(x):k for k,vs in ANNUAL.items() for x in vs}
BL={norm(x):k for k,vs in BRIDGE.items() for x in vs}
IL={norm(x):k for k,vs in INPUTS.items() for x in vs}
SL={norm(x):k for k,vs in SUMMARY.items() for x in vs}

def case_of(title,context=''):
 t=norm(title)
 if any(x in t for x in ['review','scenario','downside']):return 'review'
 if any(x in t for x in ['base','valuationoutput','original']):return 'base'
 t=norm(context)
 if any(x in t for x in ['reviewcase','reviewscenario']):return 'review'
 if any(x in t for x in ['basecase','originalcase']):return 'base'
 return None

def year_regions(rows,sheet_title):
 """Contiguous candidate year sequences plus local case labels define regions."""
 found={}
 for rn,row in enumerate(rows,1):
  tokens=[]
  for col,c in enumerate(row):
   nv=num(c.value);year=int(nv) if nv is not None and nv==int(nv) and 1<=nv<=11 else 11 if norm(c.value) in ['terminalyear','terminal','year11'] else None
   if year is not None:tokens.append((col,year))
  groups=[]
  for token in tokens:
   if not groups or token[0]>groups[-1][-1][0]+1 or (token[1]==1 and any(t[1]==1 for t in groups[-1])):groups.append([])
   groups[-1].append(token)
  groups=[g for g in groups if len(set(y for _,y in g))>=8 and {1,10}<=set(y for _,y in g)]
  left=0;regions=[]
  for g in groups:
   first,last=g[0][0],g[-1][0]
   label_text=' '.join(str(c.value) for c in row[left:first] if c.value is not None)
   local_case=case_of('',label_text) or case_of(sheet_title)
   # A number run alone is insufficient. Candidate metric labels must occupy
   # the leading label area of this block; no expected values select a block.
   metric_labels={AL[norm(c.value)] for rr in rows[rn:] for c in rr[left:first] if norm(c.value) in AL}
   if len(metric_labels)<4:left=last+1;continue
   if not local_case:raise Pending('Annual region has no explicit base/review case binding: '+sheet_title)
   years=defaultdict(list)
   for col,y in g:years[y].append(col)
   regions.append({'left':left,'right':last,'first_year_column':first,'case':local_case,'years':dict(years),'header_row':rn})
   left=last+1
  if regions:found[rn]=regions
 return found

def read(path, original=None):
 from formula_errors import audit
 errors=audit(original or path,path)
 def checked(cell,sheet):
  if cell.data_type=='e' and (sheet,cell.coordinate) not in errors['proven']:raise Pending('Native error lacks demonstrated candidate defect: '+sheet+'!'+cell.coordinate+' '+str(cell.value))
 w=openpyxl.load_workbook(path,data_only=True,read_only=True)
 a=defaultdict(list);b=defaultdict(list);inputs=defaultdict(list);summary=defaultdict(list)
 locs=defaultdict(list);input_locs=defaultdict(list);texts=[];tables=[];duplicates=set()
 for s in w:
  if s.sheet_state!='visible':continue
  rows=list(s.iter_rows());values=[[c.value for c in row] for row in rows]
  isinput=any(x in norm(s.title) for x in ['input','assumption','parameter']) or any(any(norm(v)=='companyname' for v in r) for r in values[:12])
  case=case_of(s.title,' '.join(str(v) for r in values[:2] for v in r if v is not None))
  regions=year_regions(rows,s.title);comparison=None
  for rn,row in enumerate(rows,1):
   vals=[c.value for c in row];texts.extend(str(v) for v in vals if isinstance(v,str))
   casecols={case_of('',str(v)):col for col,v in enumerate(vals) if norm(v) in ['basecase','originalcase','reviewcase','reviewscenario','downsidecase']}
   if set(casecols)=={'base','review'}:
    diffs=[col for col,v in enumerate(vals) if norm(v) in ['difference','equityvaluedifference','reviewminusbase']]
    comparison={'columns':casecols,'difference':diffs[0] if len(diffs)==1 else None,'header_row':rn}
   active=[r for hr,rs in regions.items() if hr<=rn for r in rs]
   for col,cell in enumerate(row):
    n=norm(cell.value)
    region=next((r for r in reversed(active) if r['left']<=col<r['first_year_column']),None)
    def right():
     stop=region['right']+1 if region else len(row)
     for cc in row[col+1:stop]:
      if cc.value is not None:return cc.value,cc.coordinate,cc.data_type
     return None,None,None
    if n in SL:
     value,coord,typ=right();summary[SL[n]].append(value)
    if isinput and n in IL:
     value,coord,typ=right();key=IL[n];inputs[key].append(value)
     if coord:input_locs[key].append((s.title,coord,value))
    if not isinput and n in BL:
     key=BL[n]
     if comparison and col<min(comparison['columns'].values()) and rn>comparison['header_row']:
      for cmp_case,cc in comparison['columns'].items():
       target=row[cc]
       checked(target,s.title)
       b[(cmp_case,key)].append(target.value)
       if key in ['equity','common_equity']:summary[cmp_case+'_equity'].append(target.value)
       if key=='value_per_share':summary[cmp_case+'_share'].append(target.value)
      if key in ['equity','common_equity'] and comparison['difference'] is not None:summary['difference'].append(row[comparison['difference']].value)
     elif region or case:
      value,coord,typ=right()
      if typ=='e':checked(s[coord],s.title)
      b[((region['case'] if region else case),key)].append(value)
   for region in regions.get(rn,[]):
    found=0;seen=Counter();end=next((hr-1 for hr in sorted(regions) if hr>rn and any(r['left']==region['left'] for r in regions[hr])),len(rows))
    for rrn,rr in enumerate(rows[rn:end],rn+1):
     labs=[AL[norm(c.value)] for c in rr[region['left']:region['first_year_column']] if norm(c.value) in AL]
     if not labs:continue
     if len(set(labs))!=1:raise Pending('Multiple annual metric labels in one region row')
     key=labs[0];found+=1
     for year,cols in region['years'].items():
      for col in cols:
       c=rr[col];coord=getattr(c,'coordinate',f'{get_column_letter(col+1)}{rrn}')
       checked(c,s.title)
       kk=(region['case'],key,year);a[kk].append(c.value);locs[kk].append((s.title,coord));seen[kk]+=1
       if seen[kk]>1:duplicates.add(kk)
    if found:tables.append({'sheet':s.title,'orientation':'years_in_columns','case':region['case'],'header_row':rn,'left_column':region['left']+1,'right_column':region['right']+1,'binding':'candidate consecutive year block plus local case/metric labels'})
   metrics={AL[norm(v)]:col for col,v in enumerate(vals) if norm(v) in AL}
   yearcols=[col for col,v in enumerate(vals) if norm(v) in ['year','forecastyear']]
   if len(metrics)>=4 and len(yearcols)==1:
    if not case:raise Pending('Transposed schedule case is unbound')
    seen=Counter()
    for rrn,rr in enumerate(rows[rn:],rn+1):
     yy=num(rr[yearcols[0]].value);yy=int(yy) if yy is not None and yy==int(yy) and 1<=yy<=11 else 11 if norm(rr[yearcols[0]].value) in ['terminal','terminalyear'] else None
     if yy is None:continue
     for key,col in metrics.items():
      c=rr[col];coord=getattr(c,'coordinate',f'{get_column_letter(col+1)}{rrn}')
      checked(c,s.title)
      a[(case,key,yy)].append(c.value);locs[(case,key,yy)].append((s.title,coord));seen[(case,key,yy)]+=1
      if seen[(case,key,yy)]>1:duplicates.add((case,key,yy))
    tables.append({'sheet':s.title,'orientation':'years_in_rows','case':case,'header_row':rn})
 w.close()
 if not tables:raise Pending('No safely bound labelled forecast table')
 return {'annual':a,'bridge':b,'inputs':inputs,'input_locs':input_locs,'summary':summary,'locations':locs,'tables':tables,'text':texts,'duplicates':duplicates,'formula_defects':[{'sheet':k[0],'cell':k[1],**v} for k,v in errors['roots'].items()]}

def expected(growth_delta=0,margin_delta=0,discount_delta=0):
 values={}
 for case in ['base','review']:
  o=compute(growth_delta,margin_delta,discount_delta,review=case=='review')
  for row in o['rows']:
   for key in ANNUAL:
    if key in ['pv','discount_factor'] and row['year']==11:continue
    values[('annual',case,key,row['year'])]=row[key]
  for key,val in o['bridge'].items():values[('bridge',case,key)]=val
 return values

def get(s,key):return s[key[0]].get(key[1:],[])
def matches(s,key,val):
 found=get(s,key)
 return not (key[0]=='annual' and key[1:] in s['duplicates']) and bool(found) and all(fact_equal(key,v,val) for v in found)
def literals(path):
 w=openpyxl.load_workbook(path,data_only=False,read_only=True)
 values=Counter()
 for sheet in w:
  if any(t in norm(sheet.title) for t in ['valuationoutput','reviewvaluation','reviewsummary']):continue
  for row in sheet:
   for c in row:
    if c.value is not None and c.data_type!='f':values[(type(c.value).__name__,str(c.value))]+=1
 w.close();return values

def source_rows(path):
 w=openpyxl.load_workbook(path,data_only=False,read_only=True)
 rows=Counter();params=defaultdict(list)
 for sheet in w:
  if any(t in norm(sheet.title) for t in ['valuationoutput','reviewvaluation','reviewsummary']):continue
  isinput=any(t in norm(sheet.title) for t in ['input','assumption','parameter'])
  for row in sheet:
   present=[c for c in row if c.value is not None]
   # Preserve row identities and ordered business fields; formatting is ignored.
   literal=tuple((type(c.value).__name__,str(c.value)) for c in present if c.data_type!='f')
   if literal:rows[(sheet.title,literal)]+=1
   if isinput:
    for i,c in enumerate(present[:-1]):
     if norm(c.value) in IL:params[IL[norm(c.value)]].append(str(present[i+1].value))
 w.close();return rows,dict(params)

def protected_source(actual,original):
 if not actual.exists():return False
 if sha256(actual)==sha256(original):return True
 ar,ap=source_rows(actual);br,bp=source_rows(original)
 if ap!=bp:return False
 if ar==br:return True
 if literals(actual)==literals(original):raise Pending('Reorganized historical source requires a bound field reader; unlabelled value multisets are insufficient')
 return False

def snapshot_facts(s,source_unchanged=True):
 target=expected();r2=[];r3=[];bad=[]
 for key,val in target.items():
  ok=matches(s,key,val)
  if key[0]=='annual' and key[2] in ['growth','revenue','margin','ebit','tax','nopat']:r2.append(ok)
  else:r3.append(ok)
  if not ok:bad.append({'fact':list(key),'expected':str(val),'candidate':get(s,key)})
 p=raw_inputs();r1=[]
 for k in INPUTS:
  found=s['inputs'].get(k,[])
  r1.append(all(input_equal(k,v,p[k]) for v in found) if found else source_unchanged)
 identity=' '.join(s['text']).lower()
 r1.extend([('amazon' in identity) or source_unchanged,('2018' in identity) or source_unchanged])
 summary_expected={
  'base_equity':target[('bridge','base','common_equity')], 'review_equity':target[('bridge','review','common_equity')],
  'difference':target[('bridge','review','common_equity')]-target[('bridge','base','common_equity')],
  'base_share':target[('bridge','base','value_per_share')],'review_share':target[('bridge','review','value_per_share')]}
 fallbacks={'base_equity':('bridge','base','common_equity'),'review_equity':('bridge','review','common_equity'),'base_share':('bridge','base','value_per_share'),'review_share':('bridge','review','value_per_share')}
 su=[]
 for key,val in summary_expected.items():
  found=list(s['summary'].get(key,[]))
  if key in fallbacks:found+=get(s,fallbacks[key])
  su.append(bool(found) and all(equal(v,val) for v in found))
 return {'R001':mean(r1),'R002':mean(r2),'R003':mean(r3),'R006':mean(su)}, {'baseline_mismatches':bad,'baseline_denominators':{'R001':len(r1),'R002':len(r2),'R003':len(r3),'R006':len(su)},'summary_checks':dict(zip(summary_expected,su))}

def evaluate(path,input_dir,work_dir=None):
 path=Path(path);ev={'candidate':str(path),'judge_version':'new6-a1-v1.3-proven-formula-defects','implementation':'label/year semantic reader; native LibreOffice only; original preserved'}
 status=output_status(path)
 if status:return score_profiles(TASK/'rubric.json',status=status,evidence=ev)
 if input_dir is None:return score_profiles(TASK/'rubric.json',status='JUDGE_ERROR',evidence={**ev,'reason':'Post-run source input directory required'})
 work=Path(work_dir or TASK/'metadata/evaluations');work.mkdir(parents=True,exist_ok=True)
 try:
  src=Path(input_dir)/'AmazonSept18_restore.xlsx';original=TASK/'data/input_files/AmazonSept18_restore.xlsx'
  source_unchanged=protected_source(src,original)
  fresh,receipt=recalculate_xlsx(path,work,timeout=90);s=read(fresh,path)
  facts,details=snapshot_facts(s,source_unchanged);ev.update(details);ev['candidate_formula_defects']=s['formula_defects']
  base=expected();changed=[];invariant=[];dyn=[]
  for control,shock,args in [('growth',Decimal('.01'),{'growth_delta':'.01'}),('margin',Decimal('-.01'),{'margin_delta':'-.01'}),('discount',Decimal('.005'),{'discount_delta':'.005'})]:
   controls=s['input_locs'].get(control,[])
   expected_after=expected(**args);active=[k for k in base if abs(expected_after[k]-base[k])>Decimal('0.0000001')];fixed=[k for k in base if k not in active]
   if not controls:
    changed.extend([False]*len(active));dyn.append({'control':control,'status':'EDITABLE_INPUT_NOT_DELIVERED','active_count':len(active)});continue
   if len(controls)!=1:raise Pending('Declared what-if input has multiple ambiguous editable locations: '+control)
   sn,co,val=controls[0]
   if num(val) is None:raise Pending('What-if input is not a numeric value: '+control)
   mutated=work/(control+'_input_changed.xlsx');edit(path,mutated,patches={sn:{co:float(num(val)+shock)}},clear_caches=True)
   calc,rec=recalculate_xlsx(mutated,work,timeout=90);after=read(calc,mutated);units=[];miss=[]
   for k in active:
    beforevals=get(s,k);aftervals=get(after,k);wanted=expected_after[k]-base[k]
    ok=bool(beforevals) and len(beforevals)==len(aftervals) and all(num(b) is not None and num(a) is not None and fact_equal(k,num(a)-num(b),wanted) for b,a in zip(beforevals,aftervals))
    units.append(ok)
    if not ok and len(miss)<12:miss.append({'fact':list(k),'expected_delta':str(wanted),'before':beforevals,'after':aftervals})
   changed.extend(units)
   for k in fixed:
    bv,av=get(s,k),get(after,k);invariant.append(bool(bv) and len(bv)==len(av) and all(fact_equal(k,x,y) for x,y in zip(bv,av)))
   for key in INPUTS:
    if key==control:continue
    bv,av=s['inputs'].get(key,[]),after['inputs'].get(key,[])
    if bv:invariant.append(len(bv)==len(av) and all(input_equal(key,x,y) for x,y in zip(bv,av)))
   dyn.append({'control':control,'input_change':str(shock),'sheet':sn,'cell':co,'active_fact_count':len(units),'active_correct':sum(units),'mismatches':miss,'native_receipt':rec})
  facts['R004']=mean(changed);facts['R005']=mean([source_unchanged,mean(invariant)]) if invariant else float(source_unchanged)
  ev.update(native_baseline=receipt,dynamic_tests=dyn,source_original_business_facts_preserved=source_unchanged,zero_delta_invariance_count=len(invariant),zero_delta_invariance_correct=sum(invariant),active_response_count=len(changed),candidate_tables=s['tables'],qualitative_explanation='Numeric two-case comparison measured; no keyword or prose-length score',formal_difficulty_verified=False)
  return score_profiles(TASK/'rubric.json',facts,evidence=ev)
 except Exception as exc:
  ev['reason']=type(exc).__name__+': '+str(exc)
  return score_profiles(TASK/'rubric.json',status='JUDGE_ERROR',evidence=ev)

if __name__=='__main__':
 p=argparse.ArgumentParser();p.add_argument('answer',nargs='?',default='/app/output/answer.xlsx');p.add_argument('--input-dir');p.add_argument('--result');p.add_argument('--work-dir');a=p.parse_args()
 r=evaluate(a.answer,a.input_dir,a.work_dir);out=json.dumps(r,indent=2,ensure_ascii=False,default=str)
 if a.result:Path(a.result).write_text(out)
 print(out)
