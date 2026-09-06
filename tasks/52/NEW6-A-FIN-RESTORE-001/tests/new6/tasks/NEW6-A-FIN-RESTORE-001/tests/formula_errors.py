"""Classify demonstrated candidate formula defects, never evaluate formulas.
Only direct scalar references and ordinary Excel arithmetic/IF/SUM are traced.
Unsupported functions, ranges, external references and unknown errors stay pending.
"""
import re
import openpyxl
from openpyxl.formula.tokenizer import Tokenizer
REF=re.compile(r"(?:(?:'((?:[^']|'')+)'|([^!]+))!)?(\$?[A-Z]{1,3}\$?[1-9][0-9]*)$")
SUPPORTED={'IF','SUM','ABS','MIN','MAX','ROUND','AND','OR','NOT'}

def audit(original,recalculated):
 w=openpyxl.load_workbook(original,data_only=False);v=openpyxl.load_workbook(recalculated,data_only=True)
 deps={};roots={};safe=set();errors={(s.title,c.coordinate):c.value for s in v for row in s for c in row if c.data_type=='e'}
 for s in w:
  for row in s:
   for c in row:
    key=(s.title,c.coordinate)
    if c.data_type!='f':continue
    try:ts=Tokenizer(c.value).items
    except Exception:continue
    refs=[];ok=True;functions=set()
    for t in ts:
     if t.type=='FUNC' and t.subtype=='OPEN':functions.add(t.value[:-1].upper())
     if t.type=='OPERAND' and t.subtype=='RANGE':
      m=REF.fullmatch(t.value)
      if not m:
       if ':' in t.value and '!' not in t.value:
        try:
         from openpyxl.utils.cell import range_boundaries,get_column_letter
         a,b,z,d=range_boundaries(t.value.replace('$',''))
         if (z-a+1)*(d-b+1)>500:raise ValueError()
         refs.extend((s.title,f'{get_column_letter(cc)}{rr}') for cc in range(a,z+1) for rr in range(b,d+1));continue
        except Exception:pass
       ok=False;continue
      sh=(m[1] or m[2] or s.title).replace("''", "'");co=m[3].replace('$','');refs.append((sh,co))
    if not ok or not functions<=SUPPORTED:continue
    safe.add(key);deps[key]=refs
    absent=[r for r in refs if r[0] not in w.sheetnames]
    if absent and key in errors:roots[key]={'reason':'reference_to_nonexistent_sheet','references':absent};continue
    if functions:
     # IF itself is supported. An observed #VALUE! plus a literal text cell
     # used directly by an arithmetic operator is a candidate type defect;
     # unknown functions were already excluded above.
     if errors.get(key)=='#VALUE!':
      for i,t in enumerate(ts):
       m=REF.fullmatch(t.value) if t.type=='OPERAND' and t.subtype=='RANGE' else None
       adjacent=[ts[j] for j in (i-1,i+1) if 0<=j<len(ts)]
       if m and any(q.type=='OPERATOR-INFIX' and q.value in '+-*/^' for q in adjacent):
        rr=((m[1] or m[2] or s.title).replace("''", "'"),m[3].replace('$',''))
        if rr[0] in w.sheetnames:
         z=w[rr[0]][rr[1]]
         if z.data_type!='f' and isinstance(z.value,str):
          try:float(z.value)
          except ValueError:roots[key]={'reason':'native_value_error_with_literal_text_arithmetic','reference':rr,'value':z.value}
     continue
    # Direct arithmetic on literal text is invalid Excel arithmetic. Do not
    # infer this for IF/comparison expressions or numeric-looking text.
    if all(t.type in {'OPERAND','OPERATOR-INFIX','OPERATOR-PREFIX','OPERATOR-POSTFIX','PAREN','WSPACE'} for t in ts) and any(t.type=='OPERATOR-INFIX' and t.value in '+-*/^' for t in ts) and not any(t.type=='OPERATOR-INFIX' and t.value in ['=','<>','<','>','<=','>=','&'] for t in ts):
     for r in refs:
      z=w[r[0]][r[1]]
      if z.data_type!='f' and isinstance(z.value,str):
       try:float(z.value)
       except ValueError:
        if key in errors:roots[key]={'reason':'arithmetic_uses_literal_text','reference':r,'value':z.value}
    for i,t in enumerate(ts[:-1]):
     if t.type=='OPERATOR-INFIX' and t.value=='/':
      nxt=ts[i+1];m=REF.fullmatch(nxt.value) if nxt.type=='OPERAND' and nxt.subtype=='RANGE' else None
      if m:
       r=((m[1] or m[2] or s.title).replace("''", "'"),m[3].replace('$',''))
       if r[0] in w.sheetnames and w[r[0]][r[1]].data_type!='f' and w[r[0]][r[1]].value in (None,0) and key in errors:roots[key]={'reason':'division_by_blank_or_zero','reference':r,'value':w[r[0]][r[1]].value}
  # Identify formula cycles only when workbook iteration is disabled. They
 # are dependency defects, not an approximation to Excel calculation.
 if not (w.calculation and w.calculation.iterate):
  visiting=[];done=set()
  def visit(k):
   if k in visiting:
    cyc=visiting[visiting.index(k):]
    for q in cyc:roots[q]={'reason':'circular_reference_iteration_disabled','cycle':cyc}
    return
   if k in done:return
   visiting.append(k)
   for q in deps.get(k,[]):
    if q in deps:visit(q)
   visiting.pop();done.add(k)
  for k in deps:visit(k)
 # Direct scalar aliases to a blank/zero remain transparent for division.
 def literal_alias(k,seen=None):
  seen=set() if seen is None else seen
  if k in seen:return False
  seen.add(k);c=w[k[0]][k[1]]
  if c.data_type!='f':return c.value in (None,0)
  m=REF.fullmatch(c.value[1:])
  if not m:return False
  q=((m[1] or m[2] or k[0]).replace("''", "'"),m[3].replace('$',''))
  return q[0] in w.sheetnames and literal_alias(q,seen)
 for k in deps:
  if k not in errors:continue
  ts=Tokenizer(w[k[0]][k[1]].value).items
  for i,t in enumerate(ts[:-1]):
   if t.type=='OPERATOR-INFIX' and t.value=='/':
    m=REF.fullmatch(ts[i+1].value)
    if m:
     q=((m[1] or m[2] or k[0]).replace("''", "'"),m[3].replace('$',''))
     if q[0] in w.sheetnames and literal_alias(q):roots[k]={'reason':'division_by_blank_or_zero_scalar_alias','reference':q}
 # Dependencies that are direct scalar formulas can propagate a proven
 # business error. A zero from a self-reference is not silently accepted.
 for key,refs in deps.items():
  if refs==[key] and w[key[0]][key[1]].value.replace('$','')=='='+key[1]:roots[key]={'reason':'direct_self_reference_without_initial_value'}
 proven=dict(roots)
 for _ in range(len(deps)+1):
  added=False
  for key,refs in deps.items():
   if key in proven:continue
   hits=[r for r in refs if r in proven]
   # Transparent scalar links can transmit a blank denominator, as in =B38.
   if hits and (key in errors or len(refs)==1 and w[key[0]][key[1]].value.startswith('=')):
    proven[key]={'reason':'depends_on_demonstrated_defect','references':hits};added=True
  if not added:break
 w.close();v.close();return {'proven':proven,'errors':errors,'roots':roots}
