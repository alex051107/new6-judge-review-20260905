"""Bounded A1 calibration. Explicit criterion assertions; no filename verdicts."""
from pathlib import Path
import argparse,json,sys,copy,zipfile
from lxml import etree as E
import openpyxl
from evaluate import TASK,evaluate,read,snapshot_facts,ANNUAL,BRIDGE,Pending,protected_source
from ooxml_edit import edit
from runtime import recalculate_xlsx
FIX=TASK/'fixtures';OUT=FIX/'validation';REF=TASK/'solution/reference.xlsx'
KEYS=['R001','R002','R003','R004','R005','R006']
ROW={'growth':2,'revenue':3,'margin':4,'ebit':5,'tax':6,'nopat':7,'reinvestment':8,'fcff':9,'discount_factor':13,'discount':12,'pv':14}
BR={'terminal_cashflow':16,'terminal_discount':17,'terminal_value':18,'terminal_pv':19,'forecast_pv':20,'total_pv':21,'operating_assets':24,'debt':25,'minority':26,'cash':27,'nonoperating':28,'equity':29,'options':30,'common_equity':31,'shares':32,'value_per_share':33,'price':34,'price_ratio':35}

def hide(path,names):
 with zipfile.ZipFile(path) as z:files={n:z.read(n) for n in z.namelist()}
 root=E.fromstring(files['xl/workbook.xml']);ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
 for s in root.find('m:sheets',ns):
  if s.get('name') in names:s.set('state','hidden')
 files['xl/workbook.xml']=E.tostring(root,xml_declaration=True,encoding='UTF-8')
 with zipfile.ZipFile(path,'w',zipfile.ZIP_DEFLATED) as z:
  for n,b in files.items():z.writestr(n,b)

def build(transposed_only=False):
 FIX.mkdir(exist_ok=True);OUT.mkdir(exist_ok=True)
 formulas=openpyxl.load_workbook(REF,data_only=False);cached=openpyxl.load_workbook(REF,data_only=True)
 sheets={}
 for name,source in [('Base presentation','Valuation output'),('Review presentation','Review valuation')]:
  cells={'A1':name,'A3':'Year'}
  for col,key in enumerate(ANNUAL,2):cells[f'{openpyxl.utils.get_column_letter(col)}3']=ANNUAL[key][0]
  for year in range(1,12):
   rr=year+3;cells[f'A{rr}']=year
   for col,key in enumerate(ANNUAL,2):
    sourcecell=f'{openpyxl.utils.get_column_letter(year+2)}{ROW[key]}'
    cells[f'{openpyxl.utils.get_column_letter(col)}{rr}']=None if year==11 and key in ['pv','discount_factor'] else f"='{source}'!{sourcecell}"
  for row,(key,sr) in enumerate(BR.items(),18):cells[f'A{row}']=BRIDGE[key][0];cells[f'B{row}']=f"='{source}'!B{sr}"
  sheets[name]=cells
 trans=FIX/'equivalent_transposed.xlsx';edit(REF,trans,new_sheets=sheets,clear_caches=True);hide(trans,['Valuation output','Review valuation'])
 edit(trans,FIX/'duplicate_year.xlsx',patches={'Base presentation':{'A15':1,**{f'{openpyxl.utils.get_column_letter(c)}15':sheets['Base presentation'][f'{openpyxl.utils.get_column_letter(c)}4'] for c in range(2,13)}}},clear_caches=True)
 edit(trans,FIX/'missing_year.xlsx',patches={'Base presentation':{'A8':None}},clear_caches=True)
 if transposed_only:return
 patch={}
 for sn in ['Valuation output','Review valuation']:
  patch[sn]={c.coordinate:f'=({c.value[1:]})+0' for row in formulas[sn] for c in row if c.data_type=='f'}
 edit(REF,FIX/'equivalent_formula.xlsx',patches=patch,clear_caches=True)
 flat={s.title:{c.coordinate:cached[s.title][c.coordinate].value for row in s for c in row if c.data_type=='f'} for s in formulas}
 edit(REF,FIX/'static.xlsx',patches=flat,clear_caches=True)
 edit(REF,FIX/'wrong_terminal_period.xlsx',patches={'Valuation output':{'B16':'=L9'}},clear_caches=True)
 edit(REF,FIX/'wrong_summary_delta.xlsx',patches={'Review summary':{'B5':'=B3-B4'}},clear_caches=True)
 edit(REF,FIX/'mixed_secondary_summary.xlsx',new_sheets={'Secondary final comparison':{'A1':'Equity value difference','B1':12345}},clear_caches=True)
 edit(REF,FIX/'unsupported_formula.xlsx',patches={'Valuation output':{'C3':'=_xlfn.PY("239343.75",0)'}},clear_caches=True)
 (FIX/'malformed.xlsx').write_bytes(b'not a workbook')

def source_protection_check():
 src=TASK/'data/input_files/AmazonSept18_restore.xlsx';styled=FIX/'source_style.xlsx';swapped=FIX/'source_values_swapped.xlsx'
 with zipfile.ZipFile(src) as z:files={n:z.read(n) for n in z.namelist()}
 root=E.fromstring(files['xl/styles.xml']);ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
 font=root.find('m:fonts/m:font/m:name',ns);font.set('val','Arial' if font.get('val')!='Arial' else 'Calibri')
 files['xl/styles.xml']=E.tostring(root,xml_declaration=True,encoding='UTF-8')
 with zipfile.ZipFile(styled,'w',zipfile.ZIP_DEFLATED) as z:
  for n,b in files.items():z.writestr(n,b)
 w=openpyxl.load_workbook(src,data_only=False);s=w['Input sheet']
 edit(src,swapped,patches={'Input sheet':{'B8':s['B12'].value,'B12':s['B8'].value}})
 assert protected_source(styled,src) is True
 assert protected_source(swapped,src) is False
 (OUT/'source_protection.json').write_text(json.dumps({'status':'PASS','style_only':'accepted','labelled_revenue_debt_swap':'rejected','native_recalculations':0},indent=2))

def assert_criteria(facts,loss,keep):
 assert set(loss).isdisjoint(keep) and len(set(loss))==len(loss) and len(set(keep))==len(keep)
 for k in loss:assert 0<=facts[k]<1,(k,facts)
 for k in keep:assert facts[k]==1,(k,facts)

def main():
 p=argparse.ArgumentParser();p.add_argument('--only',nargs='*');a=p.parse_args()
 full={'equivalent_transposed':([],KEYS),'equivalent_formula':([],KEYS),'static':(['R004'],['R001','R002','R003','R005','R006'])}
 partial={'wrong_terminal_period':(['R003','R006'],['R001','R002']),'wrong_summary_delta':(['R006'],['R001','R002','R003']),'mixed_secondary_summary':(['R006'],['R001','R002','R003']),'duplicate_year':(['R002','R003'],['R001','R006']),'missing_year':(['R002','R003'],['R001','R006'])}
 names=['reference',*full,*partial,'unsupported_formula','missing','malformed'];selected=names if a.only is None else a.only
 assert selected and len(selected)==len(set(selected)) and set(selected)<=set(names),'Unknown, empty or duplicate calibration case'
 if a.only is None:build()
 if not (OUT/'source_protection.json').exists():source_protection_check()
 results=[]
 for name in selected:
  dest=OUT/name;dest.mkdir(exist_ok=True)
  if name=='reference':
   previous=json.loads((TASK/'metadata/reference_judge_smoke.json').read_text());baseline=Path(previous['evidence']['native_baseline']['output'])
   facts,details=snapshot_facts(read(baseline),True);assert_criteria(facts,[],['R001','R002','R003','R006'])
   assert previous['criterion_scores']['R004']==previous['criterion_scores']['R005']==1
   result={'case':name,'mode':'Reused full native baseline and3 dynamics; current precision-aware snapshot rechecked','status':'SCORED','criterion_scores':{**facts,'R004':1,'R005':1},'normalized_score':1,'source_receipt':'metadata/reference_judge_smoke.json'}
  elif name in full:
   result=evaluate(FIX/(name+'.xlsx'),TASK/'data/input_files',dest)
   (dest/'result.json').write_text(json.dumps(result,indent=2,default=str))
   assert result['evaluation_status']=='SCORED',result.get('evidence',{}).get('reason')
   assert_criteria(result['criterion_scores'],*full[name])
   if name=='static':assert result['criterion_scores']['R004']==0 and result['normalized_score']==.65
   if name=='equivalent_transposed':assert {t['orientation'] for t in result['evidence']['candidate_tables']}=={'years_in_rows'}
   result={'case':name,'mode':'Full native baseline plus3 declared dynamics',**result}
  elif name in partial:
   fresh,receipt=recalculate_xlsx(FIX/(name+'.xlsx'),dest);facts,details=snapshot_facts(read(fresh),True)
   assert_criteria(facts,*partial[name]);result={'case':name,'mode':'Focused native baseline snapshot; dynamic behavior not claimed','criterion_scores':facts,'evidence':details,'native_receipt':receipt,'asserted_loss':partial[name][0],'asserted_keep':partial[name][1]}
  else:
   result=evaluate(FIX/(name+'.xlsx'),TASK/'data/input_files',dest);wanted='JUDGE_ERROR' if name=='unsupported_formula' else 'OUTPUT_MISSING' if name=='missing' else 'MALFORMED_OUTPUT'
   assert result['evaluation_status']==wanted,result
   assert result['normalized_score'] is None
   result={'case':name,'mode':'Delivery/parser status category',**result}
  (dest/'result.json').write_text(json.dumps(result,indent=2,default=str));results.append(result);print(name,'PASS',result.get('normalized_score'),flush=True)
 receipt={'status':'PASS','cases':len(results),'case_names':selected,'checks':[{'case':r['case'],'mode':r['mode'],'criterion_scores':r.get('criterion_scores'),'normalized_score':r.get('normalized_score'),'status':r.get('evaluation_status',r.get('status'))} for r in results],'assertions_are_criterion_specific':True,'filename_does_not_determine_pass':True,'agent_calls':0,'full_suite_repeated':False,'known_limits':'Focused baseline mutants do not claim measured dynamic scores; pending legal engines and delivery failures remain unscored.'}
 if all((OUT/n/'result.json').exists() for n in names):
  complete=[json.loads((OUT/n/'result.json').read_text()) for n in names];receipt['cases']=len(complete);receipt['case_names']=names;receipt['checks']=[{'case':r['case'],'mode':r['mode'],'criterion_scores':r.get('criterion_scores'),'normalized_score':r.get('normalized_score'),'status':r.get('evaluation_status',r.get('status'))} for r in complete]
  (OUT/'receipt.json').write_text(json.dumps(receipt,indent=2))
 (OUT/('receipt.json' if a.only is None else 'repair_receipt.json')).write_text(json.dumps(receipt,indent=2));print(json.dumps(receipt,indent=2))
if __name__=='__main__':main()
