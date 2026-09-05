"""Independent ONS source oracle. No candidate workbook is read.

Source cells are discovered from original public headers. A second OOXML parser
checks the source identity and stored numeric rates without using openpyxl.
Published-display precision is a visible project convention, not latent truth.
"""
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
import json, re, zipfile, xml.etree.ElementTree as ET
import openpyxl

TASK = Path(__file__).resolve().parents[1]
NS = {'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
def published(v):
    if isinstance(v, (int,float)) and not isinstance(v,bool):
        return str(Decimal(str(v)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP))
    return None

def xml_rows(path):
    """Independent physical OOXML source read; original LI01 is sheet 3."""
    with zipfile.ZipFile(path) as z:
        ss=[]
        if 'xl/sharedStrings.xml' in z.namelist():
            for si in ET.fromstring(z.read('xl/sharedStrings.xml')).findall('m:si',NS):
                ss.append(''.join(x.text or '' for x in si.iter('{'+NS['m']+'}t')))
        rows=[]
        for row in ET.fromstring(z.read('xl/worksheets/sheet3.xml')).findall('m:sheetData/m:row',NS):
            out={}
            for c in row.findall('m:c',NS):
                v=c.find('m:v',NS);t=c.attrib.get('t');value=None
                if v is not None:
                    value=ss[int(v.text)] if t=='s' else v.text
                elif t=='inlineStr':value=''.join(x.text or '' for x in c.iter('{'+NS['m']+'}t'))
                out[re.sub(r'\d','',c.attrib['r'])]=value
            if str(out.get('B','')).startswith(('E06','E07','E08','E09')):rows.append(out)
        return rows

def read_edition(path):
    w=openpyxl.load_workbook(path,data_only=True);s=w['LI01']
    headers={str(c.value).strip().split('\n')[0]:c.column for c in s[5] if c.value}
    assert headers['Geography code']==2
    e=headers['Employment rate'];u=headers['Unemployment rate']
    assert 'age 16 to 64' in s.cell(5,e).value and 'age 16 and older' in s.cell(5,u).value
    rows={};all_geo={}
    for r in s.iter_rows(min_row=6):
        name,code=r[0].value,r[1].value
        if not code:continue
        assert code not in all_geo, ('duplicate source code',code)
        all_geo[code]=name
        if not code.startswith(('E06','E07','E08','E09')):continue
        raw_e,raw_u=r[e-1].value,r[u-1].value
        rows[code]={'geography_code':code,'geography':name,'employment_rate':published(raw_e),
          'unemployment_rate':published(raw_u),'employment_raw':str(raw_e),'unemployment_raw':str(raw_u),
          'employment_location':f'{path.name}#LI01!{r[e-1].coordinate}',
          'unemployment_location':f'{path.name}#LI01!{r[u-1].coordinate}',
          'employment_format':r[e-1].number_format,'unemployment_format':r[u-1].number_format}
    independent=xml_rows(path)
    assert set(rows)=={r['B'] for r in independent}
    for r in independent:
        for column,metric in [('E','employment'),('G','unemployment')]:
            raw=r.get(column)
            try:val=str(Decimal(raw).quantize(Decimal('.1'),rounding=ROUND_HALF_UP))
            except Exception:val=None
            assert val==rows[r['B']][metric+'_rate'],(r['B'],metric)
    return rows,all_geo,{'employment_header':s.cell(5,e).value,'unemployment_header':s.cell(5,u).value,
      'notes':[r[1].value for r in w['Notes'].iter_rows(min_row=3)],
      'cover':[r[0].value for r in w['Cover_sheet'].iter_rows() if r[0].value]}

def recompute(input_dir=None):
    d=Path(input_dir or TASK/'data/input_files')
    earlier,geo1,notes1=read_edition(d/'ons_li01_january2024.xlsx')
    later,geo2,notes2=read_edition(d/'ons_li01_january2025.xlsx')
    panel=[];exclusions=[];provenance=[]
    for code in sorted(set(earlier)|set(later)):
        old,new=earlier.get(code),later.get(code)
        if old is None or new is None:
            exclusions.append([code,(old or new)['geography'],'both','missing geography in one edition']);continue
        vals=[old['employment_rate'],new['employment_rate'],old['unemployment_rate'],new['unemployment_rate']]
        em_ok=all(x is not None for x in vals[:2]);un_ok=all(x is not None for x in vals[2:])
        em_delta=str(Decimal(vals[1])-Decimal(vals[0])) if em_ok else None
        un_delta=str(Decimal(vals[3])-Decimal(vals[2])) if un_ok else None
        panel.append([code,new['geography'],*vals,em_delta,un_delta,'comparable' if em_ok and un_ok else 'partial_or_excluded'])
        for metric,okay in [('employment',em_ok),('unemployment',un_ok)]:
            if not okay:
                exclusions.append([code,new['geography'],metric,f"unavailable/suppressed source: January2024={old[metric+'_raw']}; January2025={new[metric+'_raw']}"])
        for edition,item in [('January 2024',old),('January 2025',new)]:
            for metric in ['employment','unemployment']:
                provenance.append([code,edition,metric,item[metric+'_raw'],item[metric+'_rate'],item[metric+'_location']])
    # Population definition for ranking is unemployment-comparable. Employment
    # unavailable does not erase an otherwise comparable unemployment pair.
    ranked=sorted((r for r in panel if r[7] is not None),key=lambda r:(-Decimal(r[7]),r[0]))[:5]
    top=[[i+1,r[0],r[1],r[7],r[6]] for i,r in enumerate(ranked)]
    # Separate sort expressed through stable ascending sort plus group inversion.
    groups={}
    for r in panel:
        if r[7] is not None:groups.setdefault(Decimal(r[7]),[]).append(r[0])
    independent=[c for change in sorted(groups,reverse=True) for c in sorted(groups[change])][:5]
    assert independent==[r[1] for r in top]
    return {'panel':panel,'exclusions':exclusions,'provenance':provenance,'top5':top,
      'eligible_union_count':len(set(earlier)|set(later)), 'employment_comparable_count':sum(r[6] is not None for r in panel),
      'unemployment_comparable_count':sum(r[7] is not None for r in panel),
      'noneligible_source_geographies':sorted((set(geo1)|set(geo2))-(set(earlier)|set(later))),
      'earlier_source_context':notes1,'later_source_context':notes2,
      'independent_verification':'All eligible original source rates and codes matched independent OOXML read; rank matched independent grouped sort.'}

if __name__=='__main__':
    out=recompute();(TASK/'solution/oracle.json').write_text(json.dumps(out,ensure_ascii=False,indent=2))
    print(json.dumps({k:out[k] for k in ['eligible_union_count','employment_comparable_count','unemployment_comparable_count','top5','independent_verification']},ensure_ascii=False,indent=2))
