"""Focused input-alias calibration, without Agent/API calls.

Eight native calculations maximum: four each for the alias-equivalent reference
and its disconnected-output mutant. Pure binding boundary checks do not recalc.
"""
from pathlib import Path
import sys,json,argparse
from decimal import Decimal
import openpyxl
TASK=Path(__file__).resolve().parents[1]
sys.path[:0]=[str(TASK/'tests'),str(TASK/'metadata')]
from evaluate import controls,checks_for,evaluate,JUDGE_VERSION
from ooxml_edit import edit

ALL=['R001','R002','R003','R004','R005','R006']

def small_book():
    w=openpyxl.Workbook();s=w.active;s.title='Native controls'
    s['D1']='Baseline';s['E1']='Scenario'
    s['B9']='Target Investment Ratio (I/Y)';s['E9']="='Meeting inputs'!$F$5"
    s['B10']='Reach target I/Y rate by year';s['E10']="='Meeting inputs'!$B$2"
    s=w.create_sheet('Meeting inputs')
    s['A1']='Scenario target investment / GDP';s['B1']=.24
    s['A2']='Transition year';s['B2']=2025;s['F5']='=$B$1'
    return w

def run(out):
    out.mkdir(parents=True,exist_ok=False);checks=[]
    def check(name,action):
        try:details=action();checks.append({'id':name,'passed':True,'details':details})
        except Exception as exc:
            checks.append({'id':name,'passed':False,'error':type(exc).__name__+': '+str(exc)})
            receipt(False);raise
    def receipt(passed):
        r={'passed':passed,'judge_version':JUDGE_VERSION,'checks':checks,'api_calls':0,'validation_budget':{'focused_suite_invocations':1,'native_recalculations_maximum':8,'independent_reviews':0,'additional_routine_hashes':0},'unchanged':'task, Oracle, weights and fact denominators; original artifacts and old receipts remain unchanged'}
        (out/'receipt.json').write_text(json.dumps(r,ensure_ascii=False,indent=2,default=str));return r
    def reference():
        raw=openpyxl.load_workbook(TASK/'solution/reference.xlsx',data_only=False)
        cached=openpyxl.load_workbook(TASK/'solution/reference.xlsx',data_only=True)
        binding=controls(cached,formula_workbook=raw);assert binding=={'target':('InputDataB_ModelSpecAssumptions','E9'),'transition':('InputDataB_ModelSpecAssumptions','E10')},binding
        details,_,_=checks_for(cached,control_bindings=binding)
        assert all(x['ok'] for k,items in details.items() if k!='R004' for x in items)
        return {'controls':binding,'static_reference_facts':'all correct; native parity established by the equivalent-reference run below'}
    check('reference_binding_and_static_facts',reference)
    def chain():
        w=small_book();got=controls(w)
        assert got=={'target':('Meeting inputs','B1'),'transition':('Meeting inputs','B2')},got
        assert any(len(x['direct_reference_chain'])==3 for x in w._new6_control_aliases)
        return {'controls':got,'aliases':w._new6_control_aliases}
    check('entry_mirror_and_reference_chain',chain)
    def pending(value,needle,patch=None):
        w=small_book();w['Native controls']['E9']=value
        if patch:patch(w)
        try:controls(w)
        except ValueError as exc:
            assert needle in str(exc),str(exc)
            return {'status':'JUDGE_ERROR','reason':str(exc)}
        raise AssertionError('Unproven control equivalence was silently accepted')
    check('same_value_without_link_is_not_merged',lambda:pending(.24,'Independent scenario target'))
    check('different_independent_controls_not_hidden',lambda:pending(.26,'Independent scenario target'))
    check('direct_reference_cycle_is_pending',lambda:pending("='Meeting inputs'!F5",'reference cycle',lambda w:w['Meeting inputs'].__setitem__('F5',"='Native controls'!E9")))
    check('external_reference_is_pending',lambda:pending("='[other.xlsx]Inputs'!B1",'External or structured'))
    check('complex_legal_control_formula_is_pending',lambda:pending("=SUM('Meeting inputs'!B1)",'outside direct-cell alias support'))

    ref=TASK/'solution/reference.xlsx';alias=out/'fixtures/alias_reference.xlsx'
    edit(ref,alias,patches={'InputDataB_ModelSpecAssumptions':{'E9':"='Meeting inputs'!$F$5",'E10':"='Meeting inputs'!$B$2"}},new_sheets={'Meeting inputs':{'A1':'Scenario target investment / GDP','B1':.24,'A2':'Transition year','B2':2025,'F5':'=$B$1'}},clear_caches=True)
    cached=openpyxl.load_workbook(ref,data_only=True)
    frozen={openpyxl.utils.get_column_letter(c)+'10':cached['Submodel1s'].cell(10,c).value for c in range(5,22)}
    assert all(isinstance(v,(int,float)) for v in frozen.values())
    disconnected=out/'fixtures/disconnected_investment_schedule.xlsx'
    edit(alias,disconnected,patches={'Submodel1s':frozen},clear_caches=True)
    for name,path,lose in [('alias_reference_full_judge',alias,False),('disconnected_schedule_loses_dynamic_facts',disconnected,True)]:
        def native_case(name=name,path=path,lose=lose):
            result=evaluate(path,out/name,completed_run=True)
            assert result['evaluation_status']=='SCORED',result['evidence']
            facts=result['criterion_scores']
            for k in ALL:
                if k=='R004' and lose:assert Decimal(facts[k])<1,facts
                else:assert Decimal(facts[k])==1,(k,facts)
            for probe in result['evidence']['dynamic_probes']:
                assert set(probe['changes'])=={'Meeting inputs'},probe['changes']
                assert set(probe['changes']['Meeting inputs'])=={'B1','B2'},probe['changes']
            if lose:
                item=next(x for x in result['evidence']['fact_units']['R004'] if x['id']=='target_026:scenario:2035:investment_share')
                assert item['ok'] is False and item['before'][0]['value']==item['after'][0]['value'],item
            return {'score_decimal':result['score_decimal'],'criterion_scores':facts,'native_calculations':4,'result':str((out/name/'evaluation.json').relative_to(out))}
        check(name,native_case)
        print(json.dumps(checks[-1],ensure_ascii=False),flush=True)
    result=receipt(True);print(json.dumps(result,ensure_ascii=False,default=str));return result

if __name__=='__main__':
    p=argparse.ArgumentParser();p.add_argument('--out',type=Path,required=True);a=p.parse_args();run(a.out)
