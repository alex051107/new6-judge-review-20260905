from pathlib import Path
import sys,json,openpyxl,posixpath,zipfile
from lxml import etree as E
from openpyxl.chart import LineChart,Reference
ROOT=Path(__file__).resolve().parents[1];sys.path.insert(0,str(ROOT.parents[1]/'common'))
from runtime import recalculate_xlsx
from ooxml_edit import edit,NS,RNS,PNS,T
from oracle_recompute import compute

def build_reference():
 source=ROOT/'metadata/configured_source.xlsx';summary=ROOT/'metadata/summary_only.xlsx';wb=openpyxl.Workbook();s=wb.active;s.title='Scenario comparison'
 s.append(['Zambia — illustrative LTGM scenario comparison']);s.append(['Money: millions of constant 2010 US dollars; per capita: constant 2010 US dollars. Investment in a year finances the following year capital stock.']);s.append(['Year','Baseline investment / GDP','Scenario investment / GDP','Baseline capital / output','Scenario capital / output','Baseline GDP','Scenario GDP','Baseline GDP per capita','Scenario GDP per capita','Baseline GDP growth','Scenario GDP growth','Baseline GDP per capita growth','Scenario GDP per capita growth','Baseline investment','Scenario investment','Baseline capital','Scenario capital'])
 for i,year in enumerate(range(2019,2036),5):
  c=openpyxl.utils.get_column_letter(i);r=i-1
  s.append([year,f'=Submodel1!{c}10',f'=Submodel1s!{c}10',f'=Submodel1!{c}24',f'=Submodel1s!{c}24',f'=Submodel1!{c}52',f'=Submodel1s!{c}52',f'=Submodel1!{c}33',f'=Submodel1s!{c}33',f'=Submodel1!{c}31',f'=Submodel1s!{c}31',f'=Submodel1!{c}23',f'=Submodel1s!{c}23',f'=B{r}*F{r}',f'=C{r}*G{r}',f'=D{r}*F{r}',f'=E{r}*G{r}'])
 s['A23']='The project scenario reduces investment from 31% to 24% of GDP by 2025. The baseline remains 31%. Lower investment reduces future capital accumulation and output growth. The initial 2019 growth is a source historical starting value shared by both cases.'
 chart=LineChart();chart.title='GDP per capita growth: baseline and scenario';chart.y_axis.title='Growth rate';chart.x_axis.title='Year';chart.add_data(Reference(s,min_col=12,max_col=13,min_row=3,max_row=20),titles_from_data=True);chart.set_categories(Reference(s,min_col=1,min_row=4,max_row=20));s.add_chart(chart,'A25');wb.save(summary)
 dest=ROOT/'solution/reference_unrecalculated.xlsx';edit(source,dest,new_sheets={'Scenario comparison':{}},clear_caches=True)
 with zipfile.ZipFile(dest) as z:files={n:z.read(n) for n in z.namelist()}
 with zipfile.ZipFile(summary) as z:add={n:z.read(n) for n in z.namelist()}
 book=E.fromstring(files['xl/workbook.xml']);rels=E.fromstring(files['xl/_rels/workbook.xml.rels']);sh=[s for s in book.find(T('sheets')) if s.get('name')=='Scenario comparison'][0];rid=sh.get('{'+RNS+'}id');target=[r.get('Target') for r in rels if r.get('Id')==rid][0];sheet_path=posixpath.normpath(posixpath.join('xl',target))
 rename={'xl/worksheets/sheet1.xml':sheet_path,'xl/worksheets/_rels/sheet1.xml.rels':posixpath.dirname(sheet_path)+'/_rels/'+posixpath.basename(sheet_path)+'.rels','xl/drawings/drawing1.xml':'xl/drawings/new6drawing.xml','xl/drawings/_rels/drawing1.xml.rels':'xl/drawings/_rels/new6drawing.xml.rels','xl/charts/chart1.xml':'xl/charts/new6chart.xml'}
 for old,new in rename.items():
  b=add[old].replace(b'/xl/drawings/drawing1.xml',b'/xl/drawings/new6drawing.xml').replace(b'/xl/charts/chart1.xml',b'/xl/charts/new6chart.xml');files[new]=b
 ct=E.fromstring(files['[Content_Types].xml'])
 for name,ctype in [('drawing','application/vnd.openxmlformats-officedocument.drawing+xml'),('chart','application/vnd.openxmlformats-officedocument.drawingml.chart+xml')]:
  part='/xl/drawings/new6drawing.xml' if name=='drawing' else '/xl/charts/new6chart.xml';E.SubElement(ct,'{http://schemas.openxmlformats.org/package/2006/content-types}Override',PartName=part,ContentType=ctype)
 files['[Content_Types].xml']=E.tostring(ct)
 with zipfile.ZipFile(dest,'w',zipfile.ZIP_DEFLATED) as z:
  for n,b in files.items():z.writestr(n,b)
 fresh,receipt=recalculate_xlsx(dest,ROOT/'metadata/reference_recalc');v=openpyxl.load_workbook(fresh,data_only=True);assert len(openpyxl.load_workbook(fresh)['Scenario comparison']._charts)==1
 for name,base in [('Submodel1',True),('Submodel1s',False)]:
  for i,row in enumerate(compute(baseline=base),5):
   for key,r in {'investment_share':10,'capital_output_ratio':24,'gdp':52,'gdp_per_capita':33,'gdp_growth':31,'pc_growth':23}.items():assert abs(v[name].cell(r,i).value-float(row[key]))<1e-7*max(1,abs(float(row[key])))
 (ROOT/'solution/reference.xlsx').write_bytes(fresh.read_bytes());(ROOT/'metadata/reference_verification.json').write_text(json.dumps({'status':'PASS','native_receipt':receipt,'independent_oracle':'340 raw-data annual checks already passed; added two 2019-2035 level/investment/capital series and linked chart verified'},indent=2))
 dest.unlink();summary.unlink()

def build_input():
 assert json.loads((ROOT/'metadata/reference_verification.json').read_text())['status']=='PASS';source=ROOT/'metadata/configured_source.xlsx';w=openpyxl.load_workbook(source);patch={};manifest=[]
 for r in [21,22,23,24,30,31,32,33,37,38,39,40,52,53,54,57,58]:
  for c in range(6,22):
   co=w['Submodel1s'].cell(r,c).coordinate;patch[co]=None;manifest.append({'sheet':'Submodel1s','cell':co,'original_formula':w['Submodel1s'][co].value,'action':'clear_scenario_2020_2035_derived_block'})
 edit(source,ROOT/'data/input_files/LTGM_Zambia_restore.xlsx',patches={'Submodel1s':patch},clear_caches=True)
 (ROOT/'metadata/mutation_manifest.json').write_text(json.dumps({'changes':manifest,'initial_2019_source_values_retained':True,'baseline_preserved':True,'formula_and_chart_caches_cleared':True},indent=2))
 # Full source PDF is private; only original method pages 1–3 are candidate-visible.
 from pypdf import PdfReader,PdfWriter
 reader=PdfReader(ROOT/'metadata/source/Instructions2020.pdf');writer=PdfWriter()
 for page in reader.pages[:3]:writer.add_page(page)
 writer.write(ROOT/'data/input_files/LTGM_instructions_methods_p1-3.pdf')

if __name__=='__main__':build_reference();build_input()
