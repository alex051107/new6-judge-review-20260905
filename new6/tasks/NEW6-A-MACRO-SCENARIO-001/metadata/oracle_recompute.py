"""LTGM v4.42 Model 1 task-specific Decimal recursion from raw country data.
Source GDP/population/demographic initial values only; no model output series.
I_t/Y_t raises K_(t+1), then GDP uses contemporaneous exogenous growth drivers.
"""
from decimal import Decimal,getcontext
from pathlib import Path
import openpyxl,json
from functools import lru_cache
getcontext().prec=40
D=lambda v:Decimal(str(v))
ROOT=Path(__file__).resolve().parent

@lru_cache(maxsize=1)
def raw_drivers():
 w=openpyxl.load_workbook(ROOT/'source/LTGMv4-42-TrainingVersion.xlsx',data_only=True);s=w['data'];rows=[r for r in range(23,242) if s.cell(r,3).value=='ZMB'];assert len(rows)==1;r=rows[0]
 def annual(a,b):
  return {int(s.cell(21,c).value):D(s.cell(r,c).value) for c in range(openpyxl.utils.column_index_from_string(a),openpyxl.utils.column_index_from_string(b)+1)}
 return {'population':annual('FE','HC'),'male_share':annual('IV','KE'),'working_share':annual('KI','LR'),'male_participation':D(s[f'DP{r}'].value),'female_participation':D(s[f'DT{r}'].value),'initial_pc':D(s[f'SW{r}'].value),'initial_pc_growth':D(s[f'SE{r}'].value),'initial_participation_growth':D(s[f'PQ{r}'].value),'provenance':{'country':'Zambia','code':'ZMB','data_row':r,'population':'FE:HC, year headers row 21','working_share':'KI:LR, ages 15-64, year headers row 21','male_share':'IV:KE, year headers row 21','initial_pc':'SW','initial_pc_growth':'SE'}}

@lru_cache(maxsize=32)
def compute(target='.24',transition_year=2025,baseline=False):
 d=raw_drivers();target=D('.31') if baseline else D(target);transition_year=int(transition_year);assert 2019<transition_year<=2050
 beta=D('.618');dep=D('.049');h=D('.006');tfp=D('.01');ky=D('2.42');pc=d['initial_pc'];result=[]
 participation=lambda y:d['male_share'][y]*d['male_participation']+(1-d['male_share'][y])*d['female_participation']
 for year in range(2019,2036):
  investment=D('.31')+(target-D('.31'))*min(D(year-2019)/D(transition_year-2019),D(1))
  pop=d['population'][year]/D(1000000);ng=d['population'][year]/d['population'][year-1]-1;wag=d['working_share'][year]/d['working_share'][year-1]-1;pg=d['initial_participation_growth'] if year==2019 else participation(year)/participation(year-1)-1
  if year==2019:kworker=None;yworker=None;gpc=d['initial_pc_growth']
  else:
   prev=result[-1];kworker=((1-dep)+prev['investment_share']/prev['capital_output_ratio'])/((1+ng)*(1+wag)*(1+pg))-1
   yworker=(1+tfp)*(1+kworker)**(1-beta)*(1+h)**beta-1
   gpc=(1+yworker)*(1+wag)*(1+pg)-1
   ky*= (1+kworker)/(1+yworker);pc*=1+gpc
  gdp=pc*pop
  result.append(dict(year=year,investment_share=investment,capital_output_ratio=ky,capital=gdp*ky,investment=gdp*investment,gdp=gdp,gdp_per_capita=pc,gdp_growth=(1+gpc)*(1+ng)-1,pc_growth=gpc,population=pop,population_growth=ng,working_share_growth=wag,participation_growth=pg,capital_per_worker_growth=kworker,gdp_per_worker_growth=yworker))
 return result

def jsonable(x):
 if isinstance(x,Decimal):return str(x)
 if isinstance(x,list):return [jsonable(v) for v in x]
 if isinstance(x,dict):return {k:jsonable(v) for k,v in x.items()}
 return x
if __name__=='__main__':print(json.dumps(jsonable({'baseline':compute(baseline=True),'scenario':compute()}),indent=2))
