"""Read retained input facts as context for a separately delivered workbook.

No source-derived output schedules are inserted; candidate output evidence stays
in its own labelled tables. The in-memory context is never saved to the answer.
"""
import re
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
INPUT_DIR=ROOT/'data/input_files'
SOURCE_CACHE={}
norm=lambda value:re.sub(r'[^a-z0-9]+','',str(value or '').lower())


DIRECT_REFERENCE=re.compile(r"^=\s*(?:(?:'((?:[^']|'')+)'|([A-Za-z_][A-Za-z0-9_.]*))!)?(\$?[A-Za-z]{1,3}\$?[1-9]\d*)\s*$")

def control_terminal(w,binding):
    """Follow direct cell aliases only; never infer aliases from equal values."""
    seen=[];names={name.casefold():name for name in w.sheetnames}
    while True:
        if binding in seen:raise ValueError('Scenario control reference cycle: '+str(seen+[binding]))
        if len(seen)>=128:raise ValueError('Scenario control direct-reference chain exceeds supported bound')
        seen.append(binding);sheet,coord=binding
        if sheet not in w:raise ValueError('Scenario control reference points to an unknown worksheet: '+sheet)
        cell=w[sheet][coord]
        if cell.data_type!='f':
            if isinstance(cell.value,bool) or not isinstance(cell.value,(int,float)):
                raise ValueError('Scenario control does not terminate at a numeric input: '+str(binding))
            return binding,seen
        formula=str(cell.value)
        if '[' in formula or ']' in formula:raise ValueError('External or structured scenario-control reference requires a supported binding: '+str(binding))
        match=DIRECT_REFERENCE.fullmatch(formula)
        if not match:raise ValueError('Scenario control uses a legal expression outside direct-cell alias support: '+str(binding))
        target_sheet=(match[1].replace("''", "'") if match[1] else match[2]) or sheet
        if target_sheet.casefold() not in names:raise ValueError('Scenario control reference points to an unknown worksheet: '+target_sheet)
        binding=(names[target_sheet.casefold()],match[3].replace('$','').upper())

def bind_controls(w, native, formula_workbook=None):
    formula_workbook=formula_workbook if formula_workbook is not None else w
    discovered={}
    for key,value in native(w).items():
        discovered[key]=list(value) if isinstance(value,list) else [value]
    names={'scenariotargetinvestmentgdp':'target','transitioncompletebyyear':'transition','transitionyear':'transition'}
    for sh in w:
        for row in sh.iter_rows(max_row=min(sh.max_row,100),max_col=min(sh.max_column,15)):
            for c in row:
                key=names.get(norm(c.value))
                if not key:continue
                values=[x for x in row[c.column:c.column+3] if (isinstance(x.value,(int,float)) and not isinstance(x.value,bool)) or x.data_type=='f']
                if len(values)!=1:raise ValueError('Labelled scenario control is not unique')
                binding=(sh.title,values[0].coordinate)
                discovered.setdefault(key,[]).append(binding)
    result={};aliases=[]
    for key,bindings in discovered.items():
        for binding in dict.fromkeys(bindings):
            terminal,chain=control_terminal(formula_workbook,binding)
            if key in result and result[key]!=terminal:
                raise ValueError('Independent scenario '+key+' controls have no direct-reference link; equal values do not establish equivalence: '+str([result[key],terminal]))
            result[key]=terminal;aliases.append({'control':key,'declared_cell':binding,'terminal_input':terminal,'direct_reference_chain':chain})
    w._new6_control_aliases=aliases
    formula_workbook._new6_control_aliases=aliases
    return result


def attach_retained_source(w):
    required=['InputDataA_GeneralAssumptions','data','Readme']
    if all(sn in w for sn in required):return
    # A partly preserved or renamed original still needs explicit semantic review.
    if any(sn in w for sn in required):raise ValueError('Partly retained source layout needs a supported binding')
    source=INPUT_DIR/'LTGM_Zambia_restore.xlsx'
    if not source.exists():raise ValueError('Actual retained input was not collected')
    if str(source) not in SOURCE_CACHE:
        from runtime import recalculate_xlsx
        calculated,receipt=recalculate_xlsx(source,Path(os.environ.get('NEW6_EVIDENCE_DIR', tempfile.gettempdir()))/'retained_input_recalc')
        SOURCE_CACHE[str(source)]=(load_workbook(source,data_only=False),load_workbook(calculated,data_only=True),receipt)
    src,computed,receipt=SOURCE_CACHE[str(source)]
    if not all(sn in src for sn in required):raise ValueError('Retained source layout cannot be safely bound')
    overrides={}
    names={'depreciationrate':['D13','E13'],'laborshare':['D16','E16'],'initialcapitaloutputratio':['I13','J13']}
    for sh in w:
        for row in sh.iter_rows(max_row=min(sh.max_row,100),max_col=min(sh.max_column,15)):
            for c in row:
                keys=names.get(norm(c.value))
                if not keys:continue
                values=[x.value for x in row[c.column:c.column+2] if isinstance(x.value,(int,float))]
                if len(values)!=1:raise ValueError('Candidate source assumption has ambiguous value')
                for key in keys:
                    if key in overrides and overrides[key]!=values[0]:raise ValueError('Contradictory candidate source assumptions')
                    overrides[key]=values[0]
    for sn in required:
        target=w.create_sheet(sn)
        section=computed[sn] if sn=='InputDataA_GeneralAssumptions' else src[sn]
        for row in section:
            for cell in row:
                if cell.value is not None:target.cell(cell.row,cell.column,cell.value)
    for cell,value in overrides.items():w['InputDataA_GeneralAssumptions'][cell]=value
    w._new6_retained_source_context=str(source)
    w._new6_retained_source_recalc=receipt
