"""A1 affected-layout calibration and same-original first-attempt adjudication."""
import json,sys,zipfile,hashlib
from pathlib import Path
import openpyxl
from lxml import etree as E
from evaluate import TASK,read,snapshot_facts,evaluate
from ooxml_edit import edit
OUT=TASK/'metadata/first_attempt_rejudge_v12';FIX=TASK/'fixtures';REF=TASK/'solution/reference.xlsx'
TRIAL=TASK.parents[1]/'harbor_jobs/new6-first-20260905T153935Z/NEW6-A-FIN-RESTORE-001__ej5SFoo'

def build_adjacent():
 w=openpyxl.load_workbook(REF,data_only=False);sheets={};cells={}
 for source,offset in [('Valuation output',0),('Review valuation',14)]:
  s=w[source]
  for row in s.iter_rows():
   for c in row:
    if c.value is None:continue
    target=f'{openpyxl.utils.get_column_letter(c.column+offset)}{c.row}'
    cells[target]=f"='{source}'!{c.coordinate}" if c.data_type=='f' else c.value
  if offset:cells['O1']='REVIEW SCENARIO'
  else:cells['A1']='BASE CASE'
 sheets['Adjacent case valuation']=cells
 sheets['Labelled case comparison']={'A1':'Valuation comparison','A3':'Metric','B3':'Base Case','C3':'Review Scenario','D3':'Difference','A4':'Value of equity','B4':"='Valuation output'!B29",'C4':"='Review valuation'!B29",'D4':'=C4-B4','A5':'Value per share','B5':"='Valuation output'!B33",'C5':"='Review valuation'!B33",'D5':'=C5-B5'}
 path=FIX/'equivalent_adjacent_cases.xlsx';edit(REF,path,new_sheets=sheets,clear_caches=True)
 with zipfile.ZipFile(path) as z:files={n:z.read(n) for n in z.namelist()}
 root=E.fromstring(files['xl/workbook.xml']);ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
 for s in root.find('m:sheets',ns):
  if s.get('name') in ['Valuation output','Review valuation']:s.set('state','hidden')
 files['xl/workbook.xml']=E.tostring(root,xml_declaration=True,encoding='UTF-8')
 with zipfile.ZipFile(path,'w',zipfile.ZIP_DEFLATED) as z:
  for n,b in files.items():z.writestr(n,b)
 return path

def main():
 OUT.mkdir(exist_ok=True);cal=[]
 if '--actual-only' not in sys.argv:
  cases=[('reference',TASK/'metadata/reference_judge_smoke.json',[],['R001','R002','R003','R006']),('equivalent_transposed',FIX/'validation/equivalent_transposed/result.json',[],['R001','R002','R003','R006']),('wrong_terminal_period',FIX/'validation/wrong_terminal_period/result.json',['R003','R006'],['R001','R002']),('mixed_secondary_summary',FIX/'validation/mixed_secondary_summary/result.json',['R006'],['R001','R002','R003']),('duplicate_year',FIX/'validation/duplicate_year/result.json',['R002','R003'],['R001','R006'])]
  for name,receipt,loss,keep in cases:
   old=json.loads(receipt.read_text());native=old.get('native_receipt') or old['evidence']['native_baseline'];facts,details=snapshot_facts(read(native['output']),True)
   for key in loss:assert facts[key]<1,(name,key,facts)
   for key in keep:assert facts[key]==1,(name,key,facts)
   cal.append({'case':name,'facts':facts,'asserted_loss':loss,'asserted_keep':keep,'native_output_reused':native['output']})
  adjacent=build_adjacent();result=evaluate(adjacent,TASK/'data/input_files',OUT/'equivalent_adjacent')
  (OUT/'equivalent_adjacent_result.json').write_text(json.dumps(result,indent=2,default=str))
  assert result['evaluation_status']=='SCORED' and result['normalized_score']==1,result
  assert {t['case'] for t in result['evidence']['candidate_tables']}=={'base','review'}
  cal.append({'case':'equivalent_adjacent_cases','score':result['normalized_score'],'facts':result['criterion_scores'],'native_recalculations':4})
  (OUT/'calibration.json').write_text(json.dumps({'status':'PASS','checks':cal,'scope':'Affected reader snapshots plus new adjacent-case full native dynamic equivalence'},indent=2))
 answer=TRIAL/'artifacts/app/output/answer.xlsx';before=hashlib.sha256(answer.read_bytes()).hexdigest()
 original=json.loads((TASK/'metadata/first_attempt.json').read_text());assert before==original['output_sha256']
 actual=evaluate(answer,TRIAL/'artifacts/app/input',OUT/'actual_recalc')
 after=hashlib.sha256(answer.read_bytes()).hexdigest();assert before==after
 (OUT/'result.json').write_text(json.dumps(actual,indent=2,default=str))
 consistency={}
 if actual.get('evidence',{}).get('native_baseline'):
  fresh=Path(actual['evidence']['native_baseline']['output']);w=openpyxl.load_workbook(fresh,data_only=True);f=openpyxl.load_workbook(answer,data_only=False)
  c15=f['Operating lease converter']['C15'];f5=w['Synthetic rating']['F5'].value;implied=w['Input sheet']['B9'].value+w['Operating lease converter']['F32'].value
  consistency={'candidate_C15_value':c15.value,'candidate_C15_is_formula':c15.data_type=='f','native_synthetic_ebit':f5,'native_formula_implied_ebit':implied,'residual':f5-implied,'consistent':abs(f5-implied)<1e-7}
  assert c15.value==.0413 and consistency['consistent'],consistency
 receipt={'status':'REJUDGED_SAME_ORIGINAL','judge_version':'new6-a1-v1.2-adjacent-cases','old_status':original['original_evaluation_status'],'old_score':original['original_score'],'same_original_sha256':before,'original_unchanged':True,'frozen_wrapper_unchanged':True,'calibration':'calibration.json','result':'result.json','actual_status':actual['evaluation_status'],'profiles':{k:{f:v[f] for f in ['evaluation_status','normalized_score','score_decimal','pass']} for k,v in actual['profiles'].items()},'criterion_scores':actual.get('criterion_scores'),'historical_engine_consistency':consistency,'agent_calls':0,'weight_or_fact_changes':False}
 (OUT/'receipt.json').write_text(json.dumps(receipt,indent=2));print(json.dumps(receipt,indent=2))
if __name__=='__main__':main()
