"""Source-labelled business facts; no gold-value search or fixed output coordinates."""
from pathlib import Path
from decimal import Decimal
from collections import defaultdict
import argparse,json,re,sys,math
import openpyxl
ROOT=Path(__file__).resolve().parents[1]
sys.path[:0]=[str(ROOT/'metadata'),str(ROOT.parents[1]/'common')]
from oracle_recompute import source,compute
from ooxml_edit import edit
from runtime import score_profiles,output_status,recalculate_xlsx,RecalcUnavailable
S=source();KINDS=['elements','summary','provisional','exclusions','qualifications']
norm=lambda x:re.sub(r'[^a-z0-9]','',str(x or '').lower())
LOOK={norm(v):(kind,r['id']) for kind in KINDS for r in S[kind] for v in [r['label']]+r['aliases']}
def number(x):
 if isinstance(x,str):
  x=x.strip().replace(',','').replace('£','')
  if x.endswith('%'):return float(x[:-1])/100
 return float(x)
def equal(a,b):
 try:return math.isfinite(number(a)) and abs(number(a)-number(b))<=.000001
 except (TypeError,ValueError):return False
HEADERS={
 'label':['item','stage','element','description','costitem','qualification','exclusion','allowance'],
 'printed':['printedamount','publishedamount','sourceamount','originalamount','printedcost','publishedcost','printedgbp'],
 'rate':['printedrate','publishedrate','originalrate','total','percenttotal','sourcepercentage'],
 'working':['workingamount','workingcost','currentamount','adjustedamount','workinggbp','calculatedamount'],
 'working_rate':['workingrate','allowancepercentage','adjustedrate','currentrate'],
 'delta':['difference','variance','reconciliation','workingminusprinted','printedminusworking'],
 'scope':['scope','status','treatment','meaning','qualification','notes','inclusion'],
 'source':['source','reference','sourcereference','document'],
 'page':['page','pdfpage','sourcepage']}
HEADERS['label']+=['elementdescription','stageitem','costcomponent']
HEADERS['printed']+=['original','originalcost','amount','amountgbp','cost','costgbp','sourcecost','published']
HEADERS['working']+=['adjusted','adjustedcost','workbook','current','reviewedcost','calculated']
HEADERS['working_rate']+=['ratefactor']
HEADERS['scope']+=['descriptionnotes','notesprovisionalsums','notessource']
HEADERS['adjustment_factor']=['adjfactor','adjustmentfactor']
HEADERS['adjustment_percent']=['adjustment']
H={v:k for k,vs in HEADERS.items() for v in vs}
ALIASES={
 'total building works':('summary','building'),
 'building works total':('summary','building'),
 'total building works estimate':('summary','building_estimate'),
 'maincontractorsohp':('summary','overheads'),
 'heatingashp':('provisional','heating'),
 'works costs involved with onerous planning conditions':('exclusions','excluded_h'),
 'increased costs foreign trade tariffs':('exclusions','excluded_n'),
 'increased costs conflict impacts':('exclusions','excluded_o'),
 'value added tax':('exclusions','excluded_p'),
 'dry rot treatment works':('exclusions','excluded_q'),
 'fire safety systems':('exclusions','excluded_services_1')}
ALIASES={norm(k):v for k,v in ALIASES.items()}
def header_role(value):
 n=norm(value)
 if n in ['oftotal','total'] and '%' in str(value):return 'display_share'
 if n in ['rate','ratepercent']:return 'generic_rate'
 if '%' in str(value):
  if n in ['original','printed','source']:return 'rate'
  if n in ['adjusted','working','current']:return 'working_rate'
  if n=='variance':return None
 if n=='source' and re.search(r'£|\bGBP\b',str(value),re.I):return 'printed'
 return H.get(n)
def bind_label(v):
 v=re.sub(r'^\s*(?:[•–-]|\d+[.)]|[a-z][)])\s*','',str(v),flags=re.I)
 n=norm(v);n=re.sub(r'^\d+','',n)
 if n in LOOK:return LOOK[n]
 if n in ALIASES:return ALIASES[n]
 # A readable abbreviated source label is sufficient only when its identity
 # remains unique; numerical values never select the identity.
 prefixes={key for token,key in LOOK.items() if len(token)>=15 and n.startswith(token)}
 if len(prefixes)==1:return next(iter(prefixes))
 hits={key for token,key in LOOK.items() if len(token)>22 and n.startswith(token[:22])}
 return next(iter(hits)) if len(hits)==1 else None
def note_records(w,found):
 """Read delivered list statements without inventing a missing source amount."""
 for sh in w:
  paras=[]
  for row in sh.iter_rows(max_row=min(sh.max_row,500),max_col=min(sh.max_column,60)):
   vals=[str(c.value) for c in row if c.value is not None]
   if vals:paras.append((row[0].row,' '.join(vals)))
  for pos,(rid,line) in enumerate(paras):
   clean=re.sub(r'^\s*(?:[•–-]|\d+[.)]|[a-z][)])\s*','',line,flags=re.I)
   vicinity=' '.join(t for rr,t in paras[max(0,pos-1):pos+2])
   context=' '.join(t for rr,t in paras[:3])
   # One clearly excluded list can enumerate several service objects.
   if re.search(r'exclud|exclus',vicinity+' '+context,re.I):
    for r in S['exclusions']:
     tokens=[norm(r['label'])]+[norm(x) for x in r['aliases']]
     hits=any(token in norm(clean) for token in tokens if len(token)>=8)
     if not hits:
      special={'excluded_h':r'onerous planning conditions','excluded_n':r'foreign trade tariff','excluded_o':r'conflict.*Ukraine.*Middle East','excluded_q':r'dry rot treatment'}
      hits=bool(r['id'] in special and re.search(special[r['id']],clean,re.I))
     key=('exclusions',r['id'])
     if hits and not any(z['sheet']==sh.title and z['row']==rid for z in found[key]):
      found[key].append({'sheet':sh.title,'row':rid,'table':0,'label':clean,'context':context,'scope':vicinity+' '+context,'cells':{},'delta_sign':1})
   # Bullet prose is also a legitimate representation of a provisional allowance.
   money=re.search(r'£\s*([\d,]+(?:\.\d+)?)([kKmM])?',clean)
   if money and re.search(r'provisional|included within elemental',context+' '+' '.join(t for rr,t in paras[:pos]),re.I):
    label=clean[:money.start()].rstrip(': @ ')
    key=bind_label(label)
    if not key and re.search(r'external doors',label,re.I):key=('provisional','external_doors')
    if not key and re.search(r'scaffolding',label,re.I):key=('provisional','scaffolding')
    if key and key[0]=='provisional' and not any(z['sheet']==sh.title and z['row']==rid for z in found[key]):
     amount=number(money.group(1))*({'k':1000,'m':1000000}.get((money.group(2) or '').lower(),1))
     found[key].append({'sheet':sh.title,'row':rid,'table':0,'label':label,'printed':amount,'scope':' '.join(t for rr,t in paras[:pos])+clean,'context':context,'cells':{'printed':None},'delta_sign':1})
def discover(w,proven_reference_failure=False,proven_arithmetic_failure=False):
 found=defaultdict(list);unbound=[];tables=[];material_unparsed=[]
 text=' '.join(str(c.value) for s in w for row in s.iter_rows(max_row=min(s.max_row,400),max_col=min(s.max_column,60)) for c in row if c.value is not None)
 for s in w:
  if s.max_row>10000 or s.max_column>500:raise ValueError('Material workbook exceeds bounded reader')
  context=' '.join(str(c.value) for row in s.iter_rows(max_row=min(s.max_row,3),max_col=min(s.max_column,30)) for c in row if c.value is not None)
  current=None;sheet_matches=0;table_row=None;section_context=context;self_check={}
  for row in s.iter_rows(max_row=min(s.max_row,500),max_col=min(s.max_column,60)):
   header={header_role(c.value):c.column for c in row if header_role(c.value)}
   if 'label' in header and len(header)>=2:
    if 'generic_rate' in header:header['working_rate' if 'working' in header else 'rate']=header.pop('generic_rate')
    if 'display_share' in header:
     col=header.pop('display_share');header['rate' if 'working' not in header or col<header['working'] else 'working_share']=col
    current=header;table_row=row[0].row;tables.append((s.title,table_row));continue
   values=[c for c in row if c.value is not None]
   if len(values)==1 and isinstance(values[0].value,str) and (str(values[0].value).rstrip().endswith(':') or re.fullmatch(r'SUMMARY COST CHAIN|ELEMENTAL COST BUILD.UP',str(values[0].value),re.I)):
    section_context=str(values[0].value);table_row=row[0].row;continue
   # Exclusion bullets are valid source statements, even outside a formal table.
   if len(values)==1 and isinstance(values[0].value,str):
    bullet_key=bind_label(values[0].value)
    if bullet_key and bullet_key[0]=='exclusions' and re.search(r'exclu',s.title+' '+section_context,re.I):
     found[bullet_key].append({'sheet':s.title,'row':row[0].row,'table':table_row,'label':str(values[0].value),'context':context+' '+section_context,'scope':section_context,'cells':{},'delta_sign':1});sheet_matches+=1;continue
   if not current:continue
   label=s.cell(row[0].row,current['label']).value
   if label is None:continue
   diagnostic=norm(label)
   if diagnostic in ['expectedtotalfrompdf','calculatedtotal','difference'] and re.search(r'reconciliation check',section_context,re.I):
    nums=[c for c in row if c.column!=current['label'] and c.value is not None]
    if len(nums)==1:self_check[diagnostic]={'value':nums[0].value,'cell':nums[0].coordinate}
    else:raise ValueError('Ambiguous labelled self-check: '+str(label))
    continue
   rate_row='rate' in norm(label) and ('adjustment_factor' in current or 'working_rate' in current)
   rate_label=re.sub(r'\s+Rate\b','',str(label),flags=re.I)
   if rate_row:key=bind_label(rate_label)
   else:key=bind_label(label)
   if not key:
    if any(isinstance(s.cell(row[0].row,col).value,(int,float)) for name,col in current.items() if name in ['printed','working']) and not re.search(r'total|sum|check',str(label),re.I):unbound.append({'sheet':s.title,'label':str(label)})
    continue
   rec={'sheet':s.title,'row':row[0].row,'table':table_row,'label':str(label),'context':context+' '+section_context,'cells':{}}
   for name,col in current.items():rec[name]=s.cell(row[0].row,col).value;rec['cells'][name]=s.cell(row[0].row,col).coordinate
   label_rate=re.search(r'\(\s*(\d+(?:\.\d+)?)\s*%\s*\)',str(label))
   if label_rate and key[0]=='summary' and rec.get('rate') is None:
    rec['rate']=number(label_rate.group(1))/100;rec['cells']['rate']=None;rec['printed_rate_label']=str(label)
   if rate_row and key[0]=='summary':
    rec['_role']='rate_control'
    if 'printed' in rec:rec['rate']=rec.pop('printed');rec['cells']['rate']=rec['cells'].pop('printed')
    rate_col=current.get('adjustment_factor',current.get('working_rate'))
    rec['working_rate']=s.cell(row[0].row,rate_col).value;rec['cells']['working_rate']=s.cell(row[0].row,rate_col).coordinate
    for f in ['working','adjustment_factor']:
     rec.pop(f,None);rec['cells'].pop(f,None)
   if 'printed' not in current:
    amounts=[c for c in row if norm(s.cell(tables[-1][1],c.column).value) in ['amount','cost','amountgbp','costgbp']]
    if amounts:rec['printed']=amounts[0].value;rec['cells']['printed']=amounts[0].coordinate
   rec['delta_sign']=-1 if current.get('delta') and norm(s.cell(tables[-1][1],current['delta']).value)=='printedminusworking' else 1
   if key[0]=='exclusions' and re.search(r'exclu',s.title+' '+section_context,re.I):rec['scope']=str(rec.get('scope',''))+' '+section_context
   found[key].append(rec);sheet_matches+=1
  if sheet_matches==0 and sum(isinstance(c.value,(int,float)) for row in s.iter_rows(max_row=min(s.max_row,500),max_col=min(s.max_column,60)) for c in row)>10:material_unparsed.append(s.title)
  if self_check:
   if set(self_check)!=set(['expectedtotalfrompdf','calculatedtotal','difference']):raise ValueError('Incomplete self-check layout: '+s.title)
   f=self_check;values={k:v['value'] for k,v in f.items()}
   found[('_self_checks',s.title)].append({'sheet':s.title,'row':0,'table':0,'label':'Reconciliation Check','cells':{},'values':values,'arithmetic_consistent':equal(values['calculatedtotal']-values['expectedtotalfrompdf'],values['difference']) if all(isinstance(x,(int,float)) for x in values.values()) else False,'agrees_with_expected':equal(values['calculatedtotal'],values['expectedtotalfrompdf'])})
 if unbound:raise ValueError('Material result labels cannot be safely bound: '+str(unbound[:3]))
 if material_unparsed and any((kind,r['id']) not in found for kind in KINDS for r in S[kind]):raise ValueError('A material candidate table remains unbound: '+str(material_unparsed))
 if not found and sum(isinstance(c.value,(int,float)) for s in w for row in s for c in row)>10:raise ValueError('Material numeric result layout cannot be safely bound')
 # Qualifications may be ordinary briefing paragraphs. Bind by explicit concepts,
 # not by looking for an expected monetary answer or requiring a private table.
 note_records(w,found)
 concepts={'scope_basis':r'works.*(?:envisaged|drawings)', 'phase':r'single.phase', 'duration':r'contract duration|duration of the contract', 'price_basis':r'(?:costs|rates|price basis).*(?:4Q24|Q4.?2024)', 'guide':r'guide.*(?:probable|cost)', 'qs':r'consultation.*quantity surveyor|consult(?:ed|ing)?\s+(?:with\s+)?(?:the\s+)?quantity surveyor', 'drawings':r'pre.?planning drawings'}
 for s in w:
  paras=[]
  for row in s.iter_rows(max_row=min(s.max_row,500),max_col=min(s.max_column,60)):
   vals=[str(c.value) for c in row if isinstance(c.value,str)]
   if vals:paras.append((row[0].row,' '.join(vals)))
  for id,pattern in concepts.items():
   if ('qualifications',id) in found:continue
   hits=[(i,t) for i,t in paras if re.search(pattern,t,re.I)]
   if hits:
    i,t=hits[0];near=' '.join(t for ri,t in paras if i<=ri<=i+2)
    found[('qualifications',id)].append({'sheet':s.title,'row':i,'table':i,'label':t,'context':near,'scope':near,'cells':{},'delta_sign':1})
 # A named source PDF plus the delivered unique business label is a readable
 # locator. The instruction does not require repeating a private page coordinate.
 source_name=S['elements'][0]['source']
 if source_name.lower() in text.lower() or (re.search(r'Falmouth',text,re.I) and re.search(r'October\s*2024',text,re.I) and re.search(r'(?:cost estimate|estimate).*revision\s*A',text,re.I)):
  for rs in found.values():
   for r in rs:r['document_locator']=source_name+' :: '+r['label']
 pending_errors=['#N/A']+([] if proven_reference_failure else ['#NAME?'])+([] if proven_arithmetic_failure else ['#VALUE!'])
 errors=[(key,r.get(field)) for key,rs in found.items() for r in rs for field in ['printed','working','working_rate','delta'] if r.get(field) in pending_errors]
 if errors:raise RecalcUnavailable('Bound required quantity has an engine/parse error: '+str(errors[:3]))
 return found,text
def records(f,k,id):return f.get((k,id),[])
def unique_records(rs):
 # Repeating a source item in a separately labelled reconciliation view is legal;
 # repeated occurrences within one table remain a duplicate. Values in every view
 # are still checked, so a contradictory final representation never disappears.
 relevant=[r for r in rs if r.get('_role')!='rate_control']
 return bool(relevant) and len({(r['sheet'],r.get('table')) for r in relevant})==len(relevant)
def numeric(f,k,id,field,value):
 rs=records(f,k,id);actual=[r for r in rs if field in r['cells'] and r.get(field) is not None];return bool(actual) and all(equal(r.get(field),value) for r in actual)
def page_ok(r,page):
 v=str(r.get('page',''))
 explicit=bool(re.search(r'(?<!\d)'+str(page)+r'(?!\d)',v)) or bool(re.search(r'(?:page|p\.?)[\s.:]*'+str(page)+r'\b',r['context'],re.I))
 # An explicit wrong page must not be rescued by a generic document citation.
 return explicit or (not v.strip() and bool(r.get('document_locator')))
def scope_text(rs):return ' '.join(str(r.get('scope',''))+' '+r.get('context','') for r in rs).lower()
QUAL={'guide':[r'guide|probable',r'cost|estimate'],'qs':[r'quantity surveyor',r'consult'],'drawings':[r'RTP',r'pre.?planning',r'no input|no other|another designer'],'scope_basis':[r'drawing',r'specification|envisaged'],'phase':[r'single phase',r'local',r'medium'],'duration':[r'45',r'week'],'price_basis':[r'4Q24|Q4.?2024',r'1Q25|Q1.?2025',r'BCIS',r'Cornwall']}
def checks(w,proven_reference_failure=False,proven_arithmetic_failure=False):
 f,text=discover(w,proven_reference_failure,proven_arithmetic_failure);o=compute();detail={k:[] for k in ['R001','R002','R003','R004','R005','R006']}
 def add(cid,id,ok,actual=None,expected=None):detail[cid].append(dict(id=id,ok=bool(ok),actual=actual,expected=str(expected) if expected is not None else None))
 for id,pattern in {'document':r'cost estimate|cost.review|order.of.cost|\bOCE\b','revision':r'revision\s*A\b|rev\s*A\b|reva','date':r'October.?2024|Oct.?24|2024.?10','project':r'Falmouth','currency':r'GBP|£|pounds sterling'}.items():add('R001',id,bool(re.search(pattern,text,re.I)))
 for kind in KINDS:
  for r in S[kind]:
   rs=records(f,kind,r['id']);unique=unique_records(rs)
   if kind in ['elements','summary','provisional']:
    add('R002',kind+':'+r['id']+':unique',unique,len([z for z in rs if z.get('_role')!='rate_control']),1)
    if isinstance(r['amount'],(int,float)):add('R002',kind+':'+r['id']+':printed',numeric(f,kind,r['id'],'printed',r['amount']),[z.get('printed') for z in rs],r['amount'])
    if r['rate'] is not None:add('R002',kind+':'+r['id']+':printed_rate',numeric(f,kind,r['id'],'rate',r['rate']),[z.get('rate') for z in rs],r['rate'])
   if kind=='provisional':add('R003',r['id']+':included_memorandum',unique and 'includ' in scope_text(rs) and ('memorandum' in scope_text(rs) or re.search(r'\b(?:in|within)\s+(?:the\s+)?elemental\s+cost',scope_text(rs))) and not any(re.search(r'additional charge|add to total|not included',str(z.get('scope','')),re.I) for z in rs))
   if kind=='exclusions':add('R003',r['id']+':excluded',unique and all(re.search(r'exclud|exclus',str(z.get('scope','')).lower()+' '+str(z.get('printed','')).lower()) and not isinstance(z.get('printed'),(int,float)) for z in rs))
   if kind=='qualifications':add('R003',r['id']+':qualification',unique and all(re.search(p,scope_text(rs),re.I) for p in QUAL[r['id']]))
   if r['id'] in ['listing','construction_risk','employer_risk','consultants']:
    add('R003',r['id']+':source_unpriced',unique and all(not isinstance(z.get('printed'),(int,float)) and (str(z.get('printed','')).strip() in ['-','—','–',''] or re.search(r'exclud|unpriced|not priced',str(z.get('printed','')),re.I)) for z in rs))
   source_ok=bool(re.search(r'Falmouth|Passmore Edwards',text,re.I)) and bool(re.search(r'\.pdf|October.?2024|Oct.?24',text,re.I))
   add('R006',kind+':'+r['id']+':source_page',unique and source_ok and all(page_ok(z,r['page']) for z in rs))
 for id,value in o['elements'].items():add('R004','element:'+id,numeric(f,'elements',id,'working',value),expected=value)
 for id,value in o['working'].items():
  add('R004','working:'+id,numeric(f,'summary',id,'working',value),expected=value)
  rs=[z for z in records(f,'summary',id) if 'delta' in z['cells']]
  self_ok=all(z['arithmetic_consistent'] for key,zs in f.items() if key[0]=='_self_checks' for z in zs) if id=='building' else True
  add('R004','reconciliation:'+id,bool(rs) and self_ok and all(equal(z.get('delta'),o['reconciliation'][id]*z['delta_sign']) for z in rs),expected=o['reconciliation'][id])
 for id,value in o['rates'].items():add('R004','working_rate:'+id,numeric(f,'summary',id,'working_rate',value),expected=value)
 return detail,f
def locate(f,kind,id,field):
 rs=records(f,kind,id);rs=[r for r in rs if field in r['cells']]
 return (rs[0]['sheet'],rs[0]['cells'][field]) if rs and rs[0]['cells'][field] else None
def edit_control(f,raw,id,value):
 if id.endswith('_rate'):
  return locate(f,'summary',id.removesuffix('_rate'),'working_rate'),float(value)
 for r in records(f,'elements',id):
  for role in ['adjustment_factor','adjustment_percent']:
   co=r['cells'].get(role)
   if co and raw[r['sheet']][co].data_type!='f':
    try:factor=number(value)/number(r.get('printed'))
    except (TypeError,ValueError,ZeroDivisionError):continue
    return (r['sheet'],co),factor-1 if role=='adjustment_percent' else factor
 where=locate(f,'elements',id,'working')
 if where and raw[where[0]][where[1]].data_type=='f':return None,float(value)
 return where,float(value)
def broken_sheet_references(w):
 """Prove ordinary formula references to absent candidate sheets, without evaluation.

 Only SUM/IF arithmetic is admitted for this narrow diagnosis. Any additional
 function leaves #NAME? pending until engine support has been checked separately.
 """
 missing=[];functions=set()
 for sh in w:
  for row in sh:
   for c in row:
    if c.data_type!='f':continue
    try:tokens=openpyxl.formula.Tokenizer(c.value).items
    except Exception:return [],False
    for token in tokens:
     if token.type=='FUNC' and token.subtype=='OPEN':functions.add(token.value[:-1].upper())
     if token.type=='OPERAND' and token.subtype=='RANGE' and '!' in token.value:
      name=token.value.rsplit('!',1)[0].strip("'").replace("''", "'")
      if '[' not in name and ':' not in name and name not in w.sheetnames:
       missing.append({'sheet':sh.title,'cell':c.coordinate,'formula':c.value,'absent_sheet':name})
 return missing,bool(missing) and functions.issubset({'SUM','IF'})
def invalid_text_arithmetic(raw,cached):
 """Identify a displayed ordinary subtraction applied to a delivered text value."""
 functions=set();bad=[]
 for sh in raw:
  for row in sh:
   for c in row:
    if c.data_type!='f':continue
    try:tokens=openpyxl.formula.Tokenizer(c.value).items
    except Exception:return [],False
    functions.update(t.value[:-1].upper() for t in tokens if t.type=='FUNC' and t.subtype=='OPEN')
    m=re.fullmatch(r'=\s*(\$?[A-Z]+\$?\d+)\s*-\s*(\$?[A-Z]+\$?\d+)\s*',c.value,re.I)
    if m and cached[sh.title][c.coordinate].value=='#VALUE!':
     a,b=[cached[sh.title][co.replace('$','')].value for co in m.groups()]
     if any(isinstance(v,str) and not v.startswith('#') and not re.fullmatch(r'-?\d+(?:\.\d+)?',v.strip()) for v in [a,b]):bad.append({'sheet':sh.title,'cell':c.coordinate,'formula':c.value,'operands':[a,b],'cached_error':'#VALUE!'})
 return bad,bool(bad) and functions.issubset({'SUM','IF','ABS'})
def evaluate(path,evidence_dir,completed_run=True):
 out=Path(evidence_dir);out.mkdir(parents=True,exist_ok=True);status=output_status(path)
 if status:
  result=score_profiles(ROOT/'rubric.json',status=status if completed_run else 'INFRA_ERROR',evidence={'completed_run':completed_run})
 else:
  try:
   raw=openpyxl.load_workbook(path,data_only=False)
   if any(c.data_type=='f' and re.search(r'LAMBDA\(|_xlfn\.PY\(',str(c.value),re.I) for s in raw for row in s for c in row):raise RecalcUnavailable('Legal formula feature is unsupported by the native adapter')
   broken_references,proven_reference_failure=broken_sheet_references(raw)
   fresh,receipt=recalculate_xlsx(path,out/'base');w=openpyxl.load_workbook(fresh,data_only=True)
   arithmetic_errors,proven_arithmetic_failure=invalid_text_arithmetic(raw,w)
   details,before=checks(w,proven_reference_failure,proven_arithmetic_failure);orig=compute();probes=[];printed_invariants=[]
   for name,changes in [('roof',{'pitched_roof':428900}),('risk',{'design_risk_rate':'.12'}),('joint',{'services':639250,'overheads_rate':'.12','inflation_rate':'.015'})]:
    patch=defaultdict(dict);bound=True
    for id,value in changes.items():
     where,actual_control_value=edit_control(before,raw,id,value)
     if where is None:bound=False;continue
     patch[where[0]][where[1]]=actual_control_value
    new=compute(changes);af={};rec=None
    if bound:
     dest=out/name/'mutated.xlsx';edit(path,dest,patches=patch,clear_caches=True);fresh2,rec=recalculate_xlsx(dest,out/name/'recalc');af,_=discover(openpyxl.load_workbook(fresh2,data_only=True),proven_reference_failure,proven_arithmetic_failure)
    for id,value in new['working'].items():
     delta=value-orig['working'][id]
     if delta==0:continue
     br=[r for r in records(before,'summary',id) if 'working' in r['cells']];ar=[r for r in records(af,'summary',id) if 'working' in r['cells']];ok=bool(br) and len(br)==len(ar) and all(isinstance(a.get('working'),(int,float)) and isinstance(b.get('working'),(int,float)) and equal(b['working']-a['working'],delta) for a,b in zip(br,ar))
     details['R005'].append(dict(id=name+':'+id,ok=ok,expected_delta=str(delta)))
    if bound:
     good=True
     for kind in ['elements','summary','provisional']:
      for r in S[kind]:
       br=records(before,kind,r['id']);ar=records(af,kind,r['id'])
       for field in ['printed','rate']:
        good=good and len(br)==len(ar) and all(a.get(field)==b.get(field) for a,b in zip(br,ar))
     printed_invariants.append(good)
    else:printed_invariants.append(True)
    probes.append(dict(name=name,changes=changes,actual_control_edits=dict(patch),bound=bound,native_receipt=rec))
   details['R002'].append(dict(id='original_figures_preserved_during_edits',ok=all(printed_invariants)))
   scores={k:str(Decimal(sum(z['ok'] for z in xs))/Decimal(len(xs))) for k,xs in details.items()}
   result=score_profiles(ROOT/'rubric.json',scores,evidence={'judge_version':'new6-c1-facts-v1.2-controls-and-notes','base_native_receipt':receipt,'dynamic_probes':probes,'fact_units':details,'candidate_self_checks':[z for key,rs in before.items() if key[0]=='_self_checks' for z in rs],'candidate_broken_sheet_references':broken_references,'candidate_invalid_text_arithmetic':arithmetic_errors,'proven_ordinary_formula_reference_failure':proven_reference_failure,'parser':'Binds visible source labels and role headers, regardless of sheet/row/column. Separate reconciliation views are checked together; explicit factor/percentage inputs and note allowances are supported. Unsupported material layouts and legal formula limits are pending, never zero.','scope':'15 element rows;12 summary rows;13 included memorandum rows;24 exclusions;7 qualifications. Dynamic denominator uses independent nonzero expected working-stage changes.'})
  except (RecalcUnavailable,ValueError,TypeError,KeyError,openpyxl.utils.exceptions.InvalidFileException) as exc:result=score_profiles(ROOT/'rubric.json',status='JUDGE_ERROR',evidence={'error_type':type(exc).__name__,'error':str(exc)})
 (out/'evaluation.json').write_text(json.dumps(result,ensure_ascii=False,indent=2,default=str));return result
if __name__=='__main__':
 p=argparse.ArgumentParser();p.add_argument('answer',nargs='?',default='/app/output/answer.xlsx');p.add_argument('--evidence-dir',default='/tmp/new6-c1-evidence');p.add_argument('--input-dir');p.add_argument('--completed-run',action='store_true');a=p.parse_args();print(json.dumps(evaluate(a.answer,a.evidence_dir,a.completed_run),default=str))
