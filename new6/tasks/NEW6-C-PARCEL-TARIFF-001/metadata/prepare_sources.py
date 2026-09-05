"""One bounded source preparation: PDF vs official XLSX, then declared excerpt."""
from pathlib import Path
from decimal import Decimal as D
import json,re,csv,hashlib,sys
import pdfplumber
from pypdf import PdfReader,PdfWriter
from openpyxl import load_workbook
from oracle_recompute import compute

ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT.parents[1]/'sources/downloads_c'
PDF=SRC/'usps_notice123_effective20260712.pdf'
XLSX=SRC/'usps_july2026_official_rates.xlsx'
URL='https://pe.usps.com/resources/PriceChange/July%202026%20Price%20Change%20-%20Notice123%20PDF%20-%207.10.26.pdf'
XURL='https://pe.usps.com/resources/PriceChange/July%202026%20Price%20Change%20-%2007082026%20-%20Notice%20123.xlsx'

def main():
    rates=[]; w=load_workbook(XLSX,data_only=True)
    with pdfplumber.open(PDF) as p:
        assert len(p.pages)==63 and 'Effective July 12, 2026' in p.pages[0].extract_text()
        for service,index,sheet in [('priority',4,'PM Retail'),('ground',6,'USPS Ground Advantage Retail')]:
            lines=p.pages[index].extract_text().splitlines();count=0
            for line in lines:
                m=re.match(r'^(\d+(?:\.\d+)?)( oz)?\s+(\$?\d+\.\d{2})',line)
                if not m:continue
                band=m[1];unit='oz' if m[2] else 'lb'
                if unit=='lb' and D(band)>10:continue
                prices=re.findall(r'\$?(\d+\.\d{2})',line[m.end(2) if m[2] else len(m[1]):])[:8]
                assert len(prices)==8,(line,prices)
                official_row=(13+int(band)) if service=='priority' else ({'4':6,'8':7,'12':8,'15.999':9}[band] if unit=='oz' else 10+int(band))
                for zone,price in enumerate(prices,1):
                    assert D(price)==D(str(w[sheet].cell(official_row,zone+1).value)),(service,band,zone,price)
                    rates.append({'service':service,'upper_bound':band,'weight_unit':unit,'upper_oz':str(D(band)*(16 if unit=='lb' else 1)),'zone':zone,'usd':price,'effective_date':'2026-07-12','source_page':index+1,'source_object':f'{service} retail, weight not over {band} {unit}, zone {zone}','official_xlsx_cell':f'{sheet}!{w[sheet].cell(official_row,zone+1).coordinate}'})
                    count+=1
            assert count==(80 if service=='priority' else 112),(service,count)
    (ROOT/'metadata/rates.json').write_text(json.dumps(rates,indent=2))
    requests=list(csv.DictReader((ROOT/'data/input_files/quote_requests.csv').open()))
    (ROOT/'metadata/oracle_expected.json').write_text(json.dumps(compute(rates,requests),indent=2))
    writer=PdfWriter();reader=PdfReader(PDF)
    for page in [0,4,6]:writer.add_page(reader.pages[page])
    excerpt=ROOT/'data/input_files/USPS_Notice123_20260712_pages1_5_7.pdf';writer.write(excerpt)
    manifest={'source_id':'usps_july2026','source_type':'official_rate_schedule_with_project_authored_quote_requests','source_url':URL,'source_file':PDF.name,'effective_date':'2026-07-12','publication_status':'Final','pdf_pages':63,'download_sha256':'3fbb08717d4ef00cd14fd5b447564a2699d841ecf82bd90c1a5f0f31c16dd9ec','private_machine_url':XURL,'private_machine_sha256':'c0a0ab9d5c2b98786063cb95b7b8a98b5046aca40c78eb93c8d5e979a28c8ba2','machine_internal_date':'7/8/2026','machine_status':'Final','source_checks':{'pdf_xlsx_price_parity':192,'rendered_visual_pages':[5,7],'visual_checks':['Priority Mail pounds, inclusive not-over header, zones 1–8 belong to left table; zone 9 excluded','Ground Advantage four ounce bands, then pounds; footnote 9 states pounds unless indicated','All footnotes retained on complete pages; declared carton conditions exclude all applicable add-ons']},'input_transformation':{'method':'page extraction only; source content unchanged','physical_pages':[1,5,7],'printed_pages':[1,5,7],'excerpt_sha256':hashlib.sha256(excerpt.read_bytes()).hexdigest(),'source_rates_csv_or_xlsx_mounted':False},'rights':{'source':'USPS publicly published rate schedule','local_use':'local reconstruction and evaluation of published factual rates with attribution','redistribution_clearance':'not independently cleared; no external publication made'},'scope':{'price_facts':192,'services':['priority','ground'],'zones':[1,2,3,4,5,6,7,8],'quote_request_count':12,'quote_request_origin':'project_authored_test_request'}}
    (ROOT/'metadata/source_manifest.json').write_text(json.dumps(manifest,indent=2,ensure_ascii=False))
    print(json.dumps({'verified_price_facts':len(rates),'oracle':compute(rates,requests)},indent=2))
if __name__=='__main__':main()
