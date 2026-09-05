"""Conservative label/identity reader, with actual OOXML chart evidence.

This reader discovers headers and locations from candidate contents. It does not
use reference cell coordinates or interpret candidate formulas. Uncached formulas
and chart reference forms outside the supported reader have explicit pending states.
"""
from pathlib import Path
from decimal import Decimal, InvalidOperation
from collections import Counter
import re, zipfile, xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.cell import range_boundaries,get_column_letter

class ParsePending(Exception):pass
def norm(v):return re.sub(r'[^a-z0-9]+','',str(v or '').lower())
ALIASES={
 'code':['geographycode','areacode','localcode','authoritycode','code'],
 'name':['geography','authority','areaname','geographyname','localauthority'],
 'employment_old':['employmentrateearlier','employmentrate2024','employment2024','earlieremploymentrate'],
 'employment_new':['employmentratelater','employmentrate2025','employment2025','lateremploymentrate'],
 'unemployment_old':['unemploymentrateearlier','unemploymentrate2024','unemployment2024','earlierunemploymentrate'],
 'unemployment_new':['unemploymentratelater','unemploymentrate2025','unemployment2025','laterunemploymentrate'],
 'employment_change':['employmentchangepp','employmentratechangepp','employmentchange','employmentppchange'],
 'unemployment_change':['unemploymentchangepp','unemploymentratechangepp','unemploymentchange','unemploymentppchange'],
 'comparison_status':['comparisonstatus','comparability','pairstatus'],
 'metric':['metric','indicator','measure','field','item'], 'value':['value','reportedvalue','result','finding'],
 'reason':['reason','exclusionreason','treatmentreason','exceptionreason'],
 'rank':['rank','position'], 'edition':['edition','release'],
 'raw':['rawvalue','storedvalue','underlyingvalue'], 'published':['publishedvalue','displayedrate','publishedrate'],
 'source':['sourcelocation','source','provenance','sourcereference'],
 'rowid':['sourcerowid','rowidentity','sourceid'], 'invoice':['invoiceno','invoice','invoiceid','invoicenumber'],
 'stock':['stockcode','productcode'], 'description':['description','productdescription'],
 'quantity':['quantity','qty'], 'date':['invoicedate','timestamp','transactiondate'],
 'price':['unitprice','unitpricegbp'], 'customer':['customerid','customer'], 'country':['country'],
 'class':['classification','status','category','entrytype'], 'amount':['recordedamountgbp','amountgbp','signedamount','recordedamount','amount'],
 'missing_customer':['missingcustomer','missingcustomerid'], 'count':['rowcount','recordcount','occurrences','count','linecount'],
 'issue':['issuetype','issue'], 'countries':['invoicecountries','countries'],
 'original_invoice':['originalinvoice','linkedinvoice','originalinvoiceno','credittooriginal']}
ALIASES.update({'sale_amount':['salesgbp','salesamount','salesvaluegbp'],'credit_amount':['creditsgbp','creditamount','creditsvaluegbp'],'exception_amount':['exceptionsgbp','exceptionamount'],'sale_count':['saleslines','salesrowcount'],'credit_count':['creditlines','creditrowcount'],'exception_count':['exceptionlines','exceptionrowcount']})
ALIASES.update({ 'component':['component','bridgecomponent','contribution','driver'],
 'bridge_amount':['bridgeamountgbp','bridgeamount','contributiongbp','contributionamount'],
 'cohort':['cohort','skucohort','membership'],
 'q_sep':['sepsalesquantity','septembersalesquantity','sepquantity','septemberquantity','qsep','q0'],
 'q_oct':['octsalesquantity','octobersalesquantity','octquantity','octoberquantity','qoct','q1'],
 'v_sep':['sepsalesvaluegbp','septembersalesvalue','sepsalesvalue','sepvalue','vsep','v0'],
 'v_oct':['octsalesvaluegbp','octobersalesvalue','octsalesvalue','octvalue','voct','v1'],
 'p_sep':['sepweightedunitprice','septemberweightedunitprice','sepweightedprice','psep','p0'],
 'p_oct':['octweightedunitprice','octoberweightedunitprice','octweightedprice','poct','p1'],
 'qty_effect':['quantityeffectgbp','quantityeffect','volumeeffectgbp','volumeeffect','quantitycontribution'],
 'price_effect':['priceandmixeffectgbp','priceandmixeffect','pricemixeffect','priceeffect','pricecontribution'],
 'entry_effect':['newskueffectgbp','newskueffect','newskusales','entryeffect'],
 'exit_effect':['exitedskueffectgbp','exitedskueffect','exitskueffect','exiteffect']})
ALIASES.update({'oct_report':['october'],'sep_report':['september']})
for key,more in {'sale_amount':['sales'],'credit_amount':['credits'],'exception_amount':['exceptions'],'sale_count':['salelines'],'q_sep':['sepqty'],'q_oct':['octqty'],'p_sep':['sepwtdprice'],'p_oct':['octwtdprice'],'qty_effect':['qtycontrib'],'price_effect':['pricemixcontrib'],'entry_effect':['newskucontrib'],'exit_effect':['exitskucontrib']}.items():ALIASES[key].extend(more)
LOOKUP={a:k for k,aliases in ALIASES.items() for a in aliases}

def tables(path,specs):
    w=openpyxl.load_workbook(path,read_only=True,data_only=True)
    result={k:[] for k in specs};text=[]
    for s in w:
        active=None;mapping=None;rows=[];start=None
        def flush():
            if active is not None:result[active].append({'sheet':s.title,'header_row':start,'rows':rows[:]})
        for rownum,row in enumerate(s.iter_rows(),1):
            values=[c.value for c in row]
            text.extend(str(x) for x in values if isinstance(x,str) and len(str(x))>1)
            # Keep visible row context for free-text/key-value reports without a
            # required "Metric / Value" header. Do not infer values from truth.
            populated=[v for v in values if v is not None]
            if len(populated)<=3 or rownum<=30:text.append(' || '.join(str(x) for x in populated))
            headers={LOOKUP[norm(v)]:i for i,v in enumerate(values) if norm(v) in LOOKUP}
            if 'metric' in headers and 'amount' in headers and 'value' not in headers:headers['value']=headers['amount']
            if 'component' in headers and 'value' in headers:headers['bridge_amount']=headers['value']
            if 'oct_report' in headers and 'sep_report' in headers:
                lead=min(headers['oct_report'],headers['sep_report'])-1
                if lead>=0:headers['metric']=lead
            matches=[k for k,required in specs.items() if set(required)<=set(headers)]
            if 'totals' in matches and 'invoices' not in matches and 'invoice' in norm(s.title):
                raise ParsePending('Invoice summary has classification, count and amount but its document identifier header cannot be safely bound.')
            if matches:
                flush();active='queue' if 'queue' in matches and 'issue' in headers else max(matches,key=lambda k:len(specs[k]));mapping=headers;rows=[];start=rownum;continue
            if active is None and populated:text.append('__UNBOUND__ '+' || '.join(str(x) for x in populated))
            if active is not None and any(v is not None for v in values):
                d={k:(values[i] if i<len(values) else None) for k,i in mapping.items()}
                d['_loc']=f'{s.title}!{rownum}';d['_cells']={k:f'{s.title}!{get_column_letter(i+1)}{rownum}' for k,i in mapping.items()}
                if active=='records' and d.get('class') is None and d.get('amount') is None:continue
                if active=='sku' and norm(d.get('stock')) in ['total','grandtotal'] and d.get('cohort') is None:continue
                rows.append(d)
        flush()
    w.close()
    # Wide tables retain candidate values; class counts may be recoverable from delivered detail.
    for wide,target,key in [('invoices_wide','invoices','invoice'),('countries_wide','countries','country'),('countries_amounts','countries','country')]:
        for region in result.pop(wide,[]):
            converted=[];totals=[]
            for r in region['rows']:
                if r.get(key) is None:continue
                is_total=norm(r.get(key)) in ['total','grandtotal','totals']
                for cls in ['sale','credit','exception']:
                    amount=r.get(cls+'_amount');count=r.get(cls+'_count')
                    if not is_total and num(amount)==0 and num(count)==0:continue
                    d={key:r.get(key),'class':cls,'count':count,'amount':amount,'_loc':r['_loc'],'_cells':r['_cells']}
                    if wide=='countries_amounts':d.update(_recover_count=True,_wide_total_count=r.get('count'))
                    (totals if is_total else converted).append(d)
            result[target].append({**region,'rows':converted})
            if wide=='invoices_wide' and totals:result['totals'].append({**region,'rows':totals})
    for region in result.pop('period_report',[]):
        converted=[]
        aliases={'sales':'Gross sales','credits':'Signed credits','netrecordedvalue':'Net recorded value'}
        for r in region['rows']:
            if norm(r.get('metric')) in aliases:
                converted.append({**r,'metric':aliases[norm(r['metric'])],'value':r.get('oct_report')})
        result['report'].append({**region,'rows':converted})
    # A reconciliation table may label the three classes and split adjacent periods.
    for region in result.pop('reconciliation',[]):
        converted=[];outside=[]
        for r in region['rows']:
            label=norm(r.get('description'))
            cls={'sales':'sale','credits':'credit','exceptions':'exception'}.get(label)
            if cls:converted.append({**r,'class':cls})
            elif 'outofperiod' in label:outside.append(r)
        if outside and all(num(r.get('count')) is not None and num(r.get('amount')) is not None for r in outside):
            converted.append({'class':'outside_scope','count':sum(num(r['count']) for r in outside),'amount':sum(num(r['amount']) for r in outside),'_loc':region['sheet'],'_cells':{}})
        result['totals'].append({**region,'rows':converted})
    for region in result['queue']:
        for r in region['rows']:
            if 'issue' not in r and str(r.get('reason') or '').strip():r['issue']='missing_customer_attribution' if 'customer' in str(r['reason']).lower() else 'business_exception'
    # A legal formula whose value is absent must not become a business zero.
    with zipfile.ZipFile(path) as z:
        has_formulas=any(re.search(rb'<(?:\w+:)?f(?:\s|>)',z.read(n)) for n in z.namelist() if re.fullmatch(r'xl/worksheets/sheet\d+.xml',n))
    if not has_formulas:return result,text
    f=openpyxl.load_workbook(path,read_only=True,data_only=False)
    v=openpyxl.load_workbook(path,read_only=True,data_only=True)
    for sf,sv in zip(f,v):
        for rf,rv in zip(sf.iter_rows(),sv.iter_rows()):
            for cf,cv in zip(rf,rv):
                if cf.data_type=='f' and (cv.value is None or cv.data_type=='e'):
                    f.close();v.close();raise ParsePending(f'Uncached or unsupported formula at {sf.title}!{cf.coordinate}')
    f.close();v.close()
    return result,text

def num(v):
    if v is None or norm(v) in {'x','c','w','low','na','unavailable','suppressed','notavailable'}:return None
    try:return Decimal(str(v).replace(',','').replace('£','').strip())
    except (InvalidOperation,ValueError):return None
def eq(a,b,tol=Decimal('0.0000001')):
    if b is None:return num(a) is None
    aa=num(a);bb=num(b)
    return aa is not None and bb is not None and abs(aa-bb)<=tol
def population(got,expected):
    a,b=Counter(got),Counter(list(expected));n=sum(b.values())
    return max(0,(n-sum((b-a).values())-sum((a-b).values()))/max(1,n))
def consensus_rows(regions,key):
    """Multiple final representations stay distinct; no last-write-wins."""
    rows=[r for t in regions for r in t['rows'] if r.get(key) is not None]
    by={}
    for r in rows:by.setdefault(str(r[key]),[]).append(r)
    return rows,by
def mean(vals):return sum(vals)/len(vals) if vals else 0

def visible_numbers(report_rows,text,aliases):
    """Bind numeric claims from explicit labels or nearby visible prose only."""
    found=[];normalized=[norm(a) for a in aliases]
    for r in report_rows:
        label=norm(r.get('metric'))
        if any(label==a or label in [a+'gbp',a+'pounds',a+'count'] for a in normalized):found.append(r.get('value'))
    patterns=['[\\s_]*'.join(re.escape(w) for w in a.split()) for a in aliases]
    pat=re.compile(r'(?:'+ '|'.join(patterns)+r')\s*(?:\([^)]*\))?[\s:|=]*(?:(?:were|was|is|are|totalled|totaled|of|amounted to)\s*)?(?:GBP\s*|£\s*)?[\s:|=]*(\(?-?\d[\d,]*(?:\.\d+)?\)?)',re.I)
    for line in text:
        if '||' in line and sum(num(part.strip()) is not None for part in line.split('||'))>1:continue
        for m in pat.finditer(line):
            # A separately labelled comparison-month endpoint is not an October report claim.
            if re.search(r'(?:september|sep|opening)\s*$',line[:m.start()],re.I):continue
            v=m[1];found.append('-'+v[1:-1] if v.startswith('(') else v)
    return found

def charts(path):
    ns={'c':'http://schemas.openxmlformats.org/drawingml/2006/chart'}
    w=openpyxl.load_workbook(path,data_only=True,read_only=True)
    def reference(f):
        m=re.fullmatch(r"(?:'((?:[^']|'')+)'|([^!]+))!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)",f or '')
        if not m:raise ParsePending('Unsupported legal chart reference: '+str(f))
        name=(m[1] or m[2]).replace("''", "'")
        if name not in w:raise ParsePending('Chart references unknown worksheet')
        a,b,c,d=range_boundaries(m[3]);return [w[name].cell(r,col).value for r in range(b,d+1) for col in range(a,c+1)]
    out=[]
    with zipfile.ZipFile(path) as z:
        for filename in z.namelist():
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',filename):
                tree=ET.fromstring(z.read(filename));series=[]
                for ser in tree.findall('.//c:ser',ns):
                    item={}
                    tx=ser.find('c:tx',ns)
                    txf=tx.find('.//c:f',ns) if tx is not None else None
                    txv=tx.find('.//c:v',ns) if tx is not None else None
                    item['name']=str(reference(txf.text)[0]) if txf is not None else txv.text if txv is not None else ''
                    for tag,label in [('cat','categories'),('val','values')]:
                        el=ser.find('c:'+tag,ns)
                        if el is None:raise ParsePending('Unsupported chart series shape')
                        f=el.find('.//c:f',ns)
                        cache=[p.find('c:v',ns).text for p in el.findall('.//c:pt',ns) if p.find('c:v',ns) is not None]
                        vals=reference(f.text) if f is not None else cache
                        if not vals:raise ParsePending('Unresolvable chart values')
                        item[label]=vals;item[label+'_cache']=cache
                    series.append(item)
                out.append({'chart':filename,'series':series})
    w.close();return out
