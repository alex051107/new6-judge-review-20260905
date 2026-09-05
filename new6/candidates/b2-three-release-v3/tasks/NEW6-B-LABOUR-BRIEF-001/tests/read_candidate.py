"""Read actual labelled candidate views; no Oracle values drive region selection."""
import re,zipfile,posixpath,xml.etree.ElementTree as ET
from collections import Counter
import openpyxl
from chart_reader import ParsePending,norm,num,eq,population,mean,formula_cache_states,charts
NS={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
CODE=re.compile(r'^E\d{8}$')

def labelled_code_list(values):
    """Bind only an explicit list label followed entirely by authority codes.

    These rows assert membership, never numerical results or an explanation.
    Unknown labels and any extra numbers/prose stay on the normal reader path.
    """
    filled=[v for v in values if v is not None]
    if not filled or not all(isinstance(v,str) for v in filled):return None
    if len(filled)==1:
        parts=filled[0].split(':',1)
        if len(parts)!=2:return None
        label,payload=parts
    else:label,payload=filled[0],' '.join(filled[1:])
    labels={
      'previousshortlistcodes':'previous','previousshortlist':'previous','priorshortlistcodes':'previous',
      'currentselectedcodes':'current','currentshortlistcodes':'current','currentshortlist':'current',
      'authoritiesretained':'retained','retainedauthorities':'retained','retainedcodes':'retained',
      'authoritiesentered':'entered','enteredauthorities':'entered','enteredcodes':'entered','newentrants':'entered',
      'authoritiesleft':'left','leftauthorities':'left','leftcodes':'left','departedauthorities':'left'}
    kind=labels.get(norm(label))
    if kind is None or not re.fullmatch(r'\s*E\d{8}(?:[\s,;|]+E\d{8})*[\s,;|]*',payload):return None
    return {'kind':kind,'codes':re.findall(r'E\d{8}',payload),'label':label}

def historic(s):
    return bool(re.search(r'\b(previous|prior|historical|history|archive|archived|old)\b',str(s),re.I))

def field(label):
    t=str(label or '').lower();n=norm(t)
    basic={'code':'code','geographycode':'code','authoritycode':'code','areacode':'code','onscode':'code',
      'authority':'name','localauthority':'name','authorityname':'name','geography':'name','geographyname':'name',
      'reason':'reason','exclusionreason':'reason','explanation':'reason','movement':'movement','shortlistchange':'movement',
      'previousrank':'old_rank','priorrank':'old_rank','currentrank':'rank','rank':'rank','position':'rank',
      'previouslyselected':'previous_selected','currentlyselected':'current_selected',
      'currenteligibility':'eligible','eligible':'eligible','continuedeterioration':'eligible'}
    if n in basic:return basic[n]
    metric='unemployment' if re.search(r'\bunemp|unemployment',t) else 'employment' if re.search(r'\bemp|employment',t) else None
    if not metric:return None
    if any(x in t for x in ['cumulative','overall','first to last','total change']):return metric+'_change_23_25'
    years=re.findall(r'20(\d\d)',t)
    change=any(x in t for x in ['change','delta','difference','increase','fall','rise','Δ'.lower(),'pp'])
    if change:
        if 'first' in t or 'earlier comparison' in t:return metric+'_change_23_24'
        if 'second' in t or 'latest comparison' in t:return metric+'_change_24_25'
        if len(years)==2 and years in [['23','24'],['24','25'],['23','25']]:return metric+'_change_'+'_'.join(years)
        short=re.findall(r'(?<!\d)(2[345])(?!\d)',t)
        if len(short)==2:return metric+'_change_'+'_'.join(short)
    elif len(years)==1 and years[0] in ['23','24','25']:return metric+'_20'+years[0]
    elif len(years)==2 and ('oct' in t and 'sep' in t):
        period={('21','22'):'2023',('22','23'):'2024',('23','24'):'2025'}.get(tuple(years))
        if period:return metric+'_'+period
    return None

def read(path):
    wb=openpyxl.load_workbook(path,data_only=True,read_only=True)
    regions=[];lines=[];unbound=[]
    for sheet in wb:
        context=sheet.title;active=None
        for ri,row in enumerate(sheet.iter_rows(),1):
            vals=[c.value for c in row];filled=[v for v in vals if v is not None]
            if not filled:continue
            line=' || '.join(str(v) for v in filled)
            entry={'sheet':sheet.title,'row':ri,'text':line,'historical':historic(sheet.title)}
            code_list=labelled_code_list(vals)
            if code_list:entry['code_list']=code_list
            lines.append(entry)
            if code_list:
                active=None
                continue
            headers={field(v):ci for ci,v in enumerate(vals) if field(v)}
            if 'code' in headers and len(headers)>1:
                comparison='movement' in headers or {'old_rank','rank'}<=set(headers) or {'previous_selected','current_selected'}<=set(headers)
                old=(historic(context) or 'old_rank' in headers) and not comparison
                role='previous' if old else 'shortlist' if 'rank' in headers or re.search(r'\bshortlist\b',context,re.I) else 'data'
                if comparison:role='movement'
                elif not old and 'reason' in headers and not any('change' in k for k in headers):role='exclusions'
                active={'sheet':sheet.title,'header_row':ri,'columns':headers,'role':role,'rows':[],'context':context}
                regions.append(active);continue
            if len(filled)==1 and not CODE.fullmatch(str(filled[0])):
                context=sheet.title+' '+str(filled[0]);active=None
            codes=[v for v in filled if CODE.fullmatch(str(v))]
            if active and vals[active['columns']['code']] is not None and CODE.fullmatch(str(vals[active['columns']['code']])):
                d={k:vals[c] if c<len(vals) else None for k,c in active['columns'].items()}
                d['_loc']=f'{sheet.title}!{ri}';d['_text']=line;active['rows'].append(d)
            elif codes and not (len(filled)<=3 and any(isinstance(x,str) and len(x)>45 and re.search(r'[a-z]{3} [a-z]{3}',x,re.I) for x in filled)):
                unbound.append({'sheet':sheet.title,'row':ri,'text':line})
    wb.close()
    has_formula,empty=formula_cache_states(path)
    if has_formula:
        f=openpyxl.load_workbook(path,data_only=False,read_only=True);v=openpyxl.load_workbook(path,data_only=True,read_only=True)
        for sf,sv in zip(f,v):
            for rf,rv in zip(sf.iter_rows(),sv.iter_rows()):
                for cf,cv in zip(rf,rv):
                    if cf.data_type=='f' and cv.value is None and (sf.title,cf.coordinate) not in empty:
                        f.close();v.close();raise ParsePending('Uncached formula requires isolated native calculation: '+sf.title+'!'+cf.coordinate)
                    if cf.data_type=='f' and cv.data_type=='e':
                        # Formula error attribution needs native/environment inspection.
                        f.close();v.close();raise ParsePending('Formula error requires attribution: '+sf.title+'!'+cf.coordinate)
        f.close();v.close()
    return regions,lines,unbound

def chart_owners(path):
    rid='{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    def target(base,t):return t.lstrip('/') if t.startswith('/') else posixpath.normpath(posixpath.dirname(base)+'/'+t)
    def relpath(p):return posixpath.dirname(p)+'/_rels/'+posixpath.basename(p)+'.rels'
    owners={}
    with zipfile.ZipFile(path) as z:
        def rels(p):
            rp=relpath(p)
            return {a.get('Id'):target(p,a.get('Target')) for a in ET.fromstring(z.read(rp))} if rp in z.namelist() else {}
        br=rels('xl/workbook.xml')
        for sheet in ET.fromstring(z.read('xl/workbook.xml')).findall('m:sheets/m:sheet',NS):
            sp=br[sheet.get(rid)];sr=rels(sp)
            for drawing in ET.fromstring(z.read(sp)).findall('m:drawing',NS):
                dp=sr[drawing.get(rid)]
                for cp in rels(dp).values():
                    if re.search(r'/charts/.*\.xml$',cp):owners[cp]=sheet.get('name')
    return owners

def read_charts(path):
    owners=chart_owners(path)
    out=[]
    with zipfile.ZipFile(path) as z:
        for c in charts(path):
            root=ET.fromstring(z.read(c['chart']))
            title=' '.join(x.text or '' for x in root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/chart}title//{http://schemas.openxmlformats.org/drawingml/2006/main}t'))
            out.append({**c,'title':title,'sheet':owners.get(c['chart'],''),'historical':historic(owners.get(c['chart'],''))})
    return out

def embedded_chart_candidates(path,lines):
    """Locate attached current-chart images requiring a visual reader.

    Attachment and visible chart context are required; arbitrary media files or
    an identified logo are not evidence of a business chart or its correctness.
    """
    rid='{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
    draw={'x':'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
          'a':'http://schemas.openxmlformats.org/drawingml/2006/main'}
    def target(base,value):return value.lstrip('/') if value.startswith('/') else posixpath.normpath(posixpath.dirname(base)+'/'+value)
    def chart_words(value):return bool(re.search(r'\b(?:charts?|graphs?|plots?|visualisations?|visualizations?)\b',re.sub(r'[_-]+',' ',str(value)),re.I))
    found=[]
    with zipfile.ZipFile(path) as z:
        def rels(part):
            rp=posixpath.dirname(part)+'/_rels/'+posixpath.basename(part)+'.rels'
            return {x.get('Id'):target(part,x.get('Target')) for x in ET.fromstring(z.read(rp)) if x.get('TargetMode')!='External'} if rp in z.namelist() else {}
        br=rels('xl/workbook.xml')
        for sheet in ET.fromstring(z.read('xl/workbook.xml')).findall('m:sheets/m:sheet',NS):
            name=sheet.get('name')
            if historic(name):continue
            sp=br[sheet.get(rid+'id')];sr=rels(sp)
            for drawing in ET.fromstring(z.read(sp)).findall('m:drawing',NS):
                dp=sr[drawing.get(rid+'id')];dr=rels(dp)
                for anchor in ET.fromstring(z.read(dp)):
                    start=anchor.find('x:from',draw)
                    row=int(start.find('x:row',draw).text)+1 if start is not None else None
                    col=int(start.find('x:col',draw).text)+1 if start is not None else None
                    nearby=[x['text'] for x in lines if x['sheet']==name and row is not None and row-4<=x['row']<=row+2]
                    for pic in anchor.findall('.//x:pic',draw):
                        props=pic.find('x:nvPicPr/x:cNvPr',draw)
                        desc=' '.join(props.get(k,'') for k in ['name','descr','title']) if props is not None else ''
                        near_chart=any(chart_words(x) for x in nearby)
                        labelled_logo=bool(re.search(r'\b(?:logo|icon|badge)\b',desc,re.I)) and not chart_words(desc) and not near_chart
                        explicit=chart_words(name) or chart_words(desc)
                        review_context=bool(re.search(r'brief|report|summary|comparison|review',name,re.I)) and near_chart
                        if labelled_logo or not (explicit or review_context):continue
                        for blip in pic.findall('.//a:blip',draw):
                            media=dr.get(blip.get(rid+'embed'))
                            if media and media in z.namelist():found.append({'sheet':name,'row':row,'column':col,'drawing':dp,'media':media,'description':desc,'nearby_labels':nearby,'basis':'attached image in explicit current chart context; image contents require a supported reader'})
    return found
