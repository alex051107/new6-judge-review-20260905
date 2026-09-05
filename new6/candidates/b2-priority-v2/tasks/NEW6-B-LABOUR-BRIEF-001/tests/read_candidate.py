"""Conservative label/identity reader, with actual OOXML chart evidence.

This reader discovers headers and locations from candidate contents. It does not
use reference cell coordinates or interpret candidate formulas. Uncached formulas
and chart reference forms outside the supported reader have explicit pending states.
"""
from pathlib import Path
from decimal import Decimal, InvalidOperation
from collections import Counter
import re, zipfile, posixpath, xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.cell import range_boundaries,get_column_letter

class ParsePending(Exception):pass
def norm(v):return re.sub(r'[^a-z0-9]+','',str(v or '').lower())
ALIASES={
 'code':['geographycode','areacode','localcode','authoritycode','code'],
 'name':['geography','authority','areaname','geographyname','localauthority'],
 'employment_old':['employmentrateearlier','employmentrate2024','employment2024','earlieremploymentrate'],
 'employment_new':['employmentratelater','employmentrate2025','employment2025','lateremploymentrate'],
 'unemployment_old':['unemploymentrateearlier','unemploymentrate2024','unemployment2024','earlierunemploymentrate'],
 'unemployment_new':['unemploymentratelater','unemploymentrate2025','unemployment2025','laterunemploymentrate'],
 'employment_change':['employmentchangepp','employmentratechangepp','employmentchange','employmentppchange'],
 'unemployment_change':['unemploymentchangepp','unemploymentratechangepp','unemploymentchange','unemploymentppchange'],
 'comparison_status':['comparisonstatus','comparability','pairstatus'],
 'metric':['metric','indicator','measure','field','item'], 'value':['value','reportedvalue','result','finding'],
 'reason':['reason','exclusionreason','treatmentreason'],
 'rank':['rank','position'], 'edition':['edition','release'],
 'raw':['rawvalue','storedvalue','underlyingvalue'], 'published':['publishedvalue','displayedrate','publishedrate'],
 'source':['sourcelocation','source','provenance','sourcereference'],
 'rowid':['sourcerowid','rowidentity','sourceid'], 'invoice':['invoiceno','invoice','invoiceid'],
 'stock':['stockcode','productcode'], 'description':['description','productdescription'],
 'quantity':['quantity','qty'], 'date':['invoicedate','timestamp','transactiondate'],
 'price':['unitprice','unitpricegbp'], 'customer':['customerid','customer'], 'country':['country'],
 'class':['classification','status','category','entrytype'], 'amount':['recordedamountgbp','amountgbp','signedamount','recordedamount','amount'],
 'missing_customer':['missingcustomer','missingcustomerid'], 'count':['rowcount','recordcount','occurrences','count'],
 'issue':['issuetype','issue'], 'countries':['invoicecountries','countries'],
 'original_invoice':['originalinvoice','linkedinvoice','originalinvoiceno','credittooriginal']}
LOOKUP={a:k for k,aliases in ALIASES.items() for a in aliases}
for field, aliases in {
    'employment_old':['Employment Rate Jan 2024','Employment Rate January 2024'],
    'employment_new':['Employment Rate Jan 2025','Employment Rate January 2025'],
    'unemployment_old':['Unemployment Rate Jan 2024','Unemployment Rate January 2024'],
    'unemployment_new':['Unemployment Rate Jan 2025','Unemployment Rate January 2025'],
    'employment_change':['Emp Change','Emp Change (pp)'],
    'unemployment_change':['Unemp Change','Unemp Change (pp)'],
    'employment_reason':['Employment Exclusion Reason'],
    'unemployment_reason':['Unemployment Exclusion Reason'],
}.items():
    for alias in aliases:LOOKUP[norm(alias)]=field

def shortlist_scope(label):
    """Bind an explicit business caption, without inspecting candidate answers."""
    text=str(label or '').lower()
    if 'shortlist' not in text:return None
    kinds=[kind for kind in ['baseline','relaxed','strict','current','live'] if re.search(r'\b'+kind+r'\b',text)]
    if len(kinds)!=1:return None
    return 'current' if kinds[0]=='live' else kinds[0]

def formula_cache_states(path):
    """An OOXML string cache <v/> is a calculated empty string, not no cache."""
    ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rid='{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    empty=set();has_formulas=False
    with zipfile.ZipFile(path) as z:
        rels={r.attrib['Id']:r.attrib['Target'] for r in ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))}
        book=ET.fromstring(z.read('xl/workbook.xml'))
        for sheet in book.findall('m:sheets/m:sheet',ns):
            target=rels[sheet.attrib[rid]]
            target=target.lstrip('/') if target.startswith('/') else posixpath.normpath('xl/'+target)
            for cell in ET.fromstring(z.read(target)).findall('.//m:c',ns):
                if cell.find('m:f',ns) is None:continue
                has_formulas=True;value=cell.find('m:v',ns)
                if cell.get('t')=='str' and value is not None and value.text in [None,'']:
                    empty.add((sheet.attrib['name'],cell.attrib['r']))
    return has_formulas,empty

def tables(path,specs):
    w=openpyxl.load_workbook(path,read_only=True,data_only=True)
    result={k:[] for k in specs};text=[]
    for s in w:
        active=None;mapping=None;rows=[];start=None;scope=None;virtual_scenario=None;wide_exclusions=False
        def flush():
            if active is not None:result[active].append({'sheet':s.title,'header_row':start,'rows':rows[:]})
        for rownum,row in enumerate(s.iter_rows(),1):
            values=[c.value for c in row]
            text.extend(str(x) for x in values if isinstance(x,str) and len(str(x))>1)
            # Keep visible row context for free-text/key-value reports without a
            # required "Metric / Value" header. Do not infer values from truth.
            populated=[v for v in values if v is not None]
            if len(populated)<=3 or rownum<=30:text.append(' || '.join(str(x) for x in populated))
            if len(populated)==1 and shortlist_scope(populated[0]):
                flush();active=None;rows=[];scope=shortlist_scope(populated[0]);continue
            headers={LOOKUP[norm(v)]:i for i,v in enumerate(values) if norm(v) in LOOKUP}
            if 'metric' in headers and 'amount' in headers and 'value' not in headers:headers['value']=headers['amount']
            virtual=None
            if 'code' in headers and 'rank' in headers:
                if scope in ['baseline','relaxed','strict'] and 'review' in specs:
                    headers['review_order']=headers['rank'];virtual=scope
                elif scope=='current' and 'current' in specs:
                    headers['current_order']=headers['rank']
            matches=[k for k,required in specs.items() if set(required)<=set(headers)]
            if virtual and set(specs['review'])-{'scenario'}<=set(headers):matches.append('review')
            wide='exclusions' in specs and {'code','employment_reason','unemployment_reason'}<=set(headers)
            if wide:matches.append('exclusions')
            if scope and 'top' in matches:matches.remove('top')
            if matches:
                flush();active='top' if 'top' in matches and 'rank' in headers else max(matches,key=lambda k:len(specs[k]));mapping=headers;rows=[];start=rownum;virtual_scenario=virtual;wide_exclusions=wide;continue
            if active is None and populated:text.append('__UNBOUND__ '+' || '.join(str(x) for x in populated))
            if active is not None and any(v is not None for v in values):
                d={k:(values[i] if i<len(values) else None) for k,i in mapping.items()}
                d['_loc']=f'{s.title}!{rownum}';d['_cells']={k:f'{s.title}!{get_column_letter(i+1)}{rownum}' for k,i in mapping.items()}
                if virtual_scenario:d['scenario']=virtual_scenario
                if wide_exclusions:
                    for metric in ['employment','unemployment']:
                        reason=d.get(metric+'_reason')
                        if reason is not None and str(reason).strip():rows.append({**d,'metric':metric,'reason':reason})
                else:rows.append(d)
        flush()
    w.close()
    # A legal formula whose value is absent must not become a business zero.
    has_formulas,empty_strings=formula_cache_states(path)
    if not has_formulas:return result,text
    f=openpyxl.load_workbook(path,read_only=True,data_only=False)
    v=openpyxl.load_workbook(path,read_only=True,data_only=True)
    for sf,sv in zip(f,v):
        for rf,rv in zip(sf.iter_rows(),sv.iter_rows()):
            for cf,cv in zip(rf,rv):
                if cf.data_type=='f' and (cv.data_type=='e' or (cv.value is None and (sf.title,cf.coordinate) not in empty_strings)):
                    f.close();v.close();raise ParsePending(f'Uncached or unsupported formula at {sf.title}!{cf.coordinate}')
    f.close();v.close()
    return result,text

def num(v):
    if v is None or norm(v) in {'x','c','w','low','na','unavailable','suppressed','notavailable'}:return None
    try:return Decimal(str(v).replace(',','').replace('£','').strip())
    except (InvalidOperation,ValueError):return None
def eq(a,b,tol=Decimal('0.0000001')):
    if b is None:return num(a) is None
    aa=num(a);bb=num(b)
    return aa is not None and bb is not None and abs(aa-bb)<=tol
def population(got,expected):
    a,b=Counter(got),Counter(list(expected));n=sum(b.values())
    return max(0,(n-sum((b-a).values())-sum((a-b).values()))/max(1,n))
def consensus_rows(regions,key):
    """Multiple final representations stay distinct; no last-write-wins."""
    rows=[r for t in regions for r in t['rows'] if r.get(key) is not None]
    by={}
    for r in rows:by.setdefault(str(r[key]),[]).append(r)
    return rows,by
def mean(vals):return sum(vals)/len(vals) if vals else 0

def visible_numbers(report_rows,text,aliases):
    """Bind numeric claims from explicit labels or nearby visible prose only."""
    found=[];normalized=[norm(a) for a in aliases]
    for r in report_rows:
        label=norm(r.get('metric'))
        if any(label==a or label in [a+'gbp',a+'pounds',a+'count'] for a in normalized):found.append(r.get('value'))
    patterns=['[\\s_]*'.join(re.escape(w) for w in a.split()) for a in aliases]
    pat=re.compile(r'(?:'+ '|'.join(patterns)+r')\s*(?:\([^)]*\))?[\s:|=]*(?:(?:were|was|is|are|totalled|totaled|of|amounted to)\s*)?(?:GBP\s*|£\s*)?[\s:|=]*(\(?-?\d[\d,]*(?:\.\d+)?\)?)',re.I)
    for line in text:
        for m in pat.finditer(line):
            v=m[1];found.append('-'+v[1:-1] if v.startswith('(') else v)
    return found

def charts(path):
    ns={'c':'http://schemas.openxmlformats.org/drawingml/2006/chart'}
    w=openpyxl.load_workbook(path,data_only=True,read_only=True)
    def reference(f):
        # LibreOffice serializes a literal series title as an Excel string expression.
        if f and re.fullmatch(r'"(?:[^"]|"")*"', f):
            return [f[1:-1].replace('""', '"')]
        m=re.fullmatch(r"(?:'((?:[^']|'')+)'|([^!]+))!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)",f or '')
        if not m:raise ParsePending('Unsupported legal chart reference: '+str(f))
        name=(m[1] or m[2]).replace("''", "'")
        if name not in w:raise ParsePending('Chart references unknown worksheet')
        a,b,c,d=range_boundaries(m[3]);return [w[name].cell(r,col).value for r in range(b,d+1) for col in range(a,c+1)]
    out=[]
    with zipfile.ZipFile(path) as z:
        for filename in z.namelist():
            if re.fullmatch(r'xl/(?:drawings/)?charts/chart[^/]+\.xml',filename):
                tree=ET.fromstring(z.read(filename));series=[]
                inherited={}
                grouped=[(id(group),ser) for group in tree.findall('.//c:plotArea/*',ns) for ser in group.findall('c:ser',ns)]
                for groupid,ser in grouped:
                    item={}
                    tx=ser.find('c:tx',ns)
                    txf=tx.find('.//c:f',ns) if tx is not None else None
                    txv=tx.find('.//c:v',ns) if tx is not None else None
                    item['name']=str(reference(txf.text)[0]) if txf is not None else txv.text if txv is not None else ''
                    for tag,label in [('cat','categories'),('val','values')]:
                        el=ser.find('c:'+tag,ns)
                        if el is None:
                            if tag=='cat' and groupid in inherited:
                                item[label]=inherited[groupid][0];item[label+'_cache']=inherited[groupid][1];item['shared_category_source']=True;continue
                            raise ParsePending('Unsupported chart series shape')
                        f=el.find('.//c:f',ns)
                        cache=[p.find('c:v',ns).text for p in el.findall('.//c:pt',ns) if p.find('c:v',ns) is not None]
                        vals=reference(f.text) if f is not None else cache
                        if not vals:raise ParsePending('Unresolvable chart values')
                        item[label]=vals;item[label+'_cache']=cache
                        if tag=='cat':inherited[groupid]=(vals,cache)
                    if len(item['categories'])!=len(item['values']):raise ParsePending('Shared chart categories have incompatible point counts')
                    series.append(item)
                out.append({'chart':filename,'series':series})
    w.close();return out
