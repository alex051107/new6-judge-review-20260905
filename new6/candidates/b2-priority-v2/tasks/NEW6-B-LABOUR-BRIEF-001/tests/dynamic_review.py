"""Real isolated native recalculation; no formula interpretation or candidate repair."""
from pathlib import Path
from decimal import Decimal
import json,os,tempfile,re
from collections import Counter
import openpyxl
from read_candidate import tables,LOOKUP,norm,num,eq,mean
from fixture_xml import Fixture
from runtime import recalculate_xlsx,sha256,RecalcUnavailable
TASK=Path(__file__).resolve().parents[1]
CONTROLS={'Unemployment threshold pp':1,'Employment decline threshold pp':1,'Review places':5}
CASES=[{'Unemployment threshold pp':2,'Employment decline threshold pp':1,'Review places':5},
       {'Unemployment threshold pp':1,'Employment decline threshold pp':5,'Review places':5},
       {'Unemployment threshold pp':1,'Employment decline threshold pp':1,'Review places':3},
       {'Unemployment threshold pp':.5,'Employment decline threshold pp':.5,'Review places':7}]
for key,labels in {'current_eligible':['current eligible','live eligible','current eligibility','eligible for current review'], 'current_selected':['current selected','live selected','current selection','current selection status'],'current_order':['current order','live order','current rank','current shortlist order']}.items():
    for label in labels:LOOKUP[norm(label)]=key
SPEC={'live':{'code','current_eligible','current_selected'},'current':{'code','current_order','employment_change','unemployment_change'}}
def snapshot(path):
    ts,text=tables(path,SPEC);out={};duplicates=0
    if not ts['live'] and any('__UNBOUND__' in x and 'code' in x.lower() and 'eligible' in x.lower() for x in text):
        raise RecalcUnavailable('Possible live screening table cannot be safely bound; manual parse required.')
    for t in ts['live']:
        counts=Counter(str(r['code']) for r in t['rows'] if r.get('code'))
        duplicates+=sum(n-1 for n in counts.values())
        for r in t['rows']:
            if not r.get('code'):continue
            for field in ['current_eligible','current_selected']:
                out.setdefault('screen:'+str(r['code'])+':'+field,[]).append(norm(r.get(field)))
    for t in ts['current']:
        counts=Counter(str(r['code']) for r in t['rows'] if r.get('code') and str(r['code']).strip() not in ['-','—',''])
        duplicates+=sum(n-1 for n in counts.values())
        for r in t['rows']:
            if not r.get('code') or str(r['code']).strip() in ['-','—','']:continue
            order=num(r.get('current_order'))
            if order is None:continue
            for field in ['code','employment_change','unemployment_change']:
                out.setdefault('rank:'+str(int(order))+':'+field,[]).append(r.get(field))
    w=openpyxl.load_workbook(path,data_only=True,read_only=True)
    for s in w:
        for row in s:
            for i,c in enumerate(row[:-1]):
                if norm(c.value) in ['currentselectedcount','liveselectedcount']:out.setdefault('count',[]).append(row[i+1].value)
    out['duplicate_count']=[duplicates]
    w.close();return out
def truth(settings):
    panel=json.loads((TASK/'solution/oracle.json').read_text())['panel'];out={};chosen=[]
    for r in panel:
        available=r[6] is not None and r[7] is not None
        eligible=available and Decimal(r[7])>=Decimal(str(settings['Unemployment threshold pp'])) and Decimal(r[6])<=-Decimal(str(settings['Employment decline threshold pp']))
        out['screen:'+r[0]+':current_eligible']='unavailable' if not available else 'yes' if eligible else 'no'
        if eligible:chosen.append(r)
    chosen.sort(key=lambda r:(-Decimal(r[7]),Decimal(r[6]),r[0]));chosen=chosen[:int(settings['Review places'])]
    codes={r[0] for r in chosen}
    for r in panel:out['screen:'+r[0]+':current_selected']='yes' if r[0] in codes else 'no'
    for i in range(1,11):
        r=chosen[i-1] if i<=len(chosen) else None
        for k,idx in [('code',0),('employment_change',6),('unemployment_change',7)]:out[f'rank:{i}:{k}']=r[idx] if r else None
    out['count']=len(chosen);out['duplicate_count']=0;return out
def correct(values,want):
    if want is None:return not values
    if not values:return False
    if num(want) is not None:return all(eq(v,want) for v in values)
    aliases={'true':'yes','1':'yes','eligible':'yes','selected':'yes','false':'no','0':'no','ineligible':'no','notselected':'no','na':'unavailable','missing':'unavailable','suppressed':'unavailable'}
    return all(aliases.get(norm(v),norm(v))==norm(want) for v in values)
def controls(path):
    w=openpyxl.load_workbook(path,data_only=False,read_only=True);found={k:[] for k in CONTROLS}
    for s in w:
        for row in s:
            for i,c in enumerate(row[:-1]):
                for key in CONTROLS:
                    # A fixed-scenario column heading can use the same words.
                    # Only a label with an adjacent numeric input is a control.
                    if norm(c.value)==norm(key) and num(row[i+1].value) is not None:
                        found[key].append((s.title,row[i+1].coordinate,row[i+1].value))
    w.close()
    return found
def invariant_snapshot(path):
    from evaluate_base import SPECS
    ts,_=tables(path,SPECS)
    # Actual static business values, independent of sheet/cell placement.
    return {k:sorted(json.dumps({a:b for a,b in row.items() if not a.startswith('_')},sort_keys=True,default=str) for t in regions for row in t['rows']) for k,regions in ts.items()}
def grade(path):
    before=sha256(path);location=controls(path)
    if any(len(v)>1 for v in location.values()):raise RecalcUnavailable('Multiple possible live input cells; cannot safely select one.')
    if any(not v for v in location.values()):return 0,{'reason':'Required labelled editable input missing','controls_found':{k:len(v) for k,v in location.items()}}
    if any(not eq(v[0][2],CONTROLS[k]) for k,v in location.items()):
        return 0,{'reason':'Initial live settings do not match the disclosed starting policy','observed':{k:v[0][2] for k,v in location.items()}}
    original=snapshot(path);expected=truth(CONTROLS);base=mean([correct(original.get(k,[]),v) for k,v in expected.items()]+[False for k in original if k not in expected]);fixed=invariant_snapshot(path)
    out=Path(os.environ.get('NEW6_EVIDENCE_DIR') or tempfile.mkdtemp(prefix='b2-dynamic-'))/'live_review';out.mkdir(parents=True,exist_ok=True)
    responses=[];invariants=[];receipts=[]
    for i,settings in enumerate(CASES):
        case=out/f'case-{i+1}';case.mkdir(exist_ok=True);mutated=case/'edited.xlsx';edit=Fixture(path)
        for key,value in settings.items():edit.cell(location[key][0][0],location[key][0][1],value)
        edit.save(mutated)
        recalced,receipt=recalculate_xlsx(mutated,case/'native',timeout=120)
        actual=snapshot(recalced);want=truth(settings);changed=[k for k in want if want[k]!=expected[k]]
        assert changed,'Dynamic case must change a scored result'
        checks=[{'fact':k,'expected':want[k],'candidate':actual.get(k,[]),'correct':correct(actual.get(k,[]),want[k])} for k in changed]
        responses.append(mean([x['correct'] for x in checks]))
        unchanged=[correct(actual.get(k,[]),want[k]) for k in want if k not in changed]+[False for k in actual if k not in want]
        invariants.append(mean([mean(unchanged),invariant_snapshot(recalced)==fixed]))
        receipts.append({'settings':settings,'edited_sha256':sha256(mutated),'recalculated_sha256':sha256(recalced),'native':receipt,'changed_denominator':len(changed),'changed_facts':checks,'invariant_credit':invariants[-1]})
    assert sha256(path)==before,'Original candidate changed'
    evidence={'original_sha256':before,'original_unchanged':True,'initial_credit':base,'response_credit':mean(responses),'invariant_credit':mean(invariants),'cases':receipts,'allocation':'10% initial live facts, 80% actual changed result facts, 10% unchanged live and fixed outputs'}
    (out/'receipt.json').write_text(json.dumps(evidence,indent=2,default=str))
    return .1*base+.8*mean(responses)+.1*mean(invariants),evidence
