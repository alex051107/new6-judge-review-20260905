"""Semantic table roles and vertical movement bridges; no Oracle-value discovery."""
from collections import defaultdict
import re
import openpyxl

def read(w,sr,legacy,reference_failure=False,arithmetic_failure=False):
 natural={'elementpackage','allowancetype','reviewprice','currentreview','originalestimate','newvalue'}
 selected=[]
 for sh in w:
  headers={sr.norm(c.value) for row in sh for c in row if isinstance(c.value,str)}
  bridge=any(re.match(r'original cost limit',str(c.value),re.I) for row in sh for c in row) and any(sr.norm(c.value)=='totalmovement' for row in sh for c in row)
  if headers&natural or bridge:selected.append((sh,bridge))
 if not selected:return legacy(w,reference_failure,arithmetic_failure)
 selected_names={s.title for s,_ in selected};remaining=openpyxl.Workbook();remaining.remove(remaining.active)
 for sh in w:
  if sh.title in selected_names:continue
  target=remaining.create_sheet(sh.title)
  for row in sh:
   for c in row:
    if c.value is not None:target.cell(c.row,c.column,c.value)
 if not remaining.worksheets:remaining.create_sheet('No other delivered tables')
 found,text=legacy(remaining,reference_failure,arithmetic_failure)
 text=' '.join(str(c.value) for sh in w for row in sh for c in row if c.value is not None)
 unbound=[]
 def put(key,rec):found.setdefault(key,[]).append(rec)
 def field(v):
  aliases={'elementpackage':'label','package':'label','allowancetype':'label','adjustment':'label','originalestimate':'printed','currentreview':'working','originalprice':'printed','reviewprice':'working','review':'working','movement':'delta','reviewrate':'working_rate','newvalue':'working','sourcenotes':'source'}
  return aliases.get(sr.norm(v),sr.header_role(v))
 def record(sh,row,headers,context,table):
  r={'sheet':sh.title,'row':row,'table':table,'context':context,'cells':{},'delta_sign':1}
  for role,col in headers.items():r[role]=sh.cell(row,col).value;r['cells'][role]=sh.cell(row,col).coordinate
  r['label']=str(r.get('label',''))
  if r.get('source'):
   page=re.search(r'\bpage\s*(\d+)\b',str(r['source']),re.I)
   if page:r['page']=page.group(1)
  return r
 for sh,vertical in selected:
  if vertical:
   records=[];primary={};component={};checks=[]
   def primary_record(role,rec):
    if role in primary:raise ValueError('Repeated labelled vertical bridge role requires review: '+sh.title+' / '+role)
    primary[role]=rec
   for row in sh:
    filled=[c for c in row if c.value is not None]
    if len(filled)<2:continue
    label=str(filled[0].value).strip();n=sr.norm(label);value=filled[1].value
    r={'sheet':sh.title,'row':row[0].row,'table':row[0].row,'context':'Vertical cost-limit movement bridge','label':label,'cells':{},'delta_sign':1}
    if n.startswith('originalcostlimit'):
     r['printed']=value;r['cells']['printed']=filled[1].coordinate;primary_record('original',r)
    elif n=='totalmovement':
     r['delta']=value;r['cells']['delta']=filled[1].coordinate;primary_record('movement',r)
    elif n.startswith('currentreviewcostlimit'):
     r['working']=value;r['cells']['working']=filled[1].coordinate;primary_record('current',r)
    elif n.startswith('reconciledcostlimit') or n.startswith('directcalculation'):
     r['working']=value;r['cells']['working']=filled[1].coordinate;records.append(r);component[n]=value
    elif n.startswith('difference'):
     component['check_difference']=value;r['value']=value;r['cells']['value']=filled[1].coordinate;put(('_bridge_check_display',sh.title),r)
    else:
     # These labels explicitly describe increments, not full stage amounts.
     allowed=r'heating.*(?:ashp|package)|ventilation.*nochange|othermepackages.*nochange|netpackagepriceadjustment|maincontractorohp(?:\d+)?|buildingworkstotalafterflowthrough|designdevelopmentriskratechange|inflationrate.*|allowanceratechanges|inflationonrevisedcostlimit'
     if not re.fullmatch(allowed,n):
      if isinstance(value,(int,float)):unbound.append({'sheet':sh.title,'label':label,'role':'vertical bridge component'})
      continue
     r['increment']=value;r['cells']['increment']=filled[1].coordinate;put(('_movement_components',sh.title),r);component[n]=value
   if set(primary)!={'original','movement','current'}:raise ValueError('Incomplete labelled vertical cost-limit bridge: '+sh.title)
   combined=primary['current']
   for name,role in [('original','printed'),('movement','delta')]:combined[role]=primary[name][role];combined['cells'][role]=primary[name]['cells'][role]
   put(('summary','vat_excluded'),combined)
   for r in records:put(('summary','vat_excluded'),r)
   totals=[component.get(k) for k in ['buildingworkstotalafterflowthrough','allowanceratechanges','inflationonrevisedcostlimit']]
   if all(isinstance(v,(int,float)) for v in totals):checks.append({'meaning':'Displayed total movement equals displayed component subtotals','actual':primary['movement']['delta'],'expected_from_candidate':sum(totals),'ok':sr.equal(primary['movement']['delta'],sum(totals))})
   check_values=[v for k,v in component.items() if k.startswith('reconciledcostlimit') or k.startswith('directcalculation')]
   if len(check_values)==2 and 'check_difference' in component:
    checks.append({'meaning':'Displayed verification difference agrees with its two displayed amounts','actual':component['check_difference'],'ok':sr.equal(component['check_difference'],check_values[1]-check_values[0])})
   for c in checks:put(('_bridge_arithmetic',sh.title),c)
   continue
  context=sh.title;headers=None;table=None;last_key=None
  for row in sh:
   populated=[c for c in row if c.value is not None]
   if not populated:continue
   h={field(c.value):c.column for c in row if field(c.value)}
   if 'label' in h and len(h)>=2:headers=h;table=row[0].row;last_key=None;continue
   if len(populated)==1 and isinstance(populated[0].value,str):context=sh.title+' '+str(populated[0].value);last_key=None;continue
   if not headers:continue
   r=record(sh,row[0].row,headers,context,table);label=r['label'];n=sr.norm(label);ctx=sr.norm(context+' '+str(r.get('source','')))
   if not label or label=='None':continue
   rate_only='working_rate' in headers and 'working' not in headers
   reference_view='sourcereferences' in sr.norm(sh.title)
   key=None;extra=[]
   if re.search(r'heating.*(?:ashp|package)',n) and not re.search(r'earlier|previous|superseded',n):
    if reference_view:key=('review_source','heating_price') if 'reviewadjustments' in ctx else ('provisional','heating')
    else:
     key=('review','heating_price')
     if r.get('printed') is not None:extra.append(('provisional','heating'))
   elif n.startswith('asbestos') and ('option' in n or 'optionsawaiting' in ctx or 'reviewadjustments' in ctx):key=('review_source' if reference_view else 'review','asbestos_option')
   elif n.startswith('othermebalance'):key=('_derived_balance','services')
   elif n=='totalmeservices':key=('elements','services')
   elif n=='heatinginclashp':key=('provisional','heating')
   elif n in ['buildingworksestimatetotal','impactonbuildingworkstotal']:key=('summary','building')
   elif n=='afterprelimsohp':key=('summary','building_estimate')
   elif n=='buildingworkstotal':
    key=('summary','building_estimate') if ('maincontractoradditions' in ctx or last_key==('summary','overheads')) else sr.bind_label(label)
   elif n=='buildingworksestimate' and 'elementaltotal' in ctx:key=('summary','building')
   elif n=='costlimit':key=('summary','vat_excluded')
   elif n=='basecostestimatecostlimit':key=('summary','base');extra.append(('summary','vat_excluded'))
   elif n.startswith('maincontractor') and ('ohp' in n or 'overheads' in n):key=('summary','overheads')
   elif n.startswith('designdevelopmentrisk'):key=('summary','design_risk')
   elif n=='projectduration':key=('qualifications','duration')
   elif n=='contractdelivery':key=('qualifications','phase')
   elif n=='pricedate':key=('qualifications','price_basis')
   elif n.startswith('professionalfees'):key=('exclusions','excluded_i')
   elif n=='planningbuildingregsfees':key=('exclusions','excluded_k')
   else:key=sr.bind_label(label)
   if not key:
    if any(isinstance(r.get(k),(int,float)) for k in ['printed','working','rate','working_rate','delta']):unbound.append({'sheet':sh.title,'label':label,'header':headers,'context':context})
    continue
   if reference_view and 'reviewadjustments' in ctx and key==('summary','design_risk') and 'rate' in n:key=('review_source','design_risk_rate')
   if reference_view and key[0]=='review_source':
    r['quoted_source_value']=r.pop('working',r.get('printed'));r['cells']['quoted_source_value']=r['cells'].pop('working',r['cells'].get('printed'))
   if rate_only or ('reviewadjustments' in ctx and key==('summary','design_risk') and 'rate' in n):
    r['_role']='rate_control'
    if 'working' in r:r['working_rate']=r.pop('working');r['cells']['working_rate']=r['cells'].pop('working')
   label_rate=re.search(r'\((\d+(?:\.\d+)?)%\)',label)
   if label_rate and key[0]=='summary':r['rate']=float(label_rate.group(1))/100
   if key[0] in ['exclusions','qualifications'] and r.get('printed') is not None and not isinstance(r['printed'],(int,float)):r['scope']=str(r.get('scope',''))+' '+str(r['printed'])
   if reference_view and r.get('source') and str(r['source']).startswith('RC:'):r['source']='review_correspondence.md :: '+str(r['source']);r['scope']=str(r.get('scope',''))+' '+str(r['source'])
   if key==('summary','building') and 'elementaltotal' in ctx:r['accepted_source_pages']=[2,5]
   if key[0]=='provisional' and 'mepackagebreakdown' in ctx:r['scope']=str(r.get('scope',''))+' Included within elemental costs; source M&E package breakdown.'
   put(key,r)
   for alias in extra:put(alias,dict(r))
   last_key=key
 if unbound:raise ValueError('Material natural-review label remains unbound: '+str(unbound[:3]))
 me_tables={(r['sheet'],r['table']) for r in found.get(('elements','services'),[])}
 for key,rs in found.items():
  if key[0]=='provisional':
   for r in rs:
    if (r['sheet'],r['table']) in me_tables:r['scope']=str(r.get('scope',''))+' Included within elemental costs; delivered M&E subtotal table.'
 for key in ['heating_price','asbestos_option']:
  cited=[r for r in found.get(('review_source',key),[]) if 'review_correspondence.md' in str(r.get('source',''))]
  if cited:
   for r in found.get(('review',key),[]):r['source']=r.get('source') or 'review_correspondence.md :: '+r['label']
 source_name=sr.S['elements'][0]['source']
 if re.search(r'Falmouth',text,re.I) and re.search(r'October\s*2024',text,re.I) and re.search(r'(?:cost estimate|estimate).*revision\s*A',text,re.I):
  for key,rs in found.items():
   if key[0].startswith('_'):continue
   for r in rs:r['document_locator']=source_name+' :: '+r['label']
 pending_errors=['#N/A']+([] if reference_failure else ['#NAME?'])+([] if arithmetic_failure else ['#VALUE!'])
 errors=[(key,r.get(field)) for key,rs in found.items() for r in rs for field in ['printed','working','working_rate','delta'] if r.get(field) in pending_errors]
 if errors:raise sr.RecalcUnavailable('Bound required quantity has an engine/parse error: '+str(errors[:3]))
 return found,text
