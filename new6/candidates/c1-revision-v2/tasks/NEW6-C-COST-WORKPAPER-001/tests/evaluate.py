"""C1 revision facts: visible semantic labels, isolated native edits, fixed units."""
from pathlib import Path
from decimal import Decimal
from collections import defaultdict
import sys,re,json,argparse
import openpyxl
ROOT=Path(__file__).resolve().parents[1]
sys.path.insert(0,str(ROOT/'metadata'))
import source_reader as sr
from oracle_recompute import compute,serial
from runtime import score_profiles,output_status,recalculate_xlsx,RecalcUnavailable
from ooxml_edit import edit

VERSION='new6-c1-review-facts-v2.1-natural-movement'
REVIEW_ALIASES={
 'heating_price':['Current heating / ASHP quotation','Current heating quote','Current heating price','Latest heating quotation','Heating package price','Revised heating / ASHP price'],
 'earlier_heating':['Earlier heating / ASHP quotation','Earlier heating quote','Previous heating quotation','Superseded heating quote'],
 'asbestos_option':['Asbestos removal option','Unapproved asbestos option','Asbestos option price','Optional asbestos removal']}
for key,aliases in REVIEW_ALIASES.items():
 for alias in aliases:sr.LOOK[sr.norm(alias)]=('review',key)
sr.H.update({x:'working' for x in ['reviewamount','currentprice','packageprice','optionprice']})
sr.H.update({x:'recomputed_original' for x in ['recalculatedoriginal','recomputedoriginal','recalculatedoriginalamount','recomputedoriginalamount']})
sr.H.update({x:'delta_recomputed' for x in ['differencefromrecalculatedoriginal','differencefromrecomputedoriginal','currentminusrecomputedoriginal']})
SCENARIOS=[('package_price',{'heating_price':468000}),('risk_percentage',{'design_risk_rate':'.14'}),('price_and_allowances',{'heating_price':445000,'overheads_rate':'.12','inflation_rate':'.015'}),('unapproved_option_price',{'asbestos_option':22000})]
def close(a,b):
 try:return abs(sr.number(a)-float(b))<=.500001
 except (TypeError,ValueError):return False
def close_delta(a,b):
 try:return abs(sr.number(a)-float(b))<=1.000001
 except (TypeError,ValueError):return False
def rows(f,kind,key):return sr.records(f,kind,key)
def vals(f,kind,key,field):return [r[field] for r in rows(f,kind,key) if r.get(field) is not None]
def number_fact(f,kind,key,field,expected,rounding=True):
 v=vals(f,kind,key,field);return bool(v) and all((close if rounding else sr.equal)(a,expected) for a in v)
def review_values(f,key):
 return [r.get('working') if r.get('working') is not None else r.get('printed') for r in rows(f,'review',key)]
def review_number(f,key,target):
 v=review_values(f,key);return sr.unique_records(rows(f,'review',key)) and bool(v) and all(sr.equal(a,target) for a in v)
def text_rows(f,key):return ' '.join(str(r.get('scope',''))+' '+r.get('label','') for r in rows(f,'review',key)).lower()
def original_model():return sr.compute()
def comparison_basis(r,key):
 # Explicit candidate labels choose the original comparison basis, never closeness
 # to one of two possible expected answers.
 labelled=('delta_recomputed' in r['cells']) or re.search(r'recomput|recalculat',r.get('context',''),re.I)
 if labelled:
  return original_model()['working'][key],r.get('delta_recomputed',r.get('delta')),r.get('recomputed_original'),'recomputed_original'
 return Decimal(str(next(x['amount'] for x in sr.S['summary'] if x['id']==key))),r.get('delta'),r.get('printed'),'printed_original'
def reconciliations(f,o):
 evidence=[]
 for key,value in o['working'].items():
  for r in rows(f,'summary',key):
   if not any(r.get(k) is not None for k in ['delta','delta_recomputed']):continue
   basis,delta,actual_basis,kind=comparison_basis(r,key)
   own_value=r.get('working')
   if own_value is None:
    current=vals(f,'summary',key,'working');own_value=current[0] if len(current)==1 else None
   sign=r.get('delta_sign',1)
   try:own_delta=sr.number(own_value)-sr.number(actual_basis)
   except (TypeError,ValueError):own_delta=None
   evidence.append({'stage':key,'location':r['sheet']+'!'+str(r['row']),'basis':kind,'actual':delta,'expected':str((value-basis)*sign),'correct':close(delta,(value-basis)*sign),'self_consistent':own_delta is not None and close_delta(delta,own_delta*sign)})
 return evidence
def checks(w,reference_failure=False,arithmetic_failure=False):
 old,f=sr.checks(w,reference_failure,arithmetic_failure);o=compute();facts={f'R00{i}':[] for i in range(1,7)}
 def add(cid,key,ok,actual=None,expected=None):facts[cid].append({'id':key,'ok':bool(ok),'actual':actual,'expected':serial(expected)})
 # Existing source extraction and source-scope semantics remain exactly those
 # of the validated C1 reader. Its old working-value scores are not reused.
 facts['R001']=old['R001']+old['R002']+old['R003']
 facts['R006']=old['R006']
 t=text_rows(f,'heating_price');early=text_rows(f,'earlier_heating');option=text_rows(f,'asbestos_option')
 add('R002','effective_heating_price',review_number(f,'heating_price',458000),review_values(f,'heating_price'),458000)
 earlier_rows=rows(f,'review','earlier_heating');earlier_numbers=[v for v in review_values(f,'earlier_heating') if v is not None]
 earlier_ok=not earlier_rows or (all(sr.equal(v,472000) for v in earlier_numbers) and bool(re.search(r'supersed|replac|obsolete|not current|no longer',early)) and not bool(re.search(r'not superseded|still current|still applicable',early)))
 add('R002','any_displayed_earlier_quote_is_superseded',earlier_ok,{'displayed':bool(earlier_rows),'values':earlier_numbers,'text':early})
 replacement_pairs=[r for r in rows(f,'elements','services') if r.get('working') is not None and r.get('printed') is not None]
 hp=review_values(f,'heating_price');old_heating=vals(f,'provisional','heating','printed')
 replacement_arithmetic=bool(replacement_pairs and hp and old_heating) and all(close_delta(sr.number(r['working'])-sr.number(r['printed']),sr.number(a)-sr.number(b)) for r in replacement_pairs for a in hp for b in old_heating)
 add('R002','same_scope_replacement',(bool(t) and bool(re.search(r'replac|substitut|in lieu|same.scope',t)) or replacement_arithmetic) and not bool(re.search(r'additional charge|add.*on top|in addition to',t)),{'text':t,'candidate_replacement_arithmetic':bool(replacement_arithmetic)})
 add('R002','unapproved_option_price',review_number(f,'asbestos_option',18000),review_values(f,'asbestos_option'),18000)
 add('R002','option_awaits_approval_and_is_outside',bool(option) and bool(re.search(r'not approved|unapproved|await|pending',option)) and bool(re.search(r'exclud|outside|not included|separate|not approved for inclusion',option)) and not bool(re.search(r'is approved|now approved|included in (?:the )?current',option)),option)
 for key,value in o['elements'].items():add('R003','element:'+key,number_fact(f,'elements',key,'working',value),vals(f,'elements',key,'working'),value)
 for key,value in o['working'].items():add('R003','working:'+key,number_fact(f,'summary',key,'working',value),vals(f,'summary',key,'working'),value)
 for key,value in o['rates'].items():add('R003','rate:'+key,number_fact(f,'summary',key,'working_rate',value,False),vals(f,'summary',key,'working_rate'),value)
 rc=reconciliations(f,o);final=[r for r in rc if r['stage']=='vat_excluded']
 add('R004','cost_limit_movement_correct',bool(final) and all(r['correct'] for r in final),final)
 add('R004','movement_matches_candidate_own_figures',bool(final) and all(r['self_consistent'] for r in final),final)
 bridge_checks=[r for key,rs in f.items() if key[0]=='_bridge_arithmetic' for r in rs]
 add('R004','no_contradictory_displayed_reconciliation',bool(final) and all(r['correct'] and r['self_consistent'] for r in rc) and all(r['ok'] for r in bridge_checks),{'reconciliations':rc,'bridge_arithmetic':bridge_checks})
 for key in ['heating_price','earlier_heating','asbestos_option']:
  rs=rows(f,'review',key)
  add('R006','correspondence:'+key,(key=='earlier_heating' and not rs) or bool(rs) and all(re.search(r'review.correspondence|project.authored|review scenario',str(r.get('source',''))+' '+r.get('context',''),re.I) for r in rs))
 # A clearly populated new review representation that the adapter cannot bind
 # stays pending; missing ordinary required output does not get invented.
 for key in REVIEW_ALIASES:
  if key=='earlier_heating':continue
  if not rows(f,'review',key):
   text=' '.join(str(c.value) for sh in w for row in sh for c in row if c.value is not None)
   if re.search(r'(?:heating|asbestos).{0,90}(?:£|\bGBP\b|\d{4,})|(?:£|\bGBP\b).{0,90}(?:heating|asbestos)',text,re.I):
    raise RecalcUnavailable('Material review-price representation cannot be bound: '+key)
 return facts,f,rc
def source_snapshot(f):
 return {(kind,row['id'],field):[r.get(field) for r in rows(f,kind,row['id'])] for kind in sr.KINDS for row in sr.S[kind] for field in ['printed','rate','scope']}
def control(f,raw,key,value):
 if key.endswith('_rate'):points=[(r['sheet'],r['cells'].get('working_rate')) for r in rows(f,'summary',key[:-5]) if r['cells'].get('working_rate')]
 else:
  rr=rows(f,'review',key);points=[(r['sheet'],r['cells'].get('working') or r['cells'].get('printed')) for r in rr]
 points=list(dict.fromkeys(p for p in points if p[1]));inputs=[p for p in points if raw[p[0]][p[1]].data_type!='f']
 if len(inputs)>1:
  explicit=[p for p in inputs if re.search(r'working calculation|input|control',p[0],re.I)]
  inputs=explicit if len(explicit)==1 else inputs
 if len(inputs)>1:raise RecalcUnavailable('Multiple editable review controls require unambiguous roles: '+key)
 where=inputs[0] if len(inputs)==1 else points[0] if len(points)==1 else None
 if where and raw[where[0]][where[1]].data_type=='f':
  raise RecalcUnavailable('Editable review control uses a formula; dependency control not yet bound: '+key)
 return where,float(value)
def evaluate(path,evidence_dir,completed_run=True,reuse_native=None,dynamic=True):
 out=Path(evidence_dir);out.mkdir(parents=True,exist_ok=True);status=output_status(path)
 if status:result=score_profiles(ROOT/'rubric.json',status=status if completed_run else 'INFRA_ERROR',evidence={'completed_run':completed_run})
 else:
  try:
   raw=openpyxl.load_workbook(path,data_only=False)
   if any(c.data_type=='f' and re.search(r'LAMBDA\(|_xlfn\.PY\(',str(c.value),re.I) for sh in raw for row in sh for c in row):raise RecalcUnavailable('Legal formula feature unsupported by the native adapter')
   broken,reference_failure=sr.broken_sheet_references(raw)
   if reuse_native:
    receipt=json.loads(Path(reuse_native).read_text())['native'];fresh=Path(receipt['output'])
   else:fresh,receipt=recalculate_xlsx(path,out/'base')
   w=openpyxl.load_workbook(fresh,data_only=True);arithmetic,arithmetic_failure=sr.invalid_text_arithmetic(raw,w)
   facts,before,rc=checks(w,reference_failure,arithmetic_failure);orig=compute();probes=[]
   for name,changes in SCENARIOS:
    patch=defaultdict(dict);bound=True
    for key,value in changes.items():
     where,v=control(before,raw,key,value)
     if where:patch[where[0]][where[1]]=v
     else:bound=False
    aft={};rec=None;new=compute(changes);after_rc=[]
    if bound and dynamic:
     dest=out/name/'mutated.xlsx';edit(path,dest,patches=patch,clear_caches=True)
     fresh2,rec=recalculate_xlsx(dest,out/name/'recalc');after_w=openpyxl.load_workbook(fresh2,data_only=True)
     aft,_=sr.discover(after_w,reference_failure,arithmetic_failure);after_rc=reconciliations(aft,new)
    for key,value in new['working'].items():
     delta=value-orig['working'][key]
     if delta==0:continue
     bv=vals(before,'summary',key,'working');av=vals(aft,'summary',key,'working')
     ok=bool(bv) and len(bv)==len(av) and all(close(a,value) and close_delta(float(a)-float(b),delta) for a,b in zip(av,bv) if isinstance(a,(int,float)) and isinstance(b,(int,float))) and all(isinstance(a,(int,float)) and isinstance(b,(int,float)) for a,b in zip(av,bv))
     facts['R005'].append({'id':name+':updated:'+key,'ok':ok,'before':bv,'after':av,'expected_after':str(value),'expected_delta':str(delta)})
    final=[r for r in after_rc if r['stage']=='vat_excluded']
    if new['working']['vat_excluded']!=orig['working']['vat_excluded']:
     facts['R005'].append({'id':name+':reconciliation_updates','ok':bool(final) and all(r['correct'] and r['self_consistent'] for r in final),'actual':final})
    unchanged=bound and bool(aft) and source_snapshot(before)==source_snapshot(aft)
    # Baseline source correctness is R001; this unit measures preservation.
    for key in orig['elements']:
     if name!='unapproved_option_price' and new['elements'][key]==orig['elements'][key]:unchanged=unchanged and vals(before,'elements',key,'working')==vals(aft,'elements',key,'working')
    for key in ['heating_price','earlier_heating','asbestos_option']:
     if key not in changes:unchanged=unchanged and review_values(before,key)==review_values(aft,key)
    if name=='unapproved_option_price':
     before_final=[r for r in rc if r['stage']=='vat_excluded']
     option_excluded=bool(aft) and all(vals(before,'summary',key,'working')==vals(aft,'summary',key,'working') for key in orig['working']) and bool(final) and [r['actual'] for r in final]==[r['actual'] for r in before_final]
     facts['R002'].append({'id':'option_price_change_does_not_approve_scope','ok':bool(option_excluded)})
    facts['R001'].append({'id':name+':sources_and_unaffected_preserved','ok':bool(unchanged)})
    readback=bound and bool(aft) and all(review_number(aft,k,v) if not k.endswith('_rate') else number_fact(aft,'summary',k[:-5],'working_rate',v,False) for k,v in changes.items())
    if not readback:
     for unit in facts['R005']:
      if unit['id'].startswith(name+':'):unit['ok']=False
    probes.append({'name':name,'declared_changes':changes,'control_edits':dict(patch),'bound':bound,'edited_input_readback':readback,'native':rec})
   denominators={k:len(v) for k,v in facts.items()}
   contract=json.loads((ROOT/'tests/fact_contract.json').read_text())
   if denominators!=contract['denominators']:raise RuntimeError('Fixed factual denominator drift: '+str(denominators))
   scores={k:str(Decimal(sum(x['ok'] for x in v))/Decimal(len(v))) for k,v in facts.items()}
   extra_claims={str(key):rs for key,rs in before.items() if key[0].startswith('_') or key[0]=='review_source'}
   result=score_profiles(ROOT/'rubric.json',scores,evidence={'judge_version':VERSION,'fact_units':facts,'denominators':denominators,'base_native_receipt':receipt,'dynamic_probes':probes,'candidate_broken_sheet_references':broken,'candidate_invalid_text_arithmetic':arithmetic,'candidate_reconciliation':rc,'additional_candidate_claims':extra_claims,'basis_policy':'Printed source or explicitly labelled recomputed original; whole-pound displays accepted; original source must remain intact.'})
  except (RecalcUnavailable,ValueError,TypeError,KeyError,openpyxl.utils.exceptions.InvalidFileException) as exc:result=score_profiles(ROOT/'rubric.json',status='JUDGE_ERROR',evidence={'error_type':type(exc).__name__,'error':str(exc)})
 (out/'evaluation.json').write_text(json.dumps(result,ensure_ascii=False,indent=2,default=str));return result
if __name__=='__main__':
 p=argparse.ArgumentParser();p.add_argument('answer',nargs='?',default='/app/output/answer.xlsx');p.add_argument('--evidence-dir',default='/tmp/new6-c1-v2');p.add_argument('--input-dir');p.add_argument('--completed-run',action='store_true');p.add_argument('--reuse-native');a=p.parse_args();print(json.dumps(evaluate(a.answer,a.evidence_dir,a.completed_run,a.reuse_native),default=str))
