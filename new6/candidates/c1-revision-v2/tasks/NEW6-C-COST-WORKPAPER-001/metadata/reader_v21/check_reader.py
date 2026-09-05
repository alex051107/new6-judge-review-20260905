"""One focused reader calibration; no new generation and no full legacy suite."""
from pathlib import Path
from decimal import Decimal
import json,sys
HERE=Path(__file__).resolve().parent
ROOT=HERE.parents[1]
sys.path[:0]=[str(ROOT/'tests'),str(ROOT/'metadata')]
from ooxml_edit import edit
from evaluate import evaluate,control,RecalcUnavailable
from source_reader import discover
import openpyxl

def engine_pending():
    r=json.loads((HERE/'checks/legal_natural_layout/evaluation.json').read_text())
    w=openpyxl.load_workbook(r['evidence']['base_native_receipt']['output'],data_only=True)
    w['Summary']['E16']='#NAME?'
    try:discover(w)
    except RecalcUnavailable as exc:row={'name':'unresolved_native_formula_pending','status':'PENDING_READER_LIMIT','error':str(exc),'passed':'engine/parse error' in str(exc)}
    else:row={'name':'unresolved_native_formula_pending','passed':False}
    assert row['passed'],row
    print(json.dumps(row),flush=True)
    return row

def main():
    fixtures=HERE/'fixtures';fixtures.mkdir(exist_ok=True)
    legal=fixtures/'natural_columns_and_vertical_bridge.xlsx'
    bridge={
      'A1':'Cost limit reconciliation',
      'A3':'Original cost limit (October 2024 Revision A)','B3':"='Summary'!B16",
      'A5':'Total Movement','B5':"='Summary'!F16",
      'A7':'Current Review Cost Limit','B7':"='Summary'!E16",
      'A10':'Reconciled Cost Limit','B10':'=B3+B5',
      'A11':'Direct Calculation','B11':"='Summary'!E16",
      'A12':'Difference (should be zero)','B12':'=B11-B10',
    }
    edit(ROOT/'solution/reference.xlsx',legal,patches={
      'Summary':{'B4':'Original Estimate','E4':'Current Review','F4':'Movement'},
      'Review decisions':{'A4':'Package','B4':'Original Price','C4':'Review Price','D4':'Notes','D5':'Updated package price. Ventilation and other works are unchanged.'}
    },new_sheets={'Movement bridge':bridge},clear_caches=True)
    wrong=fixtures/'wrong_movement.xlsx'
    edit(legal,wrong,patches={'Movement bridge':{'B5':"='Summary'!F16+1000"}},clear_caches=True)
    unknown=fixtures/'unknown_numeric_role.xlsx'
    edit(legal,unknown,patches={'Summary':{'A22':'Unspecified commercial aggregate','E22':12345}},clear_caches=True)
    duplicate=fixtures/'duplicate_current_role.xlsx'
    edit(legal,duplicate,patches={'Movement bridge':{'A15':'Current Review Cost Limit','B15':999999}},clear_caches=True)
    expected=json.loads((ROOT/'tests/fact_contract.json').read_text())['denominators']
    pending_only='--pending-only' in sys.argv
    rows=json.loads((HERE/'receipt.json').read_text())['results'][:2] if pending_only else []
    for name,path,lose in [('legal_natural_layout',legal,[]),('wrong_movement',wrong,['R003','R004','R005'])]:
      if pending_only:continue
      result=evaluate(path,HERE/'checks'/name,True)
      scores=result.get('criterion_scores',{})
      checks=[result['evaluation_status']=='SCORED',result.get('evidence',{}).get('denominators')==expected]
      checks += [Decimal(scores.get(k,'-1'))<1 for k in lose]
      checks += [Decimal(scores.get(k,'-1'))==1 for k in expected if k not in lose]
      row={'name':name,'status':result['evaluation_status'],'score':result.get('score_decimal'),'criteria':scores,'denominators':result.get('evidence',{}).get('denominators'),'passed':all(checks)}
      rows.append(row);(HERE/'receipt.json').write_text(json.dumps({'passed':all(x['passed'] for x in rows),'results':rows,'api_calls':0},indent=2))
      print(json.dumps(row),flush=True)
      if not row['passed']:raise AssertionError(name)
    native_legal=json.loads((HERE/'checks/legal_natural_layout/evaluation.json').read_text())['evidence']['base_native_receipt']['output']
    edit(native_legal,unknown,patches={'Summary':{'A22':'Unspecified commercial aggregate','E22':12345}})
    edit(native_legal,duplicate,patches={'Movement bridge':{'A15':'Current Review Cost Limit','B15':999999}})
    # The two pending safeguards concern binding, so exercise the exact reader
    # directly. Their numeric cells need no native recalculation to be unknown.
    for name,path,phrase in [('unknown_role_pending',unknown,'remains unbound'),('duplicate_role_pending',duplicate,'Repeated labelled vertical bridge role')]:
      try:discover(openpyxl.load_workbook(path,data_only=True))
      except ValueError as exc:
        row={'name':name,'status':'PENDING_READER_LIMIT','error':str(exc),'passed':phrase in str(exc)}
      else:row={'name':name,'passed':False,'error':'Unknown or duplicate role silently accepted'}
      rows.append(row);print(json.dumps(row),flush=True)
      if not row['passed']:raise AssertionError(name)
    w=openpyxl.Workbook();w.remove(w.active)
    records=[]
    for title in ['Package register A','Package register B']:
      sheet=w.create_sheet(title);sheet['B2']=458000
      records.append({'sheet':title,'cells':{'working':'B2'}})
    try:control({('review','heating_price'):records},w,'heating_price',468000)
    except RecalcUnavailable as exc:row={'name':'ambiguous_editable_control_pending','status':'PENDING_READER_LIMIT','error':str(exc),'passed':'Multiple editable review controls' in str(exc)}
    else:row={'name':'ambiguous_editable_control_pending','passed':False}
    rows.append(row);print(json.dumps(row),flush=True)
    if not row['passed']:raise AssertionError(row['name'])
    rows.append(engine_pending())
    (HERE/'receipt.json').write_text(json.dumps({'passed':all(x['passed'] for x in rows),'results':rows,'api_calls':0,'fixed_denominators':expected,'scope':'One legal layout, one wrong movement, unknown and duplicate-role fail-closed safeguards; legacy 19 fixtures not rerun.'},indent=2))

if __name__=='__main__':
    if '--engine-only' in sys.argv:
      receipt=json.loads((HERE/'receipt.json').read_text());receipt['results'].append(engine_pending());receipt['passed']=all(r['passed'] for r in receipt['results']);(HERE/'receipt.json').write_text(json.dumps(receipt,indent=2))
    else:main()
