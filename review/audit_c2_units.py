"""Independent inspection of two already-scored C2 originals; no new reward."""
from pathlib import Path
from decimal import Decimal as D
import json,sys,csv
import openpyxl
ROOT=Path(__file__).resolve().parents[1]
T=ROOT/'tasks/49/NEW6-C-PARCEL-TARIFF-001/tests/new6/tasks/NEW6-C-PARCEL-TARIFF-001'
sys.path[:0]=[str(T/'metadata'),str(T.parents[2]/'common')]
from oracle_recompute import compute
rates=json.loads((T/'metadata/rates.json').read_text());requests=list(csv.DictReader((T/'data/input_files/quote_requests.csv').open()));expected=compute(rates,requests)
rows=json.loads((ROOT/'results/trials.json').read_text());out=[]
for id in ['C2-codex-R08','C2-codex-R02']:
 r=next(r for r in rows if r['id']==id);w=openpyxl.load_workbook(ROOT/r['answer'],data_only=True);e=json.loads((ROOT/r['receipt']).read_text())['evidence']
 # R08 coordinates read from the original's labels; R02 is inventoried separately.
 if id!='C2-codex-R08':
  out.append({'id':id,'classification':'JUDGE_BINDING_REVIEW_REQUIRED','recorded_score':r['score_decimal'],'zero_dynamic_control_tests':sum(not t.get('actual_control_cells') for t in e['dynamic']),'reason':'No bound editable controls in Judge receipt; do not infer absent functionality.'});continue
 cells=[]
 for rate in rates:
  row=5+int(rate['upper_bound']) if rate['service']=='priority' else 22+['4','8','12','15.999','16','32','48','64','80','96','112','128','144','160'].index(str(D(rate['upper_oz']).normalize()) if D(rate['upper_oz'])<16 else str(int(D(rate['upper_oz']))))
  col=1+int(rate['zone']);actual=w['Rate Sheet'].cell(row,col).value
  cells.append({'cell':'Rate Sheet!'+w['Rate Sheet'].cell(row,col).coordinate,'service':rate['service'],'upper_oz':rate['upper_oz'],'actual':actual,'expected':rate['usd'],'matches':D(str(actual))==D(rate['usd'])})
 quote=[]
 for i,q in enumerate(expected['quotes'],5):
  for col,key in [(8,'priority'),(9,'ground'),(11,'selected_usd')]:
   actual=w['Quotes'].cell(i,col).value;value=q.get(key)
   if value is None:print('QUOTE_KEYS',q);raise KeyError(key)
   quote.append({'request':q['request_id'],'cell':'Quotes!'+w['Quotes'].cell(i,col).coordinate,'actual':actual,'expected':str(value),'matches':D(str(actual))==D(str(value))})
 out.append({'id':id,'classification':'CONFIRMED_JUDGE_FALSE_NEGATIVE','recorded_score':r['score_decimal'],'price_grid_matches':sum(c['matches'] for c in cells),'price_grid_total':len(cells),'quote_value_matches':sum(c['matches'] for c in quote),'quote_value_total':len(quote),'zero_dynamic_control_tests':sum(not t.get('actual_control_cells') for t in e['dynamic']),'original_editable_control':'Quotes!B5','observed_headers':['Quotes!H4 = Priority Mail Retail','Quotes!I4 = Ground Advantage Retail','Rate Sheet!A20 = Max actual weight (oz)','Rate Sheet!A26 = 16; J26 = 1 lb'],'cause':'Exact header aliases omit Retail; rate key equates nominal unit labels instead of accepting explicitly equivalent ounce bounds. Quote table silently omitted although rate table bound.','rate_checks':cells,'quote_checks':quote,'no_new_official_score':True})
p=ROOT/'results/c2-independent-audit.json';p.write_text(json.dumps(out,ensure_ascii=False,indent=2));print(json.dumps([{k:v for k,v in r.items() if k not in ['rate_checks','quote_checks']} for r in out],ensure_ascii=False,indent=2))
